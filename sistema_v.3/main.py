#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Sistema de Vendas - Auto Pecas & Mecanica v2.0"""

import sys, os, json, uuid
from datetime import datetime
from collections import defaultdict

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGridLayout, QLabel, QLineEdit, QPushButton, QTableWidget,
    QTableWidgetItem, QDialog, QFormLayout, QComboBox, QSpinBox,
    QDoubleSpinBox, QTextEdit, QMessageBox, QFileDialog, QHeaderView,
    QFrame, QDateEdit, QGroupBox, QRadioButton, QSplitter,
    QAbstractItemView, QStatusBar, QStackedWidget, QCheckBox, QColorDialog
)
from PyQt5.QtCore import Qt, QDate, QTimer, pyqtSignal
from PyQt5.QtGui import QFont, QColor, QPixmap, QBrush, QPalette

try:
    import openpyxl
    from openpyxl.styles import Font as XLFont, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    PDF_OK = True
except ImportError:
    PDF_OK = False

# ─── PATHS ────────────────────────────────────────────────────────────────────
if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(BASE_DIR, "autopecas_data.json")

# ─── PORTUGUESE WEEKDAYS ──────────────────────────────────────────────────────
DIAS_PT = {
    "Monday": "Segunda-feira", "Tuesday": "Terca-feira",
    "Wednesday": "Quarta-feira", "Thursday": "Quinta-feira",
    "Friday": "Sexta-feira", "Saturday": "Sabado", "Sunday": "Domingo",
}

PRESET_COLORS = [
    ("#FF6B35", "Laranja (Padrao)"),
    ("#E53935", "Vermelho"),
    ("#1E88E5", "Azul"),
    ("#43A047", "Verde"),
    ("#8E24AA", "Roxo"),
    ("#FFB300", "Ambar"),
    ("#00ACC1", "Ciano"),
    ("#F06292", "Rosa"),
    ("#5D4037", "Marrom"),
    ("#546E7A", "Ardosia"),
]

CATEGORIES = ["Filtros","Freios","Oleos","Motor","Eletrica","Suspensao",
               "Arrefecimento","Transmissao","Carroceria","Acessorios","Outros"]
UNITS = ["UN","JG","LT","KG","MT","CX","PC","FR","KIT","PAR"]
PAYMENTS = ["Dinheiro","Cartao de Credito","Cartao de Debito",
            "PIX","Boleto","Transferencia","Fiado / A Prazo"]



# ═════════════════════════════════════════════════════════════════════════════
#  NOTA FISCAL PDF
# ═════════════════════════════════════════════════════════════════════════════
def _gerar_nota_fiscal(path, cart, customer_name, customer_obj, subtotal,
                       disc_pct, disc_val, total, payment, observations, dm):
    """Generate a professional invoice PDF using reportlab."""
    doc = SimpleDocTemplate(
        path, pagesize=A4,
        rightMargin=15*mm, leftMargin=15*mm,
        topMargin=15*mm, bottomMargin=15*mm
    )
    styles = getSampleStyleSheet()
    accent = colors.HexColor("#FF6B35")
    dark   = colors.HexColor("#1A1A1A")
    gray   = colors.HexColor("#666666")
    light  = colors.HexColor("#F5F5F5")
    white  = colors.white

    title_style  = ParagraphStyle("Title2",  parent=styles["Normal"], fontSize=22, textColor=accent,  fontName="Helvetica-Bold", alignment=TA_CENTER, spaceAfter=2)
    sub_style    = ParagraphStyle("Sub2",    parent=styles["Normal"], fontSize=10, textColor=gray,    fontName="Helvetica",      alignment=TA_CENTER, spaceAfter=2)
    h1_style     = ParagraphStyle("H1",      parent=styles["Normal"], fontSize=11, textColor=dark,    fontName="Helvetica-Bold", spaceAfter=2)
    normal_style = ParagraphStyle("Normal2", parent=styles["Normal"], fontSize=9,  textColor=dark,    fontName="Helvetica")
    small_style  = ParagraphStyle("Small",   parent=styles["Normal"], fontSize=8,  textColor=gray,    fontName="Helvetica")
    right_style  = ParagraphStyle("Right",   parent=styles["Normal"], fontSize=10, textColor=dark,    fontName="Helvetica-Bold", alignment=TA_RIGHT)
    total_style  = ParagraphStyle("Total",   parent=styles["Normal"], fontSize=14, textColor=accent,  fontName="Helvetica-Bold", alignment=TA_RIGHT)

    story = []
    W = A4[0] - 30*mm  # usable width

    # ── HEADER ────────────────────────────────────────────────────────────────
    header_data = [[
        Paragraph("<b><font size=24 color='#FF6B35'>AUTO PECAS</font></b><br/>"
                  "<font size=10 color='#666666'>Sistema de Gestao e Vendas</font>", styles["Normal"]),
        Paragraph(
            f"<b><font size=18 color='#1A1A1A'>NOTA FISCAL</font></b><br/>"
            f"<font size=9 color='#666666'>Emitida em: {datetime.now().strftime('%d/%m/%Y  %H:%M')}</font>",
            ParagraphStyle("RH", parent=styles["Normal"], alignment=TA_RIGHT)
        )
    ]]
    header_table = Table(header_data, colWidths=[W*0.55, W*0.45])
    header_table.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#FFF8F5")),
        ("ROUNDEDCORNERS", [8,8,8,8]),
        ("BOTTOMPADDING", (0,0), (-1,-1), 10),
        ("TOPPADDING", (0,0), (-1,-1), 10),
        ("LEFTPADDING", (0,0), (0,-1), 12),
        ("RIGHTPADDING", (-1,0), (-1,-1), 12),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 6*mm))
    story.append(HRFlowable(width="100%", thickness=2, color=accent))
    story.append(Spacer(1, 4*mm))

    # ── CLIENT + INFO ─────────────────────────────────────────────────────────
    cpf_cnpj = customer_obj.get("cpf_cnpj","--") if customer_obj else "--"
    phone    = customer_obj.get("phone","--")    if customer_obj else "--"
    address  = customer_obj.get("address","--")  if customer_obj else "--"

    info_data = [[
        [Paragraph("<b>DADOS DO CLIENTE</b>", h1_style),
         Paragraph(f"Nome: {customer_name}", normal_style),
         Paragraph(f"CPF/CNPJ: {cpf_cnpj}", normal_style),
         Paragraph(f"Telefone: {phone}", normal_style),
         Paragraph(f"Endereco: {address}", normal_style)],
        [Paragraph("<b>INFORMACOES DE PAGAMENTO</b>", h1_style),
         Paragraph(f"Forma de Pagamento: <b>{payment}</b>", normal_style),
         Paragraph(f"Data de Emissao: <b>{datetime.now().strftime('%d/%m/%Y')}</b>", normal_style),
         Paragraph(f"Hora: <b>{datetime.now().strftime('%H:%M:%S')}</b>", normal_style),
         Paragraph(f"Status: <b>EMITIDA</b>", normal_style)],
    ]]
    info_table = Table(info_data, colWidths=[W*0.55, W*0.45])
    info_table.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("BACKGROUND", (0,0), (-1,-1), light),
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#DDDDDD")),
        ("TOPPADDING", (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("LEFTPADDING", (0,0), (-1,-1), 10),
        ("RIGHTPADDING", (0,0), (-1,-1), 10),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 6*mm))

    # ── ITEMS TABLE ──────────────────────────────────────────────────────────
    story.append(Paragraph("<b>ITENS DA NOTA FISCAL</b>", h1_style))
    story.append(Spacer(1, 2*mm))

    item_header = ["Cod.", "Produto / Descricao", "Qtd", "Preco Unit.", "Desc.%", "Total"]
    col_ws = [W*0.08, W*0.38, W*0.08, W*0.16, W*0.10, W*0.18]
    item_rows = [item_header]

    for it in cart:
        disc_f = 1 - it["discount"]/100
        row_total = it["quantity"] * it["unit_price"] * disc_f
        item_rows.append([
            it.get("code",""),
            it.get("name",""),
            str(it["quantity"]),
            fmtR(it["unit_price"]),
            f"{it['discount']:.1f}%",
            fmtR(row_total),
        ])

    item_table = Table(item_rows, colWidths=col_ws, repeatRows=1)
    item_table.setStyle(TableStyle([
        # Header row
        ("BACKGROUND",    (0,0), (-1,0), accent),
        ("TEXTCOLOR",     (0,0), (-1,0), white),
        ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,0), 9),
        ("ALIGN",         (0,0), (-1,0), "CENTER"),
        ("TOPPADDING",    (0,0), (-1,0), 7),
        ("BOTTOMPADDING", (0,0), (-1,0), 7),
        # Data rows
        ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",      (0,1), (-1,-1), 8),
        ("ALIGN",         (0,1), (0,-1), "CENTER"),   # code
        ("ALIGN",         (2,1), (2,-1), "CENTER"),   # qty
        ("ALIGN",         (3,1), (-1,-1), "RIGHT"),   # prices
        ("TOPPADDING",    (0,1), (-1,-1), 5),
        ("BOTTOMPADDING", (0,1), (-1,-1), 5),
        ("LEFTPADDING",   (1,0), (1,-1), 6),
        # Alternating rows
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [white, light]),
        # Grid
        ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#CCCCCC")),
        ("LINEBELOW",     (0,0), (-1,0), 1.5, accent),
    ]))
    story.append(item_table)
    story.append(Spacer(1, 5*mm))

    # ── TOTALS ────────────────────────────────────────────────────────────────
    totals_data = [
        ["", "Subtotal:", fmtR(subtotal)],
        ["", f"Desconto ({disc_pct:.1f}%):", f"- {fmtR(disc_val)}"],
        ["", "TOTAL:", fmtR(total)],
    ]
    totals_table = Table(totals_data, colWidths=[W*0.55, W*0.25, W*0.20])
    totals_table.setStyle(TableStyle([
        ("FONTNAME",      (0,0), (-1,-2), "Helvetica"),
        ("FONTSIZE",      (0,0), (-1,-2), 9),
        ("FONTNAME",      (0,-1), (-1,-1), "Helvetica-Bold"),
        ("FONTSIZE",      (1,-1), (2,-1), 13),
        ("TEXTCOLOR",     (2,-1), (2,-1), accent),
        ("ALIGN",         (1,0), (-1,-1), "RIGHT"),
        ("TOPPADDING",    (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ("LINEABOVE",     (1,-1), (-1,-1), 1.5, accent),
    ]))
    story.append(totals_table)
    story.append(Spacer(1, 4*mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#CCCCCC")))
    story.append(Spacer(1, 3*mm))

    # ── OBSERVATIONS ─────────────────────────────────────────────────────────
    if observations.strip():
        story.append(Paragraph("<b>Observacoes:</b>", h1_style))
        story.append(Paragraph(observations.strip(), normal_style))
        story.append(Spacer(1, 3*mm))

    # ── FOOTER ───────────────────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=1, color=accent))
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(
        f"<font color='#FF6B35'>AUTO PECAS & MECANICA</font>  |  "
        f"Nota emitida em {datetime.now().strftime('%d/%m/%Y as %H:%M')}  |  "
        "Sistema de Gestao v2.0",
        ParagraphStyle("Footer", parent=styles["Normal"], fontSize=8, textColor=gray, alignment=TA_CENTER)
    ))

    doc.build(story)


# ═════════════════════════════════════════════════════════════════════════════
#  DATA MANAGER
# ═════════════════════════════════════════════════════════════════════════════
class DataManager:
    def __init__(self):
        self.data = {}
        self.load()

    def load(self):
        if os.path.exists(DATA_FILE):
            try:
                with open(DATA_FILE, "r", encoding="utf-8") as f:
                    self.data = json.load(f)
                for k, v in self._defaults().items():
                    if k not in self.data:
                        self.data[k] = v
                # ensure built-in users exist
                self._ensure_builtin_users()
            except Exception:
                self.data = self._defaults()
                self._add_samples()
        else:
            self.data = self._defaults()
            self._add_samples()
            self.save()

    def _defaults(self):
        return {
            "products": [], "customers": [], "sales": [],
            "users": [],
            "settings": {"theme": "dark", "background_image": "", "accent_color": "#FF6B35"}
        }

    def _ensure_builtin_users(self):
        """Always guarantee admin and funcionario accounts exist."""
        existing = {u["username"] for u in self.data.get("users", [])}
        if "admin" not in existing:
            self.data["users"].insert(0, {
                "id": "admin-fixed", "username": "admin", "password": "admin",
                "role": "admin", "name": "Administrador", "active": True
            })
        if "funcionario" not in existing:
            self.data["users"].append({
                "id": "func-fixed", "username": "funcionario", "password": "123moto",
                "role": "operator", "name": "Funcionario", "active": True
            })

    def _add_samples(self):
        self._ensure_builtin_users()
        self.data["products"] = self._sample_products()
        self.data["customers"] = self._sample_customers()

    def save(self):
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)

    def _sample_products(self):
        rows = [
            ("FLT001","Filtro de Oleo","Filtros","Bosch","UN",15,35,50,10),
            ("FLT002","Filtro de Ar","Filtros","Mann","UN",20,45,30,8),
            ("FLT003","Filtro de Combustivel","Filtros","WIX","UN",18,40,25,5),
            ("FRE001","Pastilha de Freio Dianteira","Freios","Bosch","JG",45,95,20,5),
            ("FRE002","Disco de Freio Dianteiro","Freios","Brembo","UN",80,165,15,4),
            ("FRE003","Fluido de Freio DOT4","Freios","ATE","FR",12,28,40,10),
            ("OLE001","Oleo 5W30 Sintetico","Oleos","Castrol","LT",28,55,100,20),
            ("OLE002","Oleo 10W40 Semissintetico","Oleos","Mobil","LT",22,45,80,15),
            ("OLE003","Oleo de Cambio","Oleos","Valvoline","LT",30,60,30,8),
            ("VEL001","Vela de Ignicao","Motor","NGK","UN",12,25,5,10),
            ("BAT001","Bateria 60Ah","Eletrica","Moura","UN",280,450,10,3),
            ("ALT001","Correia Alternador","Motor","Dayco","UN",35,72,18,5),
            ("SUS001","Amortecedor Dianteiro","Suspensao","Monroe","UN",120,220,8,2),
            ("SUS002","Barra Estabilizadora","Suspensao","TRW","UN",65,130,6,2),
            ("RAD001","Radiador","Arrefecimento","Valeo","UN",350,620,4,1),
        ]
        return [{"id": str(uuid.uuid4()), "code": r[0], "name": r[1],
                 "category": r[2], "brand": r[3], "unit": r[4],
                 "cost_price": float(r[5]), "sale_price": float(r[6]),
                 "stock": r[7], "min_stock": r[8], "description": ""} for r in rows]

    def _sample_customers(self):
        return [
            {"id": str(uuid.uuid4()), "name": "Joao Carlos Silva",
             "cpf_cnpj": "123.456.789-00", "phone": "(21) 98765-4321",
             "email": "joao@email.com", "address": "Rua das Flores, 100"},
            {"id": str(uuid.uuid4()), "name": "Maria Santos Oliveira",
             "cpf_cnpj": "987.654.321-00", "phone": "(21) 91234-5678",
             "email": "maria@email.com", "address": "Av. Brasil, 200"},
            {"id": str(uuid.uuid4()), "name": "Auto Center Mega Ltda",
             "cpf_cnpj": "12.345.678/0001-90", "phone": "(21) 3333-4444",
             "email": "contato@mega.com", "address": "Rod. Dutra, 300"},
            {"id": str(uuid.uuid4()), "name": "Pedro Melo Mecanica",
             "cpf_cnpj": "45.678.901/0001-23", "phone": "(21) 97777-8888",
             "email": "pedro@mec.com", "address": "Rua Industrial, 55"},
        ]

    # PRODUCTS
    def get_products(self): return self.data["products"]
    def add_product(self, p): p["id"] = str(uuid.uuid4()); self.data["products"].append(p); self.save()
    def update_product(self, pid, u):
        for i, p in enumerate(self.data["products"]):
            if p["id"] == pid: self.data["products"][i] = u; break
        self.save()
    def delete_product(self, pid):
        self.data["products"] = [p for p in self.data["products"] if p["id"] != pid]; self.save()
    def get_product_by_id(self, pid):
        return next((p for p in self.data["products"] if p["id"] == pid), None)

    # CUSTOMERS
    def get_customers(self): return self.data["customers"]
    def add_customer(self, c): c["id"] = str(uuid.uuid4()); self.data["customers"].append(c); self.save()
    def update_customer(self, cid, u):
        for i, c in enumerate(self.data["customers"]):
            if c["id"] == cid: self.data["customers"][i] = u; break
        self.save()
    def delete_customer(self, cid):
        self.data["customers"] = [c for c in self.data["customers"] if c["id"] != cid]; self.save()
    def get_customer_by_id(self, cid):
        return next((c for c in self.data["customers"] if c["id"] == cid), None)

    # SALES
    def get_sales(self): return self.data["sales"]
    def add_sale(self, sale):
        sale["id"] = str(uuid.uuid4()); sale["date"] = datetime.now().isoformat()
        for item in sale.get("items", []):
            for p in self.data["products"]:
                if p["id"] == item["product_id"]:
                    p["stock"] = max(0, p["stock"] - item["quantity"]); break
        self.data["sales"].append(sale); self.save(); return sale["id"]

    # USERS
    def get_users(self): return self.data.get("users", [])
    def get_user_by_username(self, username):
        return next((u for u in self.get_users() if u.get("username","").lower() == username.lower()), None)
    def authenticate(self, username, password):
        u = self.get_user_by_username(username)
        return u if (u and u.get("password") == password and u.get("active", True)) else None
    def add_user(self, user):
        if self.get_user_by_username(user["username"]): return False, "Login ja existe."
        user["id"] = str(uuid.uuid4()); user["active"] = True
        self.data["users"].append(user); self.save(); return True, "Usuario criado!"
    def update_user(self, uid, upd):
        for i, u in enumerate(self.data["users"]):
            if u["id"] == uid: upd["id"] = uid; self.data["users"][i] = upd; break
        self.save()
    def delete_user(self, uid):
        self.data["users"] = [u for u in self.data["users"] if u["id"] != uid]; self.save()
    def toggle_user_active(self, uid):
        for u in self.data["users"]:
            if u["id"] == uid: u["active"] = not u.get("active", True); break
        self.save()

    # SALES CANCEL/DELETE
    def cancel_sale(self, sale_id):
        """Mark sale as cancelled and restore stock."""
        for s in self.data["sales"]:
            if s["id"] == sale_id:
                if s.get("status") == "cancelada":
                    return False, "Esta venda ja esta cancelada."
                # Restore stock for each item
                for item in s.get("items", []):
                    for p in self.data["products"]:
                        if p["id"] == item["product_id"]:
                            p["stock"] += item["quantity"]; break
                s["status"] = "cancelada"
                s["cancelled_at"] = datetime.now().isoformat()
                self.save()
                return True, "Venda cancelada com sucesso!"
        return False, "Venda nao encontrada."

    def delete_sale(self, sale_id):
        """Permanently delete a sale record."""
        self.data["sales"] = [s for s in self.data["sales"] if s["id"] != sale_id]
        self.save()

    # SETTINGS
    def get_settings(self):
        s = self.data.get("settings", {})
        s.setdefault("accent_color", "#FF6B35")
        s.setdefault("theme", "dark")
        s.setdefault("background_image", "")
        return s
    def save_settings(self, s): self.data["settings"] = s; self.save()


# ═════════════════════════════════════════════════════════════════════════════
#  STYLESHEET
# ═════════════════════════════════════════════════════════════════════════════
def _lighter(h):
    c = QColor(h); hue, s, v, a = c.getHsvF()
    c.setHsvF(hue, max(0, s * 0.8), min(1.0, v * 1.25), a); return c.name()

def _darker(h):
    c = QColor(h); hue, s, v, a = c.getHsvF()
    c.setHsvF(hue, s, v * 0.72, a); return c.name()

def build_stylesheet(theme="dark", accent="#FF6B35"):
    ah = _lighter(accent); ad = _darker(accent)
    if theme == "dark":
        bg="#1A1A1A"; surf="#252525"; surf2="#2E2E2E"; brd="#3A3A3A"
        txt="#EFEFEF"; txt2="#AAAAAA"; talt="#2A2A2A"; inbg="#333333"
        side="#1E1E1E"; card="#262626"; hdr="#1F1F1F"
        ok="#00C853"; warn="#FFD600"; err="#FF1744"
    else:
        bg="#F0F2F5"; surf="#FFFFFF"; surf2="#F8F9FA"; brd="#DADCE0"
        txt="#1A1A1A"; txt2="#666666"; talt="#F5F5F5"; inbg="#FFFFFF"
        side="#FFFFFF"; card="#FFFFFF"; hdr="#FFFFFF"
        ok="#43A047"; warn="#FB8C00"; err="#E53935"

    return f"""
QMainWindow,QDialog{{background:{bg};color:{txt};}}
QWidget{{background:{bg};color:{txt};font-family:'Segoe UI',Arial,sans-serif;font-size:13px;}}
QFrame#card{{background:{card};border:1px solid {brd};border-radius:8px;}}
QFrame#sidebar{{background:{side};border-right:1px solid {brd};}}
QLabel#title{{font-size:22px;font-weight:bold;color:{txt};}}
QLabel#subtitle{{font-size:12px;color:{txt2};}}
QLabel#metric_value{{font-size:26px;font-weight:bold;color:{accent};}}
QLabel#metric_label{{font-size:11px;color:{txt2};}}
QLabel#section_title{{font-size:15px;font-weight:bold;color:{txt};border-bottom:2px solid {accent};padding-bottom:4px;}}
QPushButton{{background:{accent};color:#FFF;border:none;border-radius:6px;padding:8px 18px;font-size:13px;font-weight:bold;}}
QPushButton:hover{{background:{ah};}}
QPushButton:pressed{{background:{ad};}}
QPushButton:disabled{{background:{brd};color:{txt2};}}
QPushButton#btn_secondary{{background:{surf2};color:{txt};border:1px solid {brd};}}
QPushButton#btn_secondary:hover{{background:{brd};}}
QPushButton#btn_danger{{background:{err};color:#FFF;}}
QPushButton#btn_danger:hover{{background:#FF4569;}}
QPushButton#btn_success{{background:{ok};color:#FFF;}}
QPushButton#btn_success:hover{{background:#33D66B;}}
QPushButton#btn_warning{{background:{warn};color:#111;}}
QPushButton#nav_btn{{background:transparent;color:{txt2};border:none;border-radius:8px;padding:10px 16px;text-align:left;font-size:13px;}}
QPushButton#nav_btn:hover{{background:{surf2};color:{txt};}}
QPushButton#nav_btn_active{{background:{accent};color:#FFF;border:none;border-radius:8px;padding:10px 16px;text-align:left;font-size:13px;font-weight:bold;}}
QLineEdit,QTextEdit,QComboBox,QSpinBox,QDoubleSpinBox,QDateEdit{{background:{inbg};color:{txt};border:1px solid {brd};border-radius:6px;padding:6px 10px;font-size:13px;}}
QLineEdit:focus,QTextEdit:focus,QComboBox:focus,QSpinBox:focus,QDoubleSpinBox:focus,QDateEdit:focus{{border:2px solid {accent};}}
QComboBox::drop-down{{border:none;width:28px;}}
QComboBox::down-arrow{{image:none;border-left:5px solid transparent;border-right:5px solid transparent;border-top:6px solid {txt2};margin-right:8px;}}
QComboBox QAbstractItemView{{background:{surf};color:{txt};border:1px solid {brd};selection-background-color:{accent};}}
QTableWidget{{background:{surf};color:{txt};border:1px solid {brd};border-radius:6px;gridline-color:{brd};alternate-background-color:{talt};}}
QTableWidget::item{{padding:6px 10px;}}
QTableWidget::item:selected{{background:{accent};color:#FFF;}}
QHeaderView::section{{background:{surf2};color:{txt};padding:8px 10px;border:none;border-right:1px solid {brd};border-bottom:2px solid {accent};font-weight:bold;font-size:12px;}}
QGroupBox{{border:1px solid {brd};border-radius:8px;margin-top:12px;padding:12px 8px 8px 8px;font-weight:bold;color:{txt};}}
QGroupBox::title{{subcontrol-origin:margin;subcontrol-position:top left;padding:0 8px;color:{accent};font-size:13px;}}
QScrollBar:vertical{{background:{surf2};width:8px;border-radius:4px;}}
QScrollBar::handle:vertical{{background:{brd};border-radius:4px;min-height:24px;}}
QScrollBar::handle:vertical:hover{{background:{accent};}}
QScrollBar:horizontal{{background:{surf2};height:8px;border-radius:4px;}}
QScrollBar::handle:horizontal{{background:{brd};border-radius:4px;min-width:24px;}}
QScrollBar::add-line,QScrollBar::sub-line{{width:0;height:0;}}
QMessageBox{{background:{surf};color:{txt};}}
QStatusBar{{background:{hdr};color:{txt2};border-top:1px solid {brd};font-size:12px;padding:2px 8px;}}
QRadioButton{{color:{txt};spacing:8px;}}
QRadioButton::indicator{{width:16px;height:16px;border-radius:8px;border:2px solid {brd};background:{inbg};}}
QRadioButton::indicator:checked{{background:{accent};border-color:{accent};}}
QCheckBox{{color:{txt};spacing:8px;}}
QCheckBox::indicator{{width:16px;height:16px;border-radius:3px;border:2px solid {brd};background:{inbg};}}
QCheckBox::indicator:checked{{background:{accent};border-color:{accent};}}
QSplitter::handle{{background:{brd};}}
"""


# ═════════════════════════════════════════════════════════════════════════════
#  HELPERS / REUSABLE WIDGETS
# ═════════════════════════════════════════════════════════════════════════════
def fmtR(v):
    return "R$ {:,.2f}".format(v).replace(",","X").replace(".",",").replace("X",".")

def fmt_date(iso_str, show_time=True):
    """Convert ISO datetime string to Brazilian format DD/MM/YYYY HH:MM."""
    if not iso_str:
        return "--"
    try:
        iso_str = iso_str.replace("T", " ").strip()
        if len(iso_str) >= 16:
            dt = datetime.strptime(iso_str[:16], "%Y-%m-%d %H:%M")
            return dt.strftime("%d/%m/%Y %H:%M") if show_time else dt.strftime("%d/%m/%Y")
        elif len(iso_str) >= 10:
            dt = datetime.strptime(iso_str[:10], "%Y-%m-%d")
            return dt.strftime("%d/%m/%Y")
        return iso_str
    except Exception:
        return iso_str

class MetricCard(QFrame):
    def __init__(self, title, value, icon="", color=None, parent=None):
        super().__init__(parent); self.setObjectName("card")
        self.setMinimumHeight(110); self.setMinimumWidth(175)
        lay = QVBoxLayout(self); lay.setContentsMargins(16,14,16,14); lay.setSpacing(6)
        top = QHBoxLayout()
        ico = QLabel(icon); ico.setFont(QFont("Segoe UI Emoji", 22))
        self.val = QLabel(str(value)); self.val.setObjectName("metric_value")
        if color: self.val.setStyleSheet(f"color:{color};font-size:26px;font-weight:bold;")
        top.addWidget(ico); top.addStretch(); top.addWidget(self.val); lay.addLayout(top)
        t = QLabel(title); t.setObjectName("metric_label"); lay.addWidget(t)
    def set_value(self, v): self.val.setText(str(v))

class SearchBar(QWidget):
    search_changed = pyqtSignal(str)
    def __init__(self, ph="Buscar...", parent=None):
        super().__init__(parent)
        lay = QHBoxLayout(self); lay.setContentsMargins(0,0,0,0)
        self.edit = QLineEdit(); self.edit.setPlaceholderText(f"  {ph}")
        self.edit.textChanged.connect(self.search_changed); lay.addWidget(self.edit)

class SectionTitle(QLabel):
    def __init__(self, text, parent=None):
        super().__init__(text, parent); self.setObjectName("section_title")
        f = QFont(); f.setPointSize(13); f.setBold(True); self.setFont(f)

class SwatchBtn(QPushButton):
    def __init__(self, color, label, parent=None):
        super().__init__(label, parent); self.color = color; self._sel = False; self._paint()
    def set_selected(self, v): self._sel = v; self._paint()
    def _paint(self):
        brd = "3px solid #FFF" if self._sel else f"2px solid {self.color}"
        fw = "bold" if self._sel else "normal"
        self.setStyleSheet(f"QPushButton{{background:{self.color};color:#FFF;border:{brd};"
                           f"border-radius:8px;padding:6px 10px;font-size:12px;font-weight:{fw};"
                           f"text-shadow:0 1px 2px rgba(0,0,0,.7);}}"
                           f"QPushButton:hover{{border:3px solid #FFF;font-weight:bold;}}")


# ═════════════════════════════════════════════════════════════════════════════
#  DIALOGS — Product / Customer / User / Settings
# ═════════════════════════════════════════════════════════════════════════════
class ProductDialog(QDialog):
    def __init__(self, parent=None, product=None):
        super().__init__(parent); self.product = product
        self.setWindowTitle("Produto"); self.setMinimumWidth(480); self.setModal(True)
        lay = QVBoxLayout(self); lay.setContentsMargins(20,20,20,20); lay.setSpacing(10)
        lay.addWidget(SectionTitle("Cadastro de Produto")); lay.addSpacing(4)
        g = QGridLayout(); g.setSpacing(9)
        self.code = QLineEdit(); self.name = QLineEdit()
        self.code.setPlaceholderText("Preenchido automaticamente pelo nome")
        self.name.textChanged.connect(self._auto_code)
        self.cat = QComboBox(); self.cat.addItems(CATEGORIES)
        self.brand = QLineEdit()
        self.unit = QComboBox(); self.unit.addItems(UNITS)
        self.cost = QDoubleSpinBox(); self.cost.setRange(0,999999); self.cost.setPrefix("R$ "); self.cost.setDecimals(2)
        self.price = QDoubleSpinBox(); self.price.setRange(0,999999); self.price.setPrefix("R$ "); self.price.setDecimals(2)
        self.stock = QSpinBox(); self.stock.setRange(0,99999)
        self.minst = QSpinBox(); self.minst.setRange(0,99999)
        self.desc = QTextEdit(); self.desc.setMaximumHeight(58)
        rows = [("Codigo *",self.code),("Nome *",self.name),("Categoria",self.cat),("Marca",self.brand),
                ("Unidade",self.unit),("Custo",self.cost),("Preco Venda *",self.price),
                ("Estoque",self.stock),("Minimo",self.minst),("Descricao",self.desc)]
        for r,(l,w) in enumerate(rows): g.addWidget(QLabel(l),r,0); g.addWidget(w,r,1)
        g.setColumnStretch(1,1); lay.addLayout(g); lay.addSpacing(8)
        if product:
            self.name.textChanged.disconnect(self._auto_code)  # disable auto when editing
            self.code.setText(product.get("code","")); self.name.setText(product.get("name",""))
            i = self.cat.findText(product.get("category","")); self.cat.setCurrentIndex(i if i>=0 else 0)
            self.brand.setText(product.get("brand",""))
            i = self.unit.findText(product.get("unit","UN")); self.unit.setCurrentIndex(i if i>=0 else 0)
            self.cost.setValue(float(product.get("cost_price",0))); self.price.setValue(float(product.get("sale_price",0)))
            self.stock.setValue(int(product.get("stock",0))); self.minst.setValue(int(product.get("min_stock",0)))
            self.desc.setPlainText(product.get("description",""))
        b = QHBoxLayout(); b.addStretch()
        cn = QPushButton("Cancelar"); cn.setObjectName("btn_secondary"); cn.clicked.connect(self.reject)
        sv = QPushButton("  Salvar Produto"); sv.clicked.connect(self._save)
        b.addWidget(cn); b.addWidget(sv); lay.addLayout(b)
    def _auto_code(self, text):
        """Generate product code automatically from the product name."""
        if not text.strip():
            self.code.clear(); return
        # Remove accents / special chars, uppercase
        import unicodedata, re
        nfkd = unicodedata.normalize("NFKD", text.upper())
        clean = "".join(c for c in nfkd if not unicodedata.combining(c))
        clean = re.sub(r"[^A-Z0-9 ]", "", clean).strip()
        words = clean.split()
        if len(words) == 0: return
        # Build prefix: first 3 chars of first word, or first char of each word (up to 3)
        if len(words) == 1:
            prefix = words[0][:4]
        elif len(words) == 2:
            prefix = words[0][:3] + words[1][:1]
        else:
            prefix = words[0][:2] + words[1][:1] + words[2][:1]
        prefix = prefix[:4].ljust(3, "X")
        # Find next available number for this prefix
        parent_dm = getattr(self.parent(), "dm", None)
        existing = set()
        if parent_dm:
            for p in parent_dm.get_products():
                existing.add(p.get("code", ""))
        num = 1
        while True:
            candidate = f"{prefix}{num:03d}"
            if candidate not in existing: break
            num += 1
        self.code.setText(candidate)

    def _save(self):
        if not self.code.text().strip(): QMessageBox.warning(self,"Atencao","Informe o codigo."); return
        if not self.name.text().strip(): QMessageBox.warning(self,"Atencao","Informe o nome."); return
        if self.price.value() <= 0: QMessageBox.warning(self,"Atencao","Informe o preco de venda."); return
        self.result_data = {"id": self.product["id"] if self.product else "",
            "code": self.code.text().strip().upper(), "name": self.name.text().strip(),
            "category": self.cat.currentText(), "brand": self.brand.text().strip(),
            "unit": self.unit.currentText(), "cost_price": self.cost.value(),
            "sale_price": self.price.value(), "stock": self.stock.value(),
            "min_stock": self.minst.value(), "description": self.desc.toPlainText().strip()}
        self.accept()

class CustomerDialog(QDialog):
    def __init__(self, parent=None, customer=None):
        super().__init__(parent); self.customer = customer
        self.setWindowTitle("Cliente"); self.setMinimumWidth(440); self.setModal(True)
        lay = QVBoxLayout(self); lay.setContentsMargins(20,20,20,20); lay.setSpacing(10)
        lay.addWidget(SectionTitle("Cadastro de Cliente")); lay.addSpacing(4)
        g = QGridLayout(); g.setSpacing(9)
        self.nm = QLineEdit(); self.doc = QLineEdit(); self.ph = QLineEdit()
        self.em = QLineEdit(); self.ad = QLineEdit()
        rows = [("Nome / Razao Social *",self.nm),("CPF / CNPJ",self.doc),
                ("Telefone",self.ph),("E-mail",self.em),("Endereco",self.ad)]
        for r,(l,w) in enumerate(rows): g.addWidget(QLabel(l),r,0); g.addWidget(w,r,1)
        g.setColumnStretch(1,1); lay.addLayout(g); lay.addSpacing(8)
        if customer:
            self.nm.setText(customer.get("name","")); self.doc.setText(customer.get("cpf_cnpj",""))
            self.ph.setText(customer.get("phone","")); self.em.setText(customer.get("email",""))
            self.ad.setText(customer.get("address",""))
        b = QHBoxLayout(); b.addStretch()
        cn = QPushButton("Cancelar"); cn.setObjectName("btn_secondary"); cn.clicked.connect(self.reject)
        sv = QPushButton("  Salvar Cliente"); sv.clicked.connect(self._save)
        b.addWidget(cn); b.addWidget(sv); lay.addLayout(b)
    def _save(self):
        if not self.nm.text().strip(): QMessageBox.warning(self,"Atencao","Informe o nome."); return
        self.result_data = {"id": self.customer["id"] if self.customer else "",
            "name": self.nm.text().strip(), "cpf_cnpj": self.doc.text().strip(),
            "phone": self.ph.text().strip(), "email": self.em.text().strip(),
            "address": self.ad.text().strip()}
        self.accept()

class UserDialog(QDialog):
    def __init__(self, parent=None, user=None):
        super().__init__(parent); self.user = user
        self.setWindowTitle("Novo Usuario" if not user else "Editar Usuario")
        self.setMinimumWidth(430); self.setModal(True)
        lay = QVBoxLayout(self); lay.setContentsMargins(24,24,24,24); lay.setSpacing(12)
        lay.addWidget(SectionTitle("Usuario do Sistema"))
        info = QLabel("  Operador: Produtos, Nova Venda, Historico e Clientes.  "
                      "   Administrador: acesso completo a todos os modulos.")
        info.setWordWrap(True); info.setObjectName("subtitle")
        info.setStyleSheet("font-size:12px;padding:8px 4px;background:transparent;"); lay.addWidget(info)
        g = QGridLayout(); g.setSpacing(10)
        self.nm = QLineEdit(); self.nm.setPlaceholderText("Nome completo")
        self.us = QLineEdit(); self.us.setPlaceholderText("Login de acesso (sem espacos)")
        self.pw = QLineEdit(); self.pw.setEchoMode(QLineEdit.Password); self.pw.setPlaceholderText("Senha")
        self.p2 = QLineEdit(); self.p2.setEchoMode(QLineEdit.Password); self.p2.setPlaceholderText("Confirme a senha")
        self.rl = QComboBox(); self.rl.addItems(["  Operador (acesso limitado)","  Administrador (acesso total)"])
        rows = [("Nome Completo *",self.nm),("Login / Usuario *",self.us),
                ("Senha *",self.pw),("Confirmar Senha",self.p2),("Perfil de Acesso",self.rl)]
        for r,(l,w) in enumerate(rows): g.addWidget(QLabel(l),r,0); g.addWidget(w,r,1)
        g.setColumnStretch(1,1); lay.addLayout(g); lay.addSpacing(8)
        if user:
            self.nm.setText(user.get("name","")); self.us.setText(user.get("username",""))
            self.us.setEnabled(False)
            self.pw.setPlaceholderText("Deixe em branco para manter a senha")
            self.p2.setPlaceholderText("Confirme apenas se alterar")
            self.rl.setCurrentIndex(1 if user.get("role")=="admin" else 0)
        b = QHBoxLayout(); b.addStretch()
        cn = QPushButton("Cancelar"); cn.setObjectName("btn_secondary"); cn.clicked.connect(self.reject)
        sv = QPushButton("  Salvar Usuario"); sv.clicked.connect(self._save)
        b.addWidget(cn); b.addWidget(sv); lay.addLayout(b)
    def _save(self):
        nm = self.nm.text().strip(); us = self.us.text().strip().lower().replace(" ","")
        pw = self.pw.text(); p2 = self.p2.text()
        if not nm: QMessageBox.warning(self,"Atencao","Informe o nome."); return
        if not us: QMessageBox.warning(self,"Atencao","Informe o login."); return
        if not self.user and not pw: QMessageBox.warning(self,"Atencao","Informe a senha."); return
        if pw and pw != p2: QMessageBox.warning(self,"Atencao","As senhas nao coincidem."); return
        self.result_data = {"id": self.user["id"] if self.user else "", "name": nm, "username": us,
            "password": pw if pw else (self.user.get("password","") if self.user else ""),
            "role": "admin" if self.rl.currentIndex()==1 else "operator", "active": True}
        self.accept()

class SettingsDialog(QDialog):
    def __init__(self, parent, dm):
        super().__init__(parent); self.dm = dm; self.settings = dm.get_settings().copy()
        self._sel = self.settings.get("accent_color","#FF6B35")
        self.setWindowTitle("  Configuracoes"); self.setMinimumWidth(560); self.setMinimumHeight(560); self.setModal(True)
        lay = QVBoxLayout(self); lay.setContentsMargins(24,24,24,24); lay.setSpacing(16)
        lay.addWidget(SectionTitle("  Configuracoes de Aparencia"))

        # THEME
        tg = QGroupBox("  Tema da Interface"); tl = QHBoxLayout(tg); tl.setSpacing(20)
        self.rdark = QRadioButton("  Modo Escuro (Dark)")
        self.rlight = QRadioButton("  Modo Claro (Light)")
        self.rdark.setChecked(self.settings.get("theme","dark") == "dark")
        self.rlight.setChecked(self.settings.get("theme","dark") == "light")
        tl.addWidget(self.rdark); tl.addWidget(self.rlight); tl.addStretch(); lay.addWidget(tg)

        # ACCENT COLOR
        cg = QGroupBox("  Cor Principal do Sistema"); cl = QVBoxLayout(cg); cl.setSpacing(10)
        cl.addWidget(QLabel("Escolha uma cor predefinida ou use o seletor personalizado:"))
        gw = QWidget(); gg = QGridLayout(gw); gg.setSpacing(8); gg.setContentsMargins(0,0,0,0)
        self._swatches = []
        for i,(color,label) in enumerate(PRESET_COLORS):
            btn = SwatchBtn(color, label)
            btn.clicked.connect(lambda _,c=color,b=btn: self._pick_preset(c,b))
            btn.set_selected(color.lower() == self._sel.lower())
            gg.addWidget(btn, i//2, i%2); self._swatches.append((btn,color))
        cl.addWidget(gw)
        cr = QHBoxLayout()
        self.cprev = QPushButton(); self.cprev.setFixedSize(40,40)
        self.cprev.setToolTip("Clique para cor personalizada")
        self.cprev.clicked.connect(self._pick_custom); self._paint_prev(); cr.addWidget(self.cprev)
        cr.addWidget(QLabel("  Cor personalizada (clique no quadrado)", objectName="subtitle"))
        cr.addStretch()
        self.hex_lbl = QLabel(self._sel.upper())
        self.hex_lbl.setStyleSheet("font-family:monospace;font-size:13px;"); cr.addWidget(self.hex_lbl)
        cl.addLayout(cr); lay.addWidget(cg)

        # BACKGROUND
        bg_g = QGroupBox("  Papel de Parede"); bl = QVBoxLayout(bg_g); bl.setSpacing(8)
        self.bgprev = QLabel(); self.bgprev.setFixedHeight(100); self.bgprev.setAlignment(Qt.AlignCenter)
        self.bgprev.setStyleSheet("border:1px dashed #666;border-radius:6px;")
        cur = self.settings.get("background_image","")
        if cur and os.path.exists(cur):
            self.bgprev.setPixmap(QPixmap(cur).scaled(440,100,Qt.KeepAspectRatio,Qt.SmoothTransformation))
        else: self.bgprev.setText("Nenhuma imagem selecionada")
        bl.addWidget(self.bgprev)
        br = QHBoxLayout()
        bc = QPushButton("  Escolher Imagem"); bc.clicked.connect(self._choose_bg)
        bclr = QPushButton("  Remover"); bclr.setObjectName("btn_secondary"); bclr.clicked.connect(self._clear_bg)
        br.addWidget(bc); br.addWidget(bclr); br.addStretch(); bl.addLayout(br)
        bl.addWidget(QLabel("Formatos: JPG, PNG, BMP, WEBP", objectName="subtitle")); lay.addWidget(bg_g)
        lay.addStretch()
        btns = QHBoxLayout(); btns.addStretch()
        cn = QPushButton("Cancelar"); cn.setObjectName("btn_secondary"); cn.clicked.connect(self.reject)
        ap = QPushButton("  Aplicar Configuracoes"); ap.clicked.connect(self._apply)
        btns.addWidget(cn); btns.addWidget(ap); lay.addLayout(btns)

    def _pick_preset(self, color, clicked):
        self._sel = color
        for btn,c in self._swatches: btn.set_selected(c.lower()==color.lower())
        self._paint_prev(); self.hex_lbl.setText(color.upper())
    def _pick_custom(self):
        c = QColorDialog.getColor(QColor(self._sel), self, "Cor Personalizada")
        if c.isValid():
            self._sel = c.name()
            for btn,_ in self._swatches: btn.set_selected(False)
            self._paint_prev(); self.hex_lbl.setText(self._sel.upper())
    def _paint_prev(self):
        self.cprev.setStyleSheet(f"QPushButton{{background:{self._sel};border:2px solid rgba(255,255,255,.4);border-radius:8px;}}"
                                 f"QPushButton:hover{{border:2px solid #FFF;}}")
    def _choose_bg(self):
        p,_ = QFileDialog.getOpenFileName(self,"Imagem de Fundo","","Imagens (*.jpg *.jpeg *.png *.bmp *.webp)")
        if p:
            self.settings["background_image"] = p
            self.bgprev.setPixmap(QPixmap(p).scaled(440,100,Qt.KeepAspectRatio,Qt.SmoothTransformation))
    def _clear_bg(self):
        self.settings["background_image"] = ""; self.bgprev.clear(); self.bgprev.setText("Nenhuma imagem selecionada")
    def _apply(self):
        self.settings["theme"] = "dark" if self.rdark.isChecked() else "light"
        self.settings["accent_color"] = self._sel
        self.dm.save_settings(self.settings); self.accept()


# ═════════════════════════════════════════════════════════════════════════════
#  DASHBOARD
# ═════════════════════════════════════════════════════════════════════════════
class DashboardTab(QWidget):
    def __init__(self, dm, parent=None):
        super().__init__(parent); self.dm = dm; self._build(); self.refresh()
    def _build(self):
        lay = QVBoxLayout(self); lay.setContentsMargins(20,20,20,20); lay.setSpacing(18)
        hdr = QHBoxLayout(); t = QLabel("  Dashboard"); t.setObjectName("title")
        self.tl = QLabel(); self.tl.setObjectName("subtitle")
        hdr.addWidget(t); hdr.addStretch(); hdr.addWidget(self.tl); lay.addLayout(hdr)
        tmr = QTimer(self); tmr.timeout.connect(self._tick); tmr.start(1000); self._tick()
        cr = QHBoxLayout(); cr.setSpacing(14)
        self.c1 = MetricCard("Vendas Hoje","R$ 0,00","","#FF6B35")
        self.c2 = MetricCard("Vendas do Mes","R$ 0,00","","#00C853")
        self.c3 = MetricCard("Produtos","0","","#FFD600")
        self.c4 = MetricCard("Estoque Critico","0","","#FF1744")
        for c in [self.c1,self.c2,self.c3,self.c4]: cr.addWidget(c)
        lay.addLayout(cr)
        bot = QHBoxLayout(); bot.setSpacing(16)
        sf = QFrame(); sf.setObjectName("card"); sl = QVBoxLayout(sf); sl.setContentsMargins(14,14,14,14); sl.setSpacing(8)
        sl.addWidget(SectionTitle("  Ultimas Vendas"))
        self.st = QTableWidget(0,5); self.st.setHorizontalHeaderLabels(["Data","Cliente","Itens","Total","Pagamento"])
        self.st.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.st.setEditTriggers(QAbstractItemView.NoEditTriggers); self.st.setAlternatingRowColors(True)
        self.st.setSelectionBehavior(QAbstractItemView.SelectRows); self.st.verticalHeader().setVisible(False)
        self.st.setMaximumHeight(260); sl.addWidget(self.st); bot.addWidget(sf,3)
        ef = QFrame(); ef.setObjectName("card"); el = QVBoxLayout(ef); el.setContentsMargins(14,14,14,14); el.setSpacing(8)
        el.addWidget(SectionTitle("  Estoque Critico"))
        self.et = QTableWidget(0,4); self.et.setHorizontalHeaderLabels(["Codigo","Produto","Qtd","Minimo"])
        self.et.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.et.setEditTriggers(QAbstractItemView.NoEditTriggers); self.et.setAlternatingRowColors(True)
        self.et.verticalHeader().setVisible(False); self.et.setMaximumHeight(260)
        el.addWidget(self.et); bot.addWidget(ef,2); lay.addLayout(bot); lay.addStretch()
    def _tick(self): self.tl.setText(datetime.now().strftime("  %d/%m/%Y     %H:%M:%S"))
    def refresh(self):
        prods = self.dm.get_products(); sales = self.dm.get_sales(); now = datetime.now()
        td = sum(s.get("total",0) for s in sales if s.get("date","")[:10]==now.strftime("%Y-%m-%d"))
        mo = sum(s.get("total",0) for s in sales if s.get("date","")[:7]==now.strftime("%Y-%m"))
        alerts = [p for p in prods if p.get("stock",0)<=p.get("min_stock",0)]
        self.c1.set_value(fmtR(td)); self.c2.set_value(fmtR(mo))
        self.c3.set_value(str(len(prods))); self.c4.set_value(str(len(alerts)))
        recent = sorted(sales,key=lambda x:x.get("date",""),reverse=True)[:10]
        self.st.setRowCount(len(recent))
        for r,s in enumerate(recent):
            self.st.setItem(r,0,QTableWidgetItem(fmt_date(s.get("date",""))))
            self.st.setItem(r,1,QTableWidgetItem(s.get("customer_name","—")))
            self.st.setItem(r,2,QTableWidgetItem(str(len(s.get("items",[])))))
            ti = QTableWidgetItem(fmtR(s.get("total",0))); ti.setForeground(QColor("#00C853")); self.st.setItem(r,3,ti)
            self.st.setItem(r,4,QTableWidgetItem(s.get("payment_method","—")))
        self.et.setRowCount(len(alerts))
        for r,p in enumerate(alerts):
            self.et.setItem(r,0,QTableWidgetItem(p.get("code",""))); self.et.setItem(r,1,QTableWidgetItem(p.get("name","")))
            si = QTableWidgetItem(str(p.get("stock",0))); si.setForeground(QColor("#FF1744")); self.et.setItem(r,2,si)
            self.et.setItem(r,3,QTableWidgetItem(str(p.get("min_stock",0))))


# ═════════════════════════════════════════════════════════════════════════════
#  PRODUCTS
# ═════════════════════════════════════════════════════════════════════════════
class ProductsTab(QWidget):
    def __init__(self, dm, parent=None):
        super().__init__(parent); self.dm = dm; self._all=[]; self._ids=[]; self._build(); self.refresh()
    def _build(self):
        lay = QVBoxLayout(self); lay.setContentsMargins(20,20,20,20); lay.setSpacing(12)
        lay.addWidget(SectionTitle("  Gestao de Produtos"))
        tb = QHBoxLayout()
        self.srch = SearchBar("Codigo, nome ou marca..."); self.srch.search_changed.connect(self._filter); tb.addWidget(self.srch,3)
        self.catf = QComboBox(); self.catf.addItem("Todas"); self.catf.addItems(CATEGORIES)
        self.catf.currentTextChanged.connect(self._filter); tb.addWidget(self.catf,1)
        self.lowcb = QCheckBox("  Critico"); self.lowcb.stateChanged.connect(self._filter); tb.addWidget(self.lowcb)
        tb.addStretch()
        ba = QPushButton("  Novo"); ba.clicked.connect(self._add); tb.addWidget(ba)
        be = QPushButton("  Editar"); be.setObjectName("btn_secondary"); be.clicked.connect(self._edit); tb.addWidget(be)
        bd = QPushButton("  Excluir"); bd.setObjectName("btn_danger"); bd.clicked.connect(self._delete); tb.addWidget(bd)
        lay.addLayout(tb)
        self.tbl = QTableWidget(0,10)
        self.tbl.setHorizontalHeaderLabels(["Codigo","Nome","Categoria","Marca","Und","Custo","Venda","Estoque","Min","Status"])
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.tbl.horizontalHeader().setSectionResizeMode(1,QHeaderView.Stretch)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers); self.tbl.setAlternatingRowColors(True)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectRows); self.tbl.verticalHeader().setVisible(False)
        self.tbl.doubleClicked.connect(self._edit)
        for c,w in enumerate([80,-1,100,100,55,90,90,70,60,85]):
            if w>0: self.tbl.setColumnWidth(c,w)
        lay.addWidget(self.tbl)
        self.cnt = QLabel(); self.cnt.setObjectName("subtitle"); lay.addWidget(self.cnt)
    def refresh(self): self._all = self.dm.get_products(); self._filter()
    def _filter(self):
        txt = self.srch.edit.text().lower(); cat = self.catf.currentText(); low = self.lowcb.isChecked()
        prods = self._all
        if txt: prods = [p for p in prods if txt in p.get("code","").lower() or txt in p.get("name","").lower() or txt in p.get("brand","").lower()]
        if cat != "Todas": prods = [p for p in prods if p.get("category","") == cat]
        if low: prods = [p for p in prods if p.get("stock",0) <= p.get("min_stock",0)]
        self.tbl.setRowCount(len(prods)); self._ids = []
        for r,p in enumerate(prods):
            self._ids.append(p["id"]); stk=p.get("stock",0); mn=p.get("min_stock",0)
            status = "Esgotado" if stk==0 else ("Baixo" if stk<=mn else "Ativo")
            self.tbl.setItem(r,0,QTableWidgetItem(p.get("code",""))); self.tbl.setItem(r,1,QTableWidgetItem(p.get("name","")))
            self.tbl.setItem(r,2,QTableWidgetItem(p.get("category",""))); self.tbl.setItem(r,3,QTableWidgetItem(p.get("brand","")))
            self.tbl.setItem(r,4,QTableWidgetItem(p.get("unit","")))
            self.tbl.setItem(r,5,QTableWidgetItem(fmtR(p.get("cost_price",0)))); self.tbl.setItem(r,6,QTableWidgetItem(fmtR(p.get("sale_price",0))))
            si = QTableWidgetItem(str(stk)); si.setTextAlignment(Qt.AlignCenter)
            si.setForeground(QColor("#FF1744" if stk==0 else "#FFD600" if stk<=mn else "#00C853")); self.tbl.setItem(r,7,si)
            mi = QTableWidgetItem(str(mn)); mi.setTextAlignment(Qt.AlignCenter); self.tbl.setItem(r,8,mi)
            sti = QTableWidgetItem(status); sti.setTextAlignment(Qt.AlignCenter)
            sti.setForeground(QColor({"Ativo":"#00C853","Baixo":"#FFD600","Esgotado":"#FF1744"}.get(status,"#AAA"))); self.tbl.setItem(r,9,sti)
            self.tbl.setRowHeight(r,36)
        self.cnt.setText(f"Exibindo {len(prods)} de {len(self._all)} produtos")
    def _sel(self):
        r = self.tbl.currentRow(); return self._ids[r] if 0<=r<len(self._ids) else None
    def _add(self):
        dlg = ProductDialog(self)
        if dlg.exec_() == QDialog.Accepted: self.dm.add_product(dlg.result_data); self.refresh()
    def _edit(self):
        pid = self._sel()
        if not pid: QMessageBox.information(self,"Atencao","Selecione um produto."); return
        p = self.dm.get_product_by_id(pid)
        if not p: return
        dlg = ProductDialog(self, p)
        if dlg.exec_() == QDialog.Accepted: dlg.result_data["id"]=pid; self.dm.update_product(pid,dlg.result_data); self.refresh()
    def _delete(self):
        pid = self._sel()
        if not pid: QMessageBox.information(self,"Atencao","Selecione um produto."); return
        p = self.dm.get_product_by_id(pid)
        if p and QMessageBox.question(self,"Confirmar",f"Excluir '{p['name']}'?",QMessageBox.Yes|QMessageBox.No)==QMessageBox.Yes:
            self.dm.delete_product(pid); self.refresh()


# ═════════════════════════════════════════════════════════════════════════════
#  SALES (PDV)
# ═════════════════════════════════════════════════════════════════════════════
class SalesTab(QWidget):
    sale_completed = pyqtSignal()
    def __init__(self, dm, parent=None):
        super().__init__(parent); self.dm=dm; self.cart=[]; self._pids=[]; self._build(); self._sp("")
    def _build(self):
        lay = QVBoxLayout(self); lay.setContentsMargins(20,20,20,20); lay.setSpacing(12)
        lay.addWidget(SectionTitle("  Nova Venda"))
        spl = QSplitter(Qt.Horizontal); spl.setHandleWidth(8)
        left = QWidget(); ll = QVBoxLayout(left); ll.setContentsMargins(0,0,0,0); ll.setSpacing(10)
        sg = QGroupBox("Adicionar Produto"); g = QGridLayout(sg); g.setSpacing(8)
        g.addWidget(QLabel("Produto:"),0,0)
        self.psr = QLineEdit(); self.psr.setPlaceholderText("Codigo ou nome...")
        self.psr.textChanged.connect(self._sp); g.addWidget(self.psr,0,1,1,3)
        self.pl = QTableWidget(0,5); self.pl.setHorizontalHeaderLabels(["Codigo","Nome","Marca","Preco","Estoque"])
        self.pl.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.pl.setEditTriggers(QAbstractItemView.NoEditTriggers); self.pl.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.pl.verticalHeader().setVisible(False); self.pl.setMaximumHeight(155)
        self.pl.selectionModel().selectionChanged.connect(self._osel); g.addWidget(self.pl,1,0,1,4)
        g.addWidget(QLabel("Qtd:"),2,0); self.qty = QSpinBox(); self.qty.setRange(1,9999); self.qty.setValue(1); g.addWidget(self.qty,2,1)
        g.addWidget(QLabel("Preco:"),2,2); self.up = QDoubleSpinBox(); self.up.setRange(0,999999); self.up.setPrefix("R$ "); self.up.setDecimals(2); g.addWidget(self.up,2,3)
        g.addWidget(QLabel("Desc. Item:"),3,0); self.idc = QDoubleSpinBox(); self.idc.setRange(0,100); self.idc.setSuffix("%"); g.addWidget(self.idc,3,1)
        bai = QPushButton("  Adicionar ao Carrinho"); bai.clicked.connect(self._add_item); g.addWidget(bai,3,2,1,2)
        ll.addWidget(sg)
        cg = QGroupBox("Carrinho de Compras"); cl = QVBoxLayout(cg)
        self.ct = QTableWidget(0,7); self.ct.setHorizontalHeaderLabels(["Codigo","Produto","Qtd","Preco","Desc%","Total",""])
        self.ct.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.ct.horizontalHeader().setSectionResizeMode(1,QHeaderView.Stretch)
        self.ct.setEditTriggers(QAbstractItemView.NoEditTriggers); self.ct.setAlternatingRowColors(True)
        self.ct.verticalHeader().setVisible(False)
        for c,w in enumerate([80,-1,50,90,60,90,40]):
            if w>0: self.ct.setColumnWidth(c,w)
        cl.addWidget(self.ct)
        br = QPushButton("  Remover Selecionado"); br.setObjectName("btn_danger"); br.clicked.connect(self._rsel); cl.addWidget(br)
        ll.addWidget(cg); spl.addWidget(left)
        right = QWidget(); rl = QVBoxLayout(right); rl.setContentsMargins(0,0,0,0); rl.setSpacing(12)
        cug = QGroupBox("  Cliente"); cul = QVBoxLayout(cug)
        self.ccb = QComboBox(); self.ccb.addItem("-- Consumidor Final --",None)
        for c in self.dm.get_customers(): self.ccb.addItem(f"{c['name']} | {c.get('cpf_cnpj','')}",c["id"])
        cul.addWidget(self.ccb); rl.addWidget(cug)
        pg = QGroupBox("  Pagamento"); pf = QFormLayout(pg)
        self.pm = QComboBox(); self.pm.addItems(PAYMENTS); pf.addRow("Forma:",self.pm)
        self.gd = QDoubleSpinBox(); self.gd.setRange(0,100); self.gd.setSuffix("%"); self.gd.valueChanged.connect(self._utot); pf.addRow("Desconto Geral:",self.gd)
        self.obs = QTextEdit(); self.obs.setMaximumHeight(58); self.obs.setPlaceholderText("Observacoes..."); pf.addRow("Obs:",self.obs)
        rl.addWidget(pg)
        tg = QGroupBox("  Totais"); tgl = QGridLayout(tg)
        def trow(lbl,attr,r,big=False):
            l=QLabel(lbl); v=QLabel("R$ 0,00"); v.setAlignment(Qt.AlignRight|Qt.AlignVCenter)
            if big: l.setStyleSheet("font-size:16px;font-weight:bold;"); v.setStyleSheet("font-size:20px;font-weight:bold;color:#FF6B35;")
            setattr(self,attr,v); tgl.addWidget(l,r,0); tgl.addWidget(v,r,1)
        trow("Subtotal:","ls",0); trow("Desconto:","ld",1)
        sep=QFrame(); sep.setFrameShape(QFrame.HLine); tgl.addWidget(sep,2,0,1,2)
        trow("TOTAL:","lt",3,True); rl.addWidget(tg); rl.addStretch()
        bf = QPushButton("  FINALIZAR VENDA"); bf.setObjectName("btn_success"); bf.setMinimumHeight(50)
        bf.setFont(QFont("Segoe UI",14,QFont.Bold)); bf.clicked.connect(self._fin); rl.addWidget(bf)
        bnf = QPushButton("  Emitir Nota Fiscal (PDF)"); bnf.setObjectName("btn_secondary"); bnf.setMinimumHeight(40)
        bnf.clicked.connect(self._emitir_nf); rl.addWidget(bnf)
        bc = QPushButton("  Limpar Venda"); bc.setObjectName("btn_secondary"); bc.clicked.connect(self._clr); rl.addWidget(bc)
        spl.addWidget(right); spl.setSizes([650,320]); lay.addWidget(spl,1)
    def _sp(self,txt=""):
        prods = self.dm.get_products()
        if txt: prods = [p for p in prods if txt.lower() in p.get("code","").lower() or txt.lower() in p.get("name","").lower()]
        prods = prods[:60]; self.pl.setRowCount(len(prods)); self._pids=[]
        for r,p in enumerate(prods):
            self._pids.append(p["id"])
            self.pl.setItem(r,0,QTableWidgetItem(p.get("code",""))); self.pl.setItem(r,1,QTableWidgetItem(p.get("name","")))
            self.pl.setItem(r,2,QTableWidgetItem(p.get("brand",""))); self.pl.setItem(r,3,QTableWidgetItem(fmtR(p.get("sale_price",0))))
            si=QTableWidgetItem(str(p.get("stock",0))); si.setForeground(QColor("#FF1744" if p.get("stock",0)==0 else "#00C853")); self.pl.setItem(r,4,si)
            self.pl.setRowHeight(r,30)
    def _osel(self):
        r=self.pl.currentRow()
        if 0<=r<len(self._pids):
            p=self.dm.get_product_by_id(self._pids[r])
            if p: self.up.setValue(p.get("sale_price",0))
    def _add_item(self):
        r=self.pl.currentRow()
        if r<0 or r>=len(self._pids): QMessageBox.information(self,"Atencao","Selecione um produto."); return
        pid=self._pids[r]; p=self.dm.get_product_by_id(pid)
        if not p: return
        qty=self.qty.value(); price=self.up.value(); disc=self.idc.value()
        if p.get("stock",0)<qty: QMessageBox.warning(self,"Estoque",f"Disponivel: {p.get('stock',0)} {p.get('unit','UN')}"); return
        for it in self.cart:
            if it["product_id"]==pid: it["quantity"]+=qty; it["unit_price"]=price; it["discount"]=disc; self._rc(); return
        self.cart.append({"product_id":pid,"code":p.get("code",""),"name":p.get("name",""),"quantity":qty,"unit_price":price,"discount":disc})
        self._rc(); self.qty.setValue(1); self.idc.setValue(0)
    def _rc(self):
        self.ct.setRowCount(len(self.cart))
        for r,it in enumerate(self.cart):
            df=1-it["discount"]/100; tot=it["quantity"]*it["unit_price"]*df
            self.ct.setItem(r,0,QTableWidgetItem(it["code"])); self.ct.setItem(r,1,QTableWidgetItem(it["name"]))
            self.ct.setItem(r,2,QTableWidgetItem(str(it["quantity"]))); self.ct.setItem(r,3,QTableWidgetItem(fmtR(it["unit_price"])))
            self.ct.setItem(r,4,QTableWidgetItem(f"{it['discount']:.1f}%"))
            ti=QTableWidgetItem(fmtR(tot)); ti.setForeground(QColor("#00C853")); self.ct.setItem(r,5,ti)
            db=QPushButton("x"); db.setFixedSize(28,28); db.setObjectName("btn_danger")
            db.clicked.connect(lambda _,i=r: self._ri(i)); self.ct.setCellWidget(r,6,db)
            self.ct.setRowHeight(r,36)
        self._utot()
    def _ri(self,r):
        if 0<=r<len(self.cart): self.cart.pop(r); self._rc()
    def _rsel(self): self._ri(self.ct.currentRow())
    def _utot(self):
        sub=sum(i["quantity"]*i["unit_price"]*(1-i["discount"]/100) for i in self.cart)
        dp=self.gd.value(); dv=sub*dp/100; tot=sub-dv
        self.ls.setText(fmtR(sub)); self.ld.setText(f"-{fmtR(dv)}"); self.lt.setText(fmtR(tot))
    def _fin(self):
        if not self.cart: QMessageBox.warning(self,"Vazio","Adicione produtos ao carrinho."); return
        sub=sum(i["quantity"]*i["unit_price"]*(1-i["discount"]/100) for i in self.cart)
        dp=self.gd.value(); dv=sub*dp/100; tot=sub-dv
        cid=self.ccb.currentData(); cn="Consumidor Final" if cid is None else self.ccb.currentText().split("|")[0].strip()
        sale={"customer_id":cid or "","customer_name":cn,"items":self.cart.copy(),"subtotal":sub,
              "discount_pct":dp,"discount_value":dv,"total":tot,"payment_method":self.pm.currentText(),
              "observations":self.obs.toPlainText(),"status":"concluida"}
        sid=self.dm.add_sale(sale); self.sale_completed.emit()
        QMessageBox.information(self,"  Venda Concluida",f"Venda registrada!\n\nCliente: {cn}\nTotal: {fmtR(tot)}\nPagamento: {self.pm.currentText()}\nID: {sid[:8]}...")
        self._clr()
    def _emitir_nf(self):
        if not self.cart:
            QMessageBox.warning(self, "Carrinho Vazio", "Adicione produtos ao carrinho antes de emitir a nota."); return
        if not PDF_OK:
            QMessageBox.critical(self, "Erro", "Biblioteca 'reportlab' nao instalada.\nExecute: pip install reportlab"); return
        sub = sum(i["quantity"]*i["unit_price"]*(1-i["discount"]/100) for i in self.cart)
        dp = self.gd.value(); dv = sub*dp/100; tot = sub-dv
        cid = self.ccb.currentData()
        cn = "Consumidor Final" if cid is None else self.ccb.currentText().split("|")[0].strip()
        cust_obj = self.dm.get_customer_by_id(cid) if cid else None
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"nota_fiscal_{ts}.pdf"
        path, _ = QFileDialog.getSaveFileName(self, "Salvar Nota Fiscal", default_name, "PDF (*.pdf)")
        if not path: return
        try:
            _gerar_nota_fiscal(path, self.cart, cn, cust_obj, sub, dp, dv, tot,
                               self.pm.currentText(), self.obs.toPlainText(), self.dm)
            reply = QMessageBox.information(self, "  Nota Fiscal Gerada",
                f"Nota fiscal salva em:\n{path}\n\nDeseja abrir o arquivo agora?",
                QMessageBox.Open | QMessageBox.Close)
            if reply == QMessageBox.Open:
                import subprocess, sys
                if sys.platform == "win32": os.startfile(path)
                elif sys.platform == "darwin": subprocess.call(["open", path])
                else: subprocess.call(["xdg-open", path])
        except Exception as e:
            QMessageBox.critical(self, "Erro ao Gerar PDF", str(e))

    def _clr(self):
        self.cart=[]; self._rc(); self.ccb.setCurrentIndex(0); self.pm.setCurrentIndex(0); self.gd.setValue(0); self.obs.clear()
    def refresh_customers(self):
        self.ccb.clear(); self.ccb.addItem("-- Consumidor Final --",None)
        for c in self.dm.get_customers(): self.ccb.addItem(f"{c['name']} | {c.get('cpf_cnpj','')}",c["id"])


# ═════════════════════════════════════════════════════════════════════════════
#  CUSTOMERS
# ═════════════════════════════════════════════════════════════════════════════
class CustomersTab(QWidget):
    def __init__(self, dm, parent=None):
        super().__init__(parent); self.dm=dm; self._all=[]; self._ids=[]; self._build(); self.refresh()
    def _build(self):
        lay = QVBoxLayout(self); lay.setContentsMargins(20,20,20,20); lay.setSpacing(12)
        lay.addWidget(SectionTitle("  Gestao de Clientes"))
        tb = QHBoxLayout()
        self.srch = SearchBar("Nome, CPF/CNPJ ou telefone..."); self.srch.search_changed.connect(self._filter); tb.addWidget(self.srch,3); tb.addStretch()
        for lbl,obj,cb in [("  Novo","ba",self._add),("  Editar","be",self._edit),("  Excluir","bd",self._del),("  Historico","bh",self._hist)]:
            btn=QPushButton(lbl)
            if lbl=="  Excluir": btn.setObjectName("btn_danger")
            elif lbl!="  Novo": btn.setObjectName("btn_secondary")
            btn.clicked.connect(cb); tb.addWidget(btn)
        lay.addLayout(tb)
        self.tbl = QTableWidget(0,6); self.tbl.setHorizontalHeaderLabels(["Nome","CPF/CNPJ","Telefone","E-mail","Endereco","Total Compras"])
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.tbl.horizontalHeader().setSectionResizeMode(0,QHeaderView.Stretch)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers); self.tbl.setAlternatingRowColors(True)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectRows); self.tbl.verticalHeader().setVisible(False)
        self.tbl.doubleClicked.connect(self._edit)
        for c,w in enumerate([-1,130,130,160,200,110]):
            if w>0: self.tbl.setColumnWidth(c,w)
        lay.addWidget(self.tbl)
        self.cnt = QLabel(); self.cnt.setObjectName("subtitle"); lay.addWidget(self.cnt)
    def refresh(self): self._all=self.dm.get_customers(); self._filter()
    def _filter(self):
        txt=self.srch.edit.text().lower()
        custs=[c for c in self._all if not txt or txt in c.get("name","").lower() or txt in c.get("cpf_cnpj","").lower() or txt in c.get("phone","").lower()]
        self.tbl.setRowCount(len(custs)); self._ids=[]
        sales=self.dm.get_sales()
        for r,c in enumerate(custs):
            self._ids.append(c["id"]); tot=sum(s.get("total",0) for s in sales if s.get("customer_id")==c["id"])
            self.tbl.setItem(r,0,QTableWidgetItem(c.get("name",""))); self.tbl.setItem(r,1,QTableWidgetItem(c.get("cpf_cnpj","")))
            self.tbl.setItem(r,2,QTableWidgetItem(c.get("phone",""))); self.tbl.setItem(r,3,QTableWidgetItem(c.get("email","")))
            self.tbl.setItem(r,4,QTableWidgetItem(c.get("address","")))
            ti=QTableWidgetItem(fmtR(tot)); ti.setForeground(QColor("#00C853")); self.tbl.setItem(r,5,ti); self.tbl.setRowHeight(r,36)
        self.cnt.setText(f"Exibindo {len(custs)} de {len(self._all)} clientes")
    def _sid(self):
        r=self.tbl.currentRow(); return self._ids[r] if 0<=r<len(self._ids) else None
    def _add(self):
        dlg=CustomerDialog(self)
        if dlg.exec_()==QDialog.Accepted: self.dm.add_customer(dlg.result_data); self.refresh()
    def _edit(self):
        cid=self._sid()
        if not cid: QMessageBox.information(self,"Atencao","Selecione um cliente."); return
        c=self.dm.get_customer_by_id(cid)
        if not c: return
        dlg=CustomerDialog(self,c)
        if dlg.exec_()==QDialog.Accepted: dlg.result_data["id"]=cid; self.dm.update_customer(cid,dlg.result_data); self.refresh()
    def _del(self):
        cid=self._sid()
        if not cid: QMessageBox.information(self,"Atencao","Selecione um cliente."); return
        c=self.dm.get_customer_by_id(cid)
        if c and QMessageBox.question(self,"Confirmar",f"Excluir '{c['name']}'?",QMessageBox.Yes|QMessageBox.No)==QMessageBox.Yes:
            self.dm.delete_customer(cid); self.refresh()
    def _hist(self):
        cid=self._sid()
        if not cid: QMessageBox.information(self,"Atencao","Selecione um cliente."); return
        c=self.dm.get_customer_by_id(cid)
        if not c: return
        sales=[s for s in self.dm.get_sales() if s.get("customer_id")==cid]
        dlg=QDialog(self); dlg.setWindowTitle(f"Historico — {c['name']}"); dlg.setMinimumSize(640,400)
        l=QVBoxLayout(dlg); l.setContentsMargins(16,16,16,16)
        l.addWidget(QLabel(f"<b>Cliente:</b> {c['name']} | <b>Total vendas:</b> {len(sales)}"))
        t=QTableWidget(len(sales),5); t.setHorizontalHeaderLabels(["Data","Itens","Subtotal","Desconto","Total"])
        t.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch); t.setEditTriggers(QAbstractItemView.NoEditTriggers); t.verticalHeader().setVisible(False)
        for r,s in enumerate(sorted(sales,key=lambda x:x.get("date",""),reverse=True)):
            t.setItem(r,0,QTableWidgetItem(fmt_date(s.get("date",""))))
            t.setItem(r,1,QTableWidgetItem(str(len(s.get("items",[])))))
            t.setItem(r,2,QTableWidgetItem(fmtR(s.get("subtotal",0)))); t.setItem(r,3,QTableWidgetItem(fmtR(s.get("discount_value",0))))
            ti=QTableWidgetItem(fmtR(s.get("total",0))); ti.setForeground(QColor("#00C853")); t.setItem(r,4,ti)
        l.addWidget(t)
        tg=sum(s.get("total",0) for s in sales)
        lb=QLabel(f"<b>Total Gasto: {fmtR(tg)}</b>"); lb.setStyleSheet("color:#FF6B35;font-size:14px;"); l.addWidget(lb)
        btn=QPushButton("Fechar"); btn.clicked.connect(dlg.accept); l.addWidget(btn); dlg.exec_()


# ═════════════════════════════════════════════════════════════════════════════
#  REPORTS — dias em Portugues + exportacao Excel
# ═════════════════════════════════════════════════════════════════════════════
class ReportsTab(QWidget):
    def __init__(self, dm, parent=None):
        super().__init__(parent); self.dm=dm; self._hdr=[]; self._data=[]; self._build()
    def _build(self):
        lay = QVBoxLayout(self); lay.setContentsMargins(20,20,20,20); lay.setSpacing(14)
        lay.addWidget(SectionTitle("  Relatorios"))
        ctrl = QHBoxLayout(); ctrl.setSpacing(10)
        ctrl.addWidget(QLabel("Relatorio:"))
        self.rt = QComboBox(); self.rt.setMinimumWidth(245)
        self.rt.addItems(["Vendas por Periodo","Vendas por Cliente","Vendas por Produto",
            "Estoque Atual","Produtos Criticos (Estoque Baixo)","Resumo Financeiro","Ranking Mais Vendidos"])
        ctrl.addWidget(self.rt)
        ctrl.addWidget(QLabel("De:"))
        self.df = QDateEdit(QDate.currentDate().addDays(-30)); self.df.setCalendarPopup(True); self.df.setDisplayFormat("dd/MM/yyyy"); ctrl.addWidget(self.df)
        ctrl.addWidget(QLabel("Ate:"))
        self.dt2 = QDateEdit(QDate.currentDate()); self.dt2.setCalendarPopup(True); self.dt2.setDisplayFormat("dd/MM/yyyy"); ctrl.addWidget(self.dt2)
        bg = QPushButton("  Gerar Relatorio"); bg.clicked.connect(self._gen); ctrl.addWidget(bg)
        if EXCEL_OK:
            bx = QPushButton("  Exportar Excel"); bx.setObjectName("btn_success"); bx.clicked.connect(self._export); ctrl.addWidget(bx)
        ctrl.addStretch(); lay.addLayout(ctrl)
        sr = QHBoxLayout(); sr.setSpacing(12)
        self.s1=MetricCard("Registros","0","","#FF6B35"); self.s2=MetricCard("Valor Total","R$ 0,00","","#00C853")
        self.s3=MetricCard("Media","R$ 0,00","","#FFD600"); self.s4=MetricCard("Periodo","0 dias","","#1E88E5")
        for c in [self.s1,self.s2,self.s3,self.s4]: c.setMaximumHeight(100); sr.addWidget(c)
        lay.addLayout(sr)
        self.tbl = QTableWidget(0,1); self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl.setAlternatingRowColors(True); self.tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl.verticalHeader().setVisible(False); self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        lay.addWidget(self.tbl,1)
        self.info = QLabel("Selecione um relatorio e clique em Gerar"); self.info.setObjectName("subtitle"); self.info.setAlignment(Qt.AlignCenter); lay.addWidget(self.info)
        self._gen()
    def _gen(self):
        rt=self.rt.currentText(); d0=self.df.date().toString("yyyy-MM-dd"); d1=self.dt2.date().toString("yyyy-MM-dd")+"T23:59:59"
        sales=[s for s in self.dm.get_sales() if d0<=s.get("date","")<=d1]
        self.s4.set_value(f"{self.df.date().daysTo(self.dt2.date())+1} dias")
        m={"Vendas por Periodo":self._period,"Vendas por Cliente":self._customer,"Vendas por Produto":self._product,
           "Estoque Atual":self._stock,"Produtos Criticos (Estoque Baixo)":self._lowstock,
           "Resumo Financeiro":self._financial,"Ranking Mais Vendidos":self._top}
        fn=m.get(rt)
        if fn: fn() if rt in ("Estoque Atual","Produtos Criticos (Estoque Baixo)") else fn(sales)
    def _st(self, headers, rows, mc=None):
        self._hdr=headers; self._data=rows; mc=mc or []
        self.tbl.setColumnCount(len(headers)); self.tbl.setHorizontalHeaderLabels(headers); self.tbl.setRowCount(len(rows))
        for r,row in enumerate(rows):
            for c,val in enumerate(row):
                it=QTableWidgetItem(str(val))
                if c in mc: it.setForeground(QColor("#00C853"))
                self.tbl.setItem(r,c,it); self.tbl.setRowHeight(r,32)
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.info.setText(f"{len(rows)} registro(s).")
    def _upd(self,count,total):
        self.s1.set_value(str(count)); self.s2.set_value(fmtR(total)); self.s3.set_value(fmtR(total/max(1,count)))
    def _period(self, sales):
        bd=defaultdict(lambda:{"count":0,"total":0})
        for s in sales: d=s.get("date","")[:10]; bd[d]["count"]+=1; bd[d]["total"]+=s.get("total",0)
        rows=[]
        for day in sorted(bd.keys(),reverse=True):
            d=datetime.strptime(day,"%Y-%m-%d")
            dia = DIAS_PT.get(d.strftime("%A"), d.strftime("%A"))  # PORTUGUES!
            rows.append([d.strftime("%d/%m/%Y"), dia, bd[day]["count"], fmtR(bd[day]["total"])])
        self._st(["Data","Dia da Semana","N Vendas","Total"],rows,[3]); self._upd(len(sales),sum(s.get("total",0) for s in sales))
    def _customer(self, sales):
        bc=defaultdict(lambda:{"count":0,"total":0,"name":""})
        for s in sales: k=s.get("customer_id","") or "cf"; bc[k]["count"]+=1; bc[k]["total"]+=s.get("total",0); bc[k]["name"]=s.get("customer_name","Consumidor Final")
        rows=sorted(bc.values(),key=lambda x:x["total"],reverse=True)
        self._st(["Cliente","N Vendas","Total"],[[r["name"],r["count"],fmtR(r["total"])] for r in rows],[2]); self._upd(len(sales),sum(s.get("total",0) for s in sales))
    def _product(self, sales):
        bp=defaultdict(lambda:{"qty":0,"total":0,"name":"","code":""})
        for s in sales:
            for it in s.get("items",[]): pid=it.get("product_id",""); bp[pid]["qty"]+=it.get("quantity",0); bp[pid]["total"]+=it["quantity"]*it["unit_price"]*(1-it.get("discount",0)/100); bp[pid]["name"]=it.get("name",""); bp[pid]["code"]=it.get("code","")
        rows=sorted(bp.values(),key=lambda x:x["total"],reverse=True)
        self._st(["Codigo","Produto","Qtd Vendida","Total"],[[r["code"],r["name"],r["qty"],fmtR(r["total"])] for r in rows],[3]); self._upd(len(rows),sum(r["total"] for r in rows))
    def _stock(self):
        prods=self.dm.get_products(); rows=[]
        for p in sorted(prods,key=lambda x:x.get("name","")):
            stk=p.get("stock",0); status="Esgotado" if stk==0 else ("Critico" if stk<=p.get("min_stock",0) else "Normal")
            rows.append([p.get("code",""),p.get("name",""),p.get("category",""),p.get("brand",""),str(stk),str(p.get("min_stock",0)),fmtR(p.get("sale_price",0)),fmtR(stk*p.get("cost_price",0)),status])
        tv=sum(p.get("stock",0)*p.get("cost_price",0) for p in prods)
        self._st(["Codigo","Produto","Categoria","Marca","Estoque","Minimo","Preco","Val.Estoque","Status"],rows,[6,7])
        self.s1.set_value(str(len(prods))); self.s2.set_value(fmtR(tv)); self.s3.set_value(fmtR(tv/max(1,len(prods)))); self.info.setText(f"{len(prods)} produtos.")
    def _lowstock(self):
        prods=[p for p in self.dm.get_products() if p.get("stock",0)<=p.get("min_stock",0)]
        rows=[[p.get("code",""),p.get("name",""),p.get("category",""),str(p.get("stock",0)),str(p.get("min_stock",0)),str(max(0,p.get("min_stock",0)-p.get("stock",0))),fmtR(p.get("sale_price",0))] for p in sorted(prods,key=lambda x:x.get("stock",0))]
        self._st(["Codigo","Produto","Categoria","Estoque","Minimo","Qtd p/Repor","Preco"],rows,[6])
        self.s1.set_value(str(len(prods))); self.s2.set_value("--"); self.s3.set_value("--"); self.info.setText(f"{len(prods)} criticos.")
    def _financial(self, sales):
        sb=sum(s.get("subtotal",0) for s in sales); sd=sum(s.get("discount_value",0) for s in sales); sl=sum(s.get("total",0) for s in sales)
        bp=defaultdict(float)
        for s in sales: bp[s.get("payment_method","--")]+=s.get("total",0)
        rows=[[m,fmtR(v),f"{v/max(1,sl)*100:.1f}%"] for m,v in sorted(bp.items(),key=lambda x:x[1],reverse=True)]
        rows+=[["-"*20,"-"*10,""],[" TOTAL BRUTO",fmtR(sb),"100%"],[" DESCONTOS",f"-{fmtR(sd)}",""],[" TOTAL LIQUIDO",fmtR(sl),""]]
        self._st(["Forma de Pagamento","Valor","% do Total"],rows,[1]); self._upd(len(sales),sl)
    def _top(self, sales):
        bp=defaultdict(lambda:{"qty":0,"total":0,"name":"","code":""})
        for s in sales:
            for it in s.get("items",[]): pid=it.get("product_id",""); bp[pid]["qty"]+=it.get("quantity",0); bp[pid]["total"]+=it["quantity"]*it["unit_price"]*(1-it.get("discount",0)/100); bp[pid]["name"]=it.get("name",""); bp[pid]["code"]=it.get("code","")
        rs=sorted(bp.values(),key=lambda x:x["qty"],reverse=True)[:20]
        self._st(["Pos","Codigo","Produto","Qtd Vendida","Total"],[[f"#{i+1}",r["code"],r["name"],r["qty"],fmtR(r["total"])] for i,r in enumerate(rs)],[4]); self._upd(len(rs),sum(r["total"] for r in rs))
    def _export(self):
        if not EXCEL_OK: QMessageBox.warning(self,"Erro","openpyxl nao esta instalado."); return
        if not self._data: QMessageBox.information(self,"Atencao","Gere um relatorio antes de exportar."); return
        ts=datetime.now().strftime("%Y%m%d_%H%M%S")
        sn=self.rt.currentText().replace(" ","_").replace("/","_").replace("(","").replace(")","")
        path,_=QFileDialog.getSaveFileName(self,"Salvar Excel",f"relatorio_{sn}_{ts}.xlsx","Excel (*.xlsx)")
        if not path: return
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="Relatorio"
        hf=PatternFill(start_color="FF6B35",end_color="FF6B35",fill_type="solid")
        hfont=XLFont(bold=True,color="FFFFFF",size=11)
        halign=Alignment(horizontal="center",vertical="center")
        thin=Side(style="thin",color="CCCCCC"); bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
        altf=PatternFill(start_color="F5F5F5",end_color="F5F5F5",fill_type="solid")
        nc=len(self._hdr); lc=get_column_letter(nc)
        ws.merge_cells(f"A1:{lc}1")
        t=ws["A1"]; t.value=f"SISTEMA AUTO PECAS -- {self.rt.currentText().upper()}"
        t.font=XLFont(bold=True,size=14,color="FF6B35"); t.alignment=Alignment(horizontal="center"); ws.row_dimensions[1].height=28
        ws.merge_cells(f"A2:{lc}2")
        p=ws["A2"]; p.value=f"Periodo: {self.df.date().toString('dd/MM/yyyy')} a {self.dt2.date().toString('dd/MM/yyyy')}  |  Gerado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        p.font=XLFont(size=10,color="888888"); p.alignment=Alignment(horizontal="center"); ws.row_dimensions[2].height=18
        for c,h in enumerate(self._hdr,1):
            cell=ws.cell(row=4,column=c,value=h); cell.fill=hf; cell.font=hfont; cell.alignment=halign; cell.border=bdr
        ws.row_dimensions[4].height=22
        for r,row in enumerate(self._data,5):
            for c,val in enumerate(row,1):
                cell=ws.cell(row=r,column=c,value=val); cell.alignment=Alignment(horizontal="left",vertical="center"); cell.border=bdr
                if (r-5)%2==1: cell.fill=altf
            ws.row_dimensions[r].height=18
        for col in ws.columns:
            ml=0; cl2=get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value: ml=max(ml,len(str(cell.value)))
                except: pass
            ws.column_dimensions[cl2].width=min(50,max(10,ml+4))
        lr=len(self._data)+6
        ws.cell(row=lr,column=1,value="Total de Registros:").font=XLFont(bold=True); ws.cell(row=lr,column=2,value=self.s1.val.text())
        ws.cell(row=lr+1,column=1,value="Valor Total:").font=XLFont(bold=True); ws.cell(row=lr+1,column=2,value=self.s2.val.text())
        try: wb.save(path); QMessageBox.information(self,"  Exportado!",f"Salvo em:\n{path}")
        except Exception as e: QMessageBox.critical(self,"Erro",str(e))


# ═════════════════════════════════════════════════════════════════════════════
#  SALES HISTORY
# ═════════════════════════════════════════════════════════════════════════════
class HistoryTab(QWidget):
    def __init__(self, dm, user=None, parent=None):
        super().__init__(parent)
        self.dm = dm
        self.user = user or {}
        self.is_admin = self.user.get("role") == "admin"
        self._sales = []
        self._build()
        self.refresh()

    def set_user(self, user):
        self.user = user or {}
        self.is_admin = self.user.get("role") == "admin"

    def _build(self):
        lay = QVBoxLayout(self); lay.setContentsMargins(20,20,20,20); lay.setSpacing(12)
        lay.addWidget(SectionTitle("  Historico de Vendas"))
        tb = QHBoxLayout()
        self.srch = SearchBar("Cliente ou ID..."); self.srch.search_changed.connect(self.refresh)
        tb.addWidget(self.srch, 2)
        tb.addWidget(QLabel("De:"))
        self.df = QDateEdit(QDate.currentDate().addDays(-30))
        self.df.setCalendarPopup(True); self.df.setDisplayFormat("dd/MM/yyyy"); tb.addWidget(self.df)
        tb.addWidget(QLabel("Ate:"))
        self.dt = QDateEdit(QDate.currentDate())
        self.dt.setCalendarPopup(True); self.dt.setDisplayFormat("dd/MM/yyyy"); tb.addWidget(self.dt)
        bf = QPushButton("  Filtrar"); bf.clicked.connect(self.refresh); tb.addWidget(bf)
        self.show_cancelled = QCheckBox("Mostrar canceladas")
        self.show_cancelled.stateChanged.connect(self.refresh); tb.addWidget(self.show_cancelled)
        tb.addStretch()
        bd = QPushButton("  Detalhar"); bd.setObjectName("btn_secondary"); bd.clicked.connect(self._detail); tb.addWidget(bd)
        self.btn_cancel = QPushButton("  Cancelar Venda"); self.btn_cancel.setObjectName("btn_warning")
        self.btn_cancel.clicked.connect(self._cancel_sale); tb.addWidget(self.btn_cancel)
        self.btn_delete = QPushButton("  Excluir Venda"); self.btn_delete.setObjectName("btn_danger")
        self.btn_delete.clicked.connect(self._delete_sale); tb.addWidget(self.btn_delete)
        lay.addLayout(tb)
        self.tbl = QTableWidget(0, 8)
        self.tbl.setHorizontalHeaderLabels(["Data/Hora","Cliente","Itens","Subtotal","Desconto","Total","Pagamento","Status"])
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers); self.tbl.setAlternatingRowColors(True)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectRows); self.tbl.verticalHeader().setVisible(False)
        self.tbl.setColumnWidth(0, 140); self.tbl.setColumnWidth(7, 100)
        self.tbl.doubleClicked.connect(self._detail)
        lay.addWidget(self.tbl, 1)
        self.info = QLabel(); self.info.setObjectName("subtitle"); lay.addWidget(self.info)

    def refresh(self):
        d0 = self.df.date().toString("yyyy-MM-dd")
        d1 = self.dt.date().toString("yyyy-MM-dd") + "T23:59:59"
        txt = self.srch.edit.text().lower()
        show_all = self.show_cancelled.isChecked()
        sales = [s for s in self.dm.get_sales() if d0 <= s.get("date","") <= d1]
        if not show_all:
            sales = [s for s in sales if s.get("status","concluida") != "cancelada"]
        if txt:
            sales = [s for s in sales if txt in s.get("customer_name","").lower()
                     or txt in s.get("id","").lower()]
        self._sales = sorted(sales, key=lambda x: x.get("date",""), reverse=True)
        self.tbl.setRowCount(len(self._sales))
        tot = 0
        for r, s in enumerate(self._sales):
            cancelled = s.get("status","") == "cancelada"
            if not cancelled: tot += s.get("total", 0)
            self.tbl.setItem(r, 0, QTableWidgetItem(fmt_date(s.get("date",""))))
            self.tbl.setItem(r, 1, QTableWidgetItem(s.get("customer_name","--")))
            self.tbl.setItem(r, 2, QTableWidgetItem(str(len(s.get("items",[])))))
            self.tbl.setItem(r, 3, QTableWidgetItem(fmtR(s.get("subtotal",0))))
            self.tbl.setItem(r, 4, QTableWidgetItem(fmtR(s.get("discount_value",0))))
            ti = QTableWidgetItem(fmtR(s.get("total",0)))
            ti.setForeground(QColor("#888888" if cancelled else "#00C853"))
            self.tbl.setItem(r, 5, ti)
            self.tbl.setItem(r, 6, QTableWidgetItem(s.get("payment_method","--")))
            st = QTableWidgetItem("Cancelada" if cancelled else "Concluida")
            st.setTextAlignment(Qt.AlignCenter)
            st.setForeground(QColor("#FF1744" if cancelled else "#00C853"))
            self.tbl.setItem(r, 7, st)
            if cancelled:
                for c in range(8):
                    cell = self.tbl.item(r, c)
                    if cell and c != 7: cell.setForeground(QColor("#888888"))
            self.tbl.setRowHeight(r, 36)
        self.info.setText(f"{len(self._sales)} venda(s) | Total do periodo: {fmtR(tot)}")

    def _sel_sale(self):
        r = self.tbl.currentRow()
        if r < 0 or r >= len(self._sales):
            QMessageBox.information(self, "Atencao", "Selecione uma venda na tabela.")
            return None
        return self._sales[r]

    def _request_auth(self, action):
        dlg = QDialog(self); dlg.setWindowTitle("Autorizacao Necessaria")
        dlg.setFixedWidth(390); dlg.setModal(True)
        lay = QVBoxLayout(dlg); lay.setContentsMargins(24,24,24,24); lay.setSpacing(14)
        lay.addWidget(QLabel("  Autorizacao Necessaria",
                             styleSheet="font-size:15px;font-weight:bold;"))
        lay.addWidget(QLabel(
            f"Para {action.lower()} esta venda, informe a senha de autorizacao do administrador.",
            wordWrap=True, styleSheet="font-size:12px;"))
        form = QHBoxLayout(); form.addWidget(QLabel("Senha:"))
        self._pwd_e = QLineEdit(); self._pwd_e.setEchoMode(QLineEdit.Password)
        self._pwd_e.setPlaceholderText("Senha de autorizacao")
        self._pwd_e.setMinimumHeight(38); form.addWidget(self._pwd_e, 1); lay.addLayout(form)
        self._pwd_err = QLabel(""); self._pwd_err.setStyleSheet("color:#FF1744;font-size:12px;")
        lay.addWidget(self._pwd_err)
        btns = QHBoxLayout(); btns.addStretch()
        bcl = QPushButton("Cancelar"); bcl.setObjectName("btn_secondary"); bcl.clicked.connect(dlg.reject)
        bok = QPushButton(f"  Confirmar {action}")
        def _check():
            if self._pwd_e.text() == "admin": dlg.accept()
            else: self._pwd_err.setText("  Senha incorreta!"); self._pwd_e.clear(); self._pwd_e.setFocus()
        bok.clicked.connect(_check); self._pwd_e.returnPressed.connect(_check)
        btns.addWidget(bcl); btns.addWidget(bok); lay.addLayout(btns)
        return dlg.exec_() == QDialog.Accepted

    def _cancel_sale(self):
        s = self._sel_sale()
        if not s: return
        if s.get("status") == "cancelada":
            QMessageBox.information(self, "Aviso", "Esta venda ja esta cancelada."); return
        if not self.is_admin:
            if not self._request_auth("Cancelar Venda"): return
        msg = "Cancelar a venda abaixo?\n\nData: {}\nCliente: {}\nTotal: {}\n\nO estoque dos produtos sera restaurado automaticamente.".format(fmt_date(s.get('date','')), s.get('customer_name','--'), fmtR(s.get('total',0)))
        if QMessageBox.question(self, "Confirmar Cancelamento", msg,
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            ok, msg2 = self.dm.cancel_sale(s["id"])
            if ok: QMessageBox.information(self, "  Cancelada", msg2); self.refresh()
            else: QMessageBox.warning(self, "Erro", msg2)

    def _delete_sale(self):
        s = self._sel_sale()
        if not s: return
        if not self.is_admin:
            if not self._request_auth("Excluir Venda"): return
        msg = "ATENCAO: Acao IRREVERSIVEL!\n\nExcluir permanentemente?\n\nData: " + fmt_date(s.get('date','')) + "\nCliente: " + s.get('customer_name','--') + "\nTotal: " + fmtR(s.get('total',0)) + "\n\nDica: cancele antes de excluir para restaurar o estoque."
        if QMessageBox.question(self, "  Confirmar Exclusao", msg,
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            self.dm.delete_sale(s["id"])
            QMessageBox.information(self, "  Excluida", "Venda excluida permanentemente.")
            self.refresh()

    def _detail(self):
        r = self.tbl.currentRow()
        if r < 0 or r >= len(self._sales): return
        s = self._sales[r]; cancelled = s.get("status","") == "cancelada"
        dlg = QDialog(self); dlg.setWindowTitle(f"Venda #{s.get('id','')[:8]}")
        dlg.setMinimumSize(580, 460)
        l = QVBoxLayout(dlg); l.setContentsMargins(16,16,16,16); l.setSpacing(10)
        if cancelled:
            bn = QLabel("  VENDA CANCELADA  "); bn.setAlignment(Qt.AlignCenter)
            bn.setStyleSheet("background:#FF1744;color:#FFF;font-weight:bold;font-size:13px;padding:8px;border-radius:6px;")
            l.addWidget(bn)
            if s.get("cancelled_at"):
                l.addWidget(QLabel(f"<b>Cancelada em:</b> {fmt_date(s.get('cancelled_at',''))}",
                                   styleSheet="color:#FF1744;font-size:12px;"))
        l.addWidget(QLabel(f"<b>Data:</b> {fmt_date(s.get('date',''))}  "
                           f"<b>Cliente:</b> {s.get('customer_name','--')}  "
                           f"<b>Pagamento:</b> {s.get('payment_method','--')}"))
        t = QTableWidget(len(s.get("items",[])), 5)
        t.setHorizontalHeaderLabels(["Codigo","Produto","Qtd","Preco","Total"])
        t.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        t.setEditTriggers(QAbstractItemView.NoEditTriggers); t.verticalHeader().setVisible(False)
        for ri, it in enumerate(s.get("items",[])):
            rt = it["quantity"]*it["unit_price"]*(1-it.get("discount",0)/100)
            t.setItem(ri,0,QTableWidgetItem(it.get("code",""))); t.setItem(ri,1,QTableWidgetItem(it.get("name","")))
            t.setItem(ri,2,QTableWidgetItem(str(it["quantity"]))); t.setItem(ri,3,QTableWidgetItem(fmtR(it["unit_price"])))
            ti = QTableWidgetItem(fmtR(rt))
            ti.setForeground(QColor("#888888" if cancelled else "#00C853")); t.setItem(ri,4,ti)
        l.addWidget(t)
        sl = QHBoxLayout(); sl.addStretch()
        sl.addWidget(QLabel(f"Sub: {fmtR(s.get('subtotal',0))}  Desc: -{fmtR(s.get('discount_value',0))}  "
                            f"<b>TOTAL: {fmtR(s.get('total',0))}</b>"))
        l.addLayout(sl)
        if s.get("observations",""): l.addWidget(QLabel(f"<b>Obs:</b> {s['observations']}"))
        btns = QHBoxLayout()
        if not cancelled:
            bcl2 = QPushButton("  Cancelar esta Venda"); bcl2.setObjectName("btn_warning")
            bcl2.clicked.connect(lambda: (dlg.accept(), self._cancel_sale())); btns.addWidget(bcl2)
        bdl2 = QPushButton("  Excluir Venda"); bdl2.setObjectName("btn_danger")
        bdl2.clicked.connect(lambda: (dlg.accept(), self._delete_sale())); btns.addWidget(bdl2)
        btns.addStretch()
        bclose = QPushButton("Fechar"); bclose.setObjectName("btn_secondary")
        bclose.clicked.connect(dlg.accept); btns.addWidget(bclose)
        l.addLayout(btns)
        dlg.exec_()

# ═════════════════════════════════════════════════════════════════════════════
#  USERS TAB  (admin only)
# ═════════════════════════════════════════════════════════════════════════════
class UsersTab(QWidget):
    def __init__(self, dm, parent=None):
        super().__init__(parent); self.dm=dm; self._ids=[]; self._build(); self.refresh()
    def _build(self):
        lay = QVBoxLayout(self); lay.setContentsMargins(20,20,20,20); lay.setSpacing(12)
        lay.addWidget(SectionTitle("  Gerenciamento de Usuarios"))
        info=QFrame(); info.setObjectName("card"); info.setStyleSheet("QFrame#card{border-left:4px solid #1E88E5;}")
        il=QHBoxLayout(info); il.setContentsMargins(14,10,14,10)
        it=QLabel("<b>Controle de Acesso</b><br>"
            "  <b>Administrador</b>: acesso completo (todas as abas + configuracoes)<br>"
            "  <b>Operador</b>: somente Produtos, Nova Venda, Historico de Vendas e Clientes")
        it.setWordWrap(True); it.setObjectName("subtitle"); il.addWidget(it); lay.addWidget(info)
        tb=QHBoxLayout(); tb.addStretch()
        ba=QPushButton("  Novo Usuario"); ba.clicked.connect(self._add); tb.addWidget(ba)
        be=QPushButton("  Editar"); be.setObjectName("btn_secondary"); be.clicked.connect(self._edit); tb.addWidget(be)
        bt=QPushButton("  Ativar/Desativar"); bt.setObjectName("btn_warning"); bt.clicked.connect(self._toggle); tb.addWidget(bt)
        bd=QPushButton("  Excluir"); bd.setObjectName("btn_danger"); bd.clicked.connect(self._delete); tb.addWidget(bd)
        lay.addLayout(tb)
        self.tbl=QTableWidget(0,5); self.tbl.setHorizontalHeaderLabels(["Nome","Login","Perfil","Status","ID"])
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.tbl.horizontalHeader().setSectionResizeMode(0,QHeaderView.Stretch)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers); self.tbl.setAlternatingRowColors(True)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectRows); self.tbl.verticalHeader().setVisible(False)
        for c,w in enumerate([-1,130,120,100,220]):
            if w>0: self.tbl.setColumnWidth(c,w)
        lay.addWidget(self.tbl)
        self.cnt=QLabel(); self.cnt.setObjectName("subtitle"); lay.addWidget(self.cnt)
    def refresh(self):
        users=self.dm.get_users(); self.tbl.setRowCount(len(users)); self._ids=[]
        for r,u in enumerate(users):
            self._ids.append(u["id"])
            self.tbl.setItem(r,0,QTableWidgetItem(u.get("name","")))
            self.tbl.setItem(r,1,QTableWidgetItem(u.get("username","")))
            rl="  Admin" if u.get("role")=="admin" else "  Operador"
            ri=QTableWidgetItem(rl); ri.setForeground(QColor("#FFD600" if u.get("role")=="admin" else "#1E88E5")); self.tbl.setItem(r,2,ri)
            act=u.get("active",True); si=QTableWidgetItem("  Ativo" if act else "  Inativo")
            si.setForeground(QColor("#00C853" if act else "#FF1744")); self.tbl.setItem(r,3,si)
            self.tbl.setItem(r,4,QTableWidgetItem(u.get("id","")[:20]+"...")); self.tbl.setRowHeight(r,40)
        self.cnt.setText(f"{len(users)} usuario(s)")
    def _sid(self):
        r=self.tbl.currentRow(); return self._ids[r] if 0<=r<len(self._ids) else None
    def _protected(self,uid): return uid in ("admin-fixed","func-fixed")
    def _add(self):
        dlg=UserDialog(self)
        if dlg.exec_()==QDialog.Accepted:
            ok,msg=self.dm.add_user(dlg.result_data)
            QMessageBox.information(self,"  Sucesso",msg) if ok else QMessageBox.warning(self,"Erro",msg)
            self.refresh()
    def _edit(self):
        uid=self._sid()
        if not uid: QMessageBox.information(self,"Atencao","Selecione um usuario."); return
        u=next((x for x in self.dm.get_users() if x["id"]==uid),None)
        if not u: return
        dlg=UserDialog(self,u)
        if dlg.exec_()==QDialog.Accepted: dlg.result_data["id"]=uid; self.dm.update_user(uid,dlg.result_data); self.refresh()
    def _toggle(self):
        uid=self._sid()
        if not uid: QMessageBox.information(self,"Atencao","Selecione um usuario."); return
        if self._protected(uid): QMessageBox.warning(self,"Protegido","Nao e possivel desativar este usuario."); return
        self.dm.toggle_user_active(uid); self.refresh()
    def _delete(self):
        uid=self._sid()
        if not uid: QMessageBox.information(self,"Atencao","Selecione um usuario."); return
        if self._protected(uid): QMessageBox.warning(self,"Protegido","Nao e possivel excluir este usuario."); return
        u=next((x for x in self.dm.get_users() if x["id"]==uid),None)
        if u and QMessageBox.question(self,"Confirmar",f"Excluir usuario '{u['username']}'?",QMessageBox.Yes|QMessageBox.No)==QMessageBox.Yes:
            self.dm.delete_user(uid); self.refresh()


# ═════════════════════════════════════════════════════════════════════════════
#  MAIN WINDOW
# ═════════════════════════════════════════════════════════════════════════════
class MainWindow(QMainWindow):
    def __init__(self, dm, user):
        super().__init__(); self.dm=dm; self.user=user; self.is_admin=user.get("role")=="admin"
        s=dm.get_settings(); self._theme=s.get("theme","dark"); self._bg=s.get("background_image",""); self._accent=s.get("accent_color","#FF6B35")
        self.setWindowTitle(f"  Sistema Auto Pecas & Mecanica  --  {user.get('name','')}")
        self.setMinimumSize(1200,750); self.resize(1400,860); self._build(); self._apply_theme()
    def _build(self):
        central=QWidget(); self.setCentralWidget(central)
        ml=QHBoxLayout(central); ml.setContentsMargins(0,0,0,0); ml.setSpacing(0)
        self.sidebar=QFrame(); self.sidebar.setObjectName("sidebar"); self.sidebar.setFixedWidth(228)
        sl=QVBoxLayout(self.sidebar); sl.setContentsMargins(12,16,12,12); sl.setSpacing(4)
        lf=QFrame(); ll=QVBoxLayout(lf); ll.setContentsMargins(8,8,8,16)
        li=QLabel("  "); li.setFont(QFont("Segoe UI Emoji",28)); li.setAlignment(Qt.AlignCenter); ll.addWidget(li)
        lt=QLabel("AUTO PECAS"); lt.setAlignment(Qt.AlignCenter); lt.setStyleSheet("font-size:14px;font-weight:bold;letter-spacing:2px;"); ll.addWidget(lt)
        rl=QLabel("  Admin" if self.is_admin else "  Operador"); rl.setAlignment(Qt.AlignCenter)
        rl.setStyleSheet(f"font-size:11px;color:{'#FFD600' if self.is_admin else '#1E88E5'};"); ll.addWidget(rl)
        sep=QFrame(); sep.setFrameShape(QFrame.HLine); ll.addWidget(sep); sl.addWidget(lf)
        self.content=QStackedWidget()
        self.t_dash=DashboardTab(self.dm); self.t_prod=ProductsTab(self.dm)
        self.t_sale=SalesTab(self.dm); self.t_sale.sale_completed.connect(self._on_sale)
        self.t_hist=HistoryTab(self.dm, self.user); self.t_cust=CustomersTab(self.dm)
        self.t_rep=ReportsTab(self.dm); self.t_usr=UsersTab(self.dm)
        all_tabs=[
            ("  Dashboard",   self.t_dash, True),
            ("  Produtos",     self.t_prod, False),
            ("  Nova Venda",   self.t_sale, False),
            ("  Historico",    self.t_hist, False),
            ("  Clientes",     self.t_cust, False),
            ("  Relatorios",   self.t_rep,  True),
            ("  Usuarios",     self.t_usr,  True),
        ]
        self._nav=[]
        for label,widget,admin_only in all_tabs:
            idx=self.content.count(); self.content.addWidget(widget)
            if not admin_only or self.is_admin: self._nav.append((label,idx,widget))
        self._btns=[]
        for label,idx,_ in self._nav:
            btn=QPushButton(label); btn.setObjectName("nav_btn"); btn.setMinimumHeight(42)
            btn.clicked.connect(lambda _,i=idx: self._go(i)); sl.addWidget(btn); self._btns.append((btn,idx))
        sl.addStretch()
        sep2=QFrame(); sep2.setFrameShape(QFrame.HLine); sl.addWidget(sep2)
        ol=QLabel("OPCOES"); ol.setObjectName("metric_label"); ol.setContentsMargins(8,4,0,4); sl.addWidget(ol)
        if self.is_admin:
            bs=QPushButton("  Configuracoes"); bs.setObjectName("nav_btn"); bs.setMinimumHeight(42); bs.clicked.connect(self._settings); sl.addWidget(bs)
            bb=QPushButton("  Backup"); bb.setObjectName("nav_btn"); bb.setMinimumHeight(42); bb.clicked.connect(self._backup); sl.addWidget(bb)
        btc=QPushButton("  Trocar Conta"); btc.setObjectName("nav_btn"); btc.setMinimumHeight(42); btc.clicked.connect(self._switch_account); sl.addWidget(btc)
        bo=QPushButton("  Sair"); bo.setObjectName("nav_btn"); bo.setMinimumHeight(42); bo.clicked.connect(self._logout); sl.addWidget(bo)
        ml.addWidget(self.sidebar); ml.addWidget(self.content,1)
        sb=QStatusBar(); self.setStatusBar(sb)
        role_s="Administrador" if self.is_admin else "Operador"
        sb.showMessage(f"  {self.user.get('name','')} ({role_s})  |    {DATA_FILE}  |    {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        if self._nav: self._go(self._nav[0][1])
    def _go(self, idx):
        self.content.setCurrentIndex(idx)
        for btn,i in self._btns:
            btn.setObjectName("nav_btn_active" if i==idx else "nav_btn")
            btn.style().unpolish(btn); btn.style().polish(btn)
        tab_refresh={0:self.t_dash,1:self.t_prod,3:self.t_hist,4:self.t_cust,6:self.t_usr}
        t=tab_refresh.get(idx)
        if t and hasattr(t,"refresh"): t.refresh()
    def _on_sale(self):
        self.t_dash.refresh(); self.t_hist.refresh(); self.t_prod.refresh(); self.t_sale.refresh_customers()
    def _settings(self):
        dlg=SettingsDialog(self,self.dm)
        if dlg.exec_()==QDialog.Accepted:
            s=self.dm.get_settings(); self._theme=s.get("theme","dark"); self._bg=s.get("background_image",""); self._accent=s.get("accent_color","#FF6B35")
            self._apply_theme(); QMessageBox.information(self,"  Aplicado","Configuracoes salvas!")
    def _apply_theme(self):
        QApplication.instance().setStyleSheet(build_stylesheet(self._theme,self._accent))
        if self._bg and os.path.exists(self._bg):
            pix=QPixmap(self._bg); pal=self.palette()
            scaled=pix.scaled(self.size(),Qt.KeepAspectRatioByExpanding,Qt.SmoothTransformation)
            pal.setBrush(QPalette.Window,QBrush(scaled)); self.setPalette(pal); self.setAutoFillBackground(True)
        else: self.setPalette(QApplication.instance().palette()); self.setAutoFillBackground(False)
    def resizeEvent(self,e):
        super().resizeEvent(e)
        if self._bg and os.path.exists(self._bg): self._apply_theme()
    def _backup(self):
        p,_=QFileDialog.getSaveFileName(self,"Backup",f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json","JSON (*.json)")
        if p:
            import shutil
            try: shutil.copy2(DATA_FILE,p); QMessageBox.information(self,"  Backup OK",f"Salvo em:\n{p}")
            except Exception as e: QMessageBox.critical(self,"Erro",str(e))
    def _switch_account(self):
        """Close current session and show login without closing the app."""
        reply = QMessageBox.question(
            self, "Trocar Conta",
            f"Deseja trocar de conta?\n\nUsuario atual: {self.user.get('name', '')}",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            self.close()
            win = LoginWindow()
            win.show()
            scr = QApplication.instance().primaryScreen().geometry()
            win.move((scr.width() - win.width()) // 2, (scr.height() - win.height()) // 2)

    def _logout(self):
        if QMessageBox.question(self,"Sair","Deseja sair do sistema?",QMessageBox.Yes|QMessageBox.No)==QMessageBox.Yes:
            self.close(); w=LoginWindow(); w.show()
            scr=QApplication.instance().primaryScreen().geometry(); w.move((scr.width()-w.width())//2,(scr.height()-w.height())//2)


# ═════════════════════════════════════════════════════════════════════════════
#  LOGIN WINDOW
# ═════════════════════════════════════════════════════════════════════════════
class LoginWindow(QWidget):
    def __init__(self):
        super().__init__(); self.dm=DataManager()
        self.setWindowTitle("Login"); self.setFixedSize(420,560)
        self.setWindowFlags(Qt.FramelessWindowHint); self._dp=None; self._mw=None
        QApplication.instance().setStyleSheet(build_stylesheet("dark","#FF6B35"))
        lay=QVBoxLayout(self); lay.setContentsMargins(0,0,0,0)
        bg=QFrame(); bg.setStyleSheet("QFrame{background:qlineargradient(x1:0,y1:0,x2:1,y2:1,stop:0 #1A1A1A,stop:1 #2A1A0A);border-radius:14px;}")
        bl=QVBoxLayout(bg); bl.setContentsMargins(40,28,40,28); bl.setSpacing(0)
        tb=QHBoxLayout(); tb.addWidget(QLabel("  ",styleSheet="color:#555;font-size:11px;")); tb.addStretch()
        cl=QPushButton("x"); cl.setFixedSize(28,28); cl.setStyleSheet("background:transparent;color:#888;font-size:14px;border:none;"); cl.clicked.connect(QApplication.quit)
        tb.addWidget(cl); bl.addLayout(tb); bl.addSpacing(6)
        li=QLabel("  "); li.setFont(QFont("Segoe UI Emoji",44)); li.setAlignment(Qt.AlignCenter); bl.addWidget(li)
        lt=QLabel("AUTO PECAS"); lt.setAlignment(Qt.AlignCenter); lt.setStyleSheet("font-size:22px;font-weight:bold;color:#FF6B35;letter-spacing:4px;"); bl.addWidget(lt)
        ls=QLabel("Sistema de Gestao e Vendas"); ls.setAlignment(Qt.AlignCenter); ls.setStyleSheet("font-size:12px;color:#888;margin-bottom:20px;"); bl.addWidget(ls)
        bl.addSpacing(16)
        ff=QFrame(); ff.setStyleSheet("QFrame{background:rgba(255,255,255,0.05);border:1px solid rgba(255,255,255,0.1);border-radius:10px;}")
        fl=QVBoxLayout(ff); fl.setContentsMargins(20,20,20,20); fl.setSpacing(14)
        ist="QLineEdit{background:rgba(255,255,255,0.08);border:1px solid rgba(255,255,255,0.15);border-radius:8px;color:white;padding:8px 14px;font-size:14px;}QLineEdit:focus{border:2px solid #FF6B35;}"
        lbs="font-size:11px;color:#888;font-weight:bold;letter-spacing:1px;"
        fl.addWidget(QLabel("  USUARIO",styleSheet=lbs))
        self.ue=QLineEdit(); self.ue.setPlaceholderText("Login de acesso"); self.ue.setMinimumHeight(42); self.ue.setStyleSheet(ist)
        fl.addWidget(self.ue)
        fl.addWidget(QLabel("  SENHA",styleSheet=lbs))
        self.pe=QLineEdit(); self.pe.setEchoMode(QLineEdit.Password); self.pe.setPlaceholderText("Senha"); self.pe.setMinimumHeight(42); self.pe.setStyleSheet(ist)
        self.pe.returnPressed.connect(self._login); fl.addWidget(self.pe)
        bl.addWidget(ff); bl.addSpacing(12)
        self.el=QLabel(""); self.el.setAlignment(Qt.AlignCenter); self.el.setStyleSheet("color:#FF1744;font-size:12px;"); bl.addWidget(self.el)
        bl.addSpacing(4)
        btn=QPushButton("ENTRAR  >"); btn.setMinimumHeight(48)
        btn.setStyleSheet("QPushButton{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #FF6B35,stop:1 #FF8C5A);color:white;border:none;border-radius:10px;font-size:15px;font-weight:bold;letter-spacing:2px;}QPushButton:hover{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #FF8C5A,stop:1 #FFAA80);}QPushButton:pressed{background:#CC4F1F;}")
        btn.clicked.connect(self._login); bl.addWidget(btn); bl.addStretch()
        vl=QLabel("v2.0  |  Funcionario: funcionario / 123moto")
        vl.setAlignment(Qt.AlignCenter); vl.setStyleSheet("font-size:10px;color:#444;margin-top:8px;"); bl.addWidget(vl)
        lay.addWidget(bg)
    def _login(self):
        u=self.ue.text().strip(); p=self.pe.text()
        if not u or not p: self.el.setText("  Preencha o login e a senha."); return
        user=self.dm.authenticate(u,p)
        if user:
            self.el.setText(""); self._mw=MainWindow(self.dm,user); self._mw.show()
            scr=QApplication.instance().primaryScreen().geometry()
            self._mw.move((scr.width()-self._mw.width())//2,(scr.height()-self._mw.height())//2); self.close()
        else: self.el.setText("  Usuario ou senha incorretos!"); self.pe.clear(); self.pe.setFocus()
    def mousePressEvent(self,e):
        if e.button()==Qt.LeftButton: self._dp=e.globalPos()-self.pos(); e.accept()
    def mouseMoveEvent(self,e):
        if e.buttons()==Qt.LeftButton and self._dp: self.move(e.globalPos()-self._dp); e.accept()


# ═════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═════════════════════════════════════════════════════════════════════════════
def main():
    app=QApplication(sys.argv); app.setApplicationName("AutoPecas Sistema"); app.setApplicationVersion("2.0.0")
    try: app.setAttribute(Qt.AA_EnableHighDpiScaling,True); app.setAttribute(Qt.AA_UseHighDpiPixmaps,True)
    except: pass
    win=LoginWindow(); win.show()
    scr=app.primaryScreen().geometry(); win.move((scr.width()-win.width())//2,(scr.height()-win.height())//2)
    sys.exit(app.exec_())

if __name__=="__main__":
    main()
