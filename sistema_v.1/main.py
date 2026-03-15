#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════╗
║       SISTEMA DE VENDAS - AUTO PEÇAS & MECÂNICA          ║
║                      Versão 1.0.0                        ║
╚══════════════════════════════════════════════════════════╝
"""

import sys
import os
import json
import uuid
from datetime import datetime, timedelta
from collections import defaultdict

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGridLayout, QLabel, QLineEdit, QPushButton, QTableWidget,
    QTableWidgetItem, QTabWidget, QDialog, QFormLayout, QComboBox,
    QSpinBox, QDoubleSpinBox, QTextEdit, QMessageBox, QFileDialog,
    QHeaderView, QFrame, QScrollArea, QSizePolicy, QDateEdit,
    QGroupBox, QButtonGroup, QRadioButton, QSplitter, QListWidget,
    QListWidgetItem, QAbstractItemView, QMenu, QAction, QToolBar,
    QStatusBar, QStackedWidget, QCheckBox, QSlider
)
from PyQt5.QtCore import Qt, QDate, QSize, QTimer, pyqtSignal, QObject, QPoint
from PyQt5.QtGui import (
    QFont, QColor, QPixmap, QPainter, QBrush, QLinearGradient,
    QPen, QIcon, QPalette, QImage, QCursor
)

try:
    import openpyxl
    from openpyxl.styles import Font as XLFont, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# ═══════════════════════════════════════════════════════════════════
#  CAMINHOS
# ═══════════════════════════════════════════════════════════════════

if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DATA_FILE = os.path.join(BASE_DIR, "autopecas_data.json")

# ═══════════════════════════════════════════════════════════════════
#  DATA MANAGER
# ═══════════════════════════════════════════════════════════════════

class DataManager:
    def __init__(self):
        self.data = {}
        self.load()

    def load(self):
        if os.path.exists(DATA_FILE):
            try:
                with open(DATA_FILE, 'r', encoding='utf-8') as f:
                    self.data = json.load(f)
                for k, v in self._defaults().items():
                    if k not in self.data:
                        self.data[k] = v
            except Exception:
                self.data = self._defaults()
                self.data["products"] = self._sample_products()
                self.data["customers"] = self._sample_customers()
        else:
            self.data = self._defaults()
            self.data["products"] = self._sample_products()
            self.data["customers"] = self._sample_customers()
            self.save()

    def _defaults(self):
        return {
            "products": [], "customers": [], "sales": [],
            "settings": {"theme": "dark", "background_image": ""}
        }

    def save(self):
        with open(DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)

    def _sample_products(self):
        cats = [
            ("FLT001","Filtro de Óleo","Filtros","Bosch","UN",15,35,50,10,"Filtro de óleo universal"),
            ("FLT002","Filtro de Ar","Filtros","Mann","UN",20,45,30,8,"Filtro de ar elemento seco"),
            ("FLT003","Filtro de Combustível","Filtros","WIX","UN",18,40,25,5,"Filtro de combustível injetado"),
            ("FRE001","Pastilha de Freio Dianteira","Freios","Bosch","JG",45,95,20,5,"Par pastilha dianteira"),
            ("FRE002","Disco de Freio Dianteiro","Freios","Brembo","UN",80,165,15,4,"Disco freio ventilado"),
            ("FRE003","Fluido de Freio DOT4","Freios","ATE","FR",12,28,40,10,"Fluido de freio 500ml"),
            ("OLE001","Óleo 5W30 Sintético","Óleos","Castrol","LT",28,55,100,20,"Óleo motor sintético"),
            ("OLE002","Óleo 10W40 Semissintético","Óleos","Mobil","LT",22,45,80,15,"Óleo motor semissintético"),
            ("OLE003","Óleo de Câmbio","Óleos","Valvoline","LT",30,60,30,8,"Óleo câmbio automático"),
            ("VEL001","Vela de Ignição","Motor","NGK","UN",12,25,5,10,"Vela ignição iridium"),
            ("BAT001","Bateria 60Ah","Elétrica","Moura","UN",280,450,10,3,"Bateria 60Ah selada"),
            ("ALT001","Correia Alternador","Motor","Dayco","UN",35,72,18,5,"Correia poly-v alternador"),
            ("SUS001","Amortecedor Dianteiro","Suspensão","Monroe","UN",120,220,8,2,"Amortecedor a gás dianteiro"),
            ("SUS002","Barra Estabilizadora","Suspensão","TRW","UN",65,130,6,2,"Buchas barra estabilizadora"),
            ("RAD001","Radiador","Arrefecimento","Valeo","UN",350,620,4,1,"Radiador alumínio completo"),
        ]
        prods = []
        for c in cats:
            prods.append({
                "id": str(uuid.uuid4()),
                "code": c[0], "name": c[1], "category": c[2],
                "brand": c[3], "unit": c[4], "cost_price": float(c[5]),
                "sale_price": float(c[6]), "stock": c[7], "min_stock": c[8],
                "description": c[9]
            })
        return prods

    def _sample_customers(self):
        return [
            {"id": str(uuid.uuid4()), "name": "João Carlos Silva",
             "cpf_cnpj": "123.456.789-00", "phone": "(21) 98765-4321",
             "email": "joao@email.com", "address": "Rua das Flores, 100 - Centro"},
            {"id": str(uuid.uuid4()), "name": "Maria Santos Oliveira",
             "cpf_cnpj": "987.654.321-00", "phone": "(21) 91234-5678",
             "email": "maria@email.com", "address": "Av. Brasil, 200 - Barra"},
            {"id": str(uuid.uuid4()), "name": "Auto Center Mega Ltda",
             "cpf_cnpj": "12.345.678/0001-90", "phone": "(21) 3333-4444",
             "email": "contato@megaauto.com", "address": "Rod. Presidente Dutra, 300"},
            {"id": str(uuid.uuid4()), "name": "Pedro Melo Mecânica",
             "cpf_cnpj": "45.678.901/0001-23", "phone": "(21) 97777-8888",
             "email": "pedro@mecanica.com", "address": "Rua Industrial, 55 - Nova Iguaçu"},
        ]

    # --- PRODUCTS ---
    def get_products(self): return self.data["products"]
    def add_product(self, p):
        p["id"] = str(uuid.uuid4()); self.data["products"].append(p); self.save()
    def update_product(self, pid, upd):
        for i, p in enumerate(self.data["products"]):
            if p["id"] == pid: self.data["products"][i] = upd; break
        self.save()
    def delete_product(self, pid):
        self.data["products"] = [p for p in self.data["products"] if p["id"] != pid]; self.save()
    def get_product_by_id(self, pid):
        for p in self.data["products"]:
            if p["id"] == pid: return p
        return None
    def get_product_by_code(self, code):
        for p in self.data["products"]:
            if p["code"].lower() == code.lower(): return p
        return None

    # --- CUSTOMERS ---
    def get_customers(self): return self.data["customers"]
    def add_customer(self, c):
        c["id"] = str(uuid.uuid4()); self.data["customers"].append(c); self.save()
    def update_customer(self, cid, upd):
        for i, c in enumerate(self.data["customers"]):
            if c["id"] == cid: self.data["customers"][i] = upd; break
        self.save()
    def delete_customer(self, cid):
        self.data["customers"] = [c for c in self.data["customers"] if c["id"] != cid]; self.save()
    def get_customer_by_id(self, cid):
        for c in self.data["customers"]:
            if c["id"] == cid: return c
        return None

    # --- SALES ---
    def get_sales(self): return self.data["sales"]
    def add_sale(self, sale):
        sale["id"] = str(uuid.uuid4())
        sale["date"] = datetime.now().isoformat()
        for item in sale.get("items", []):
            for p in self.data["products"]:
                if p["id"] == item["product_id"]:
                    p["stock"] = max(0, p["stock"] - item["quantity"]); break
        self.data["sales"].append(sale); self.save()
        return sale["id"]

    # --- SETTINGS ---
    def get_settings(self):
        return self.data.get("settings", {"theme": "dark", "background_image": ""})
    def save_settings(self, s):
        self.data["settings"] = s; self.save()


# ═══════════════════════════════════════════════════════════════════
#  THEME MANAGER
# ═══════════════════════════════════════════════════════════════════

DARK_COLORS = {
    "bg": "#1A1A1A", "surface": "#252525", "surface2": "#2E2E2E",
    "border": "#3A3A3A", "text": "#EFEFEF", "text2": "#AAAAAA",
    "accent": "#FF6B35", "accent_hover": "#FF8C5A", "accent_dark": "#CC4F1F",
    "success": "#00C853", "warning": "#FFD600", "danger": "#FF1744",
    "table_alt": "#2A2A2A", "input_bg": "#333333", "sidebar": "#1E1E1E",
    "card": "#262626", "header": "#1F1F1F",
}
LIGHT_COLORS = {
    "bg": "#F0F2F5", "surface": "#FFFFFF", "surface2": "#F8F9FA",
    "border": "#DADCE0", "text": "#1A1A1A", "text2": "#666666",
    "accent": "#E64A19", "accent_hover": "#FF5722", "accent_dark": "#BF360C",
    "success": "#43A047", "warning": "#FB8C00", "danger": "#E53935",
    "table_alt": "#F5F5F5", "input_bg": "#FFFFFF", "sidebar": "#FFFFFF",
    "card": "#FFFFFF", "header": "#FFFFFF",
}

def build_stylesheet(theme="dark"):
    c = DARK_COLORS if theme == "dark" else LIGHT_COLORS
    return f"""
QMainWindow, QDialog {{
    background-color: {c['bg']};
    color: {c['text']};
}}
QWidget {{
    background-color: {c['bg']};
    color: {c['text']};
    font-family: 'Segoe UI', Arial, sans-serif;
    font-size: 13px;
}}
QFrame#card {{
    background-color: {c['card']};
    border: 1px solid {c['border']};
    border-radius: 8px;
}}
QFrame#sidebar {{
    background-color: {c['sidebar']};
    border-right: 1px solid {c['border']};
}}
QLabel#title {{
    font-size: 22px;
    font-weight: bold;
    color: {c['text']};
}}
QLabel#subtitle {{
    font-size: 13px;
    color: {c['text2']};
}}
QLabel#metric_value {{
    font-size: 26px;
    font-weight: bold;
    color: {c['accent']};
}}
QLabel#metric_label {{
    font-size: 11px;
    color: {c['text2']};
    text-transform: uppercase;
}}
QLabel#section_title {{
    font-size: 15px;
    font-weight: bold;
    color: {c['text']};
    border-bottom: 2px solid {c['accent']};
    padding-bottom: 4px;
}}
QPushButton {{
    background-color: {c['accent']};
    color: #FFFFFF;
    border: none;
    border-radius: 6px;
    padding: 8px 18px;
    font-size: 13px;
    font-weight: bold;
}}
QPushButton:hover {{
    background-color: {c['accent_hover']};
}}
QPushButton:pressed {{
    background-color: {c['accent_dark']};
}}
QPushButton:disabled {{
    background-color: {c['border']};
    color: {c['text2']};
}}
QPushButton#btn_secondary {{
    background-color: {c['surface2']};
    color: {c['text']};
    border: 1px solid {c['border']};
}}
QPushButton#btn_secondary:hover {{
    background-color: {c['border']};
}}
QPushButton#btn_danger {{
    background-color: {c['danger']};
    color: #FFFFFF;
}}
QPushButton#btn_danger:hover {{
    background-color: #FF4569;
}}
QPushButton#btn_success {{
    background-color: {c['success']};
    color: #FFFFFF;
}}
QPushButton#btn_success:hover {{
    background-color: #33D66B;
}}
QPushButton#btn_warning {{
    background-color: {c['warning']};
    color: #1A1A1A;
}}
QPushButton#nav_btn {{
    background-color: transparent;
    color: {c['text2']};
    border: none;
    border-radius: 8px;
    padding: 10px 16px;
    text-align: left;
    font-size: 13px;
    font-weight: normal;
}}
QPushButton#nav_btn:hover {{
    background-color: {c['surface2']};
    color: {c['text']};
}}
QPushButton#nav_btn_active {{
    background-color: {c['accent']};
    color: #FFFFFF;
    border: none;
    border-radius: 8px;
    padding: 10px 16px;
    text-align: left;
    font-size: 13px;
    font-weight: bold;
}}
QLineEdit, QTextEdit, QComboBox, QSpinBox, QDoubleSpinBox, QDateEdit {{
    background-color: {c['input_bg']};
    color: {c['text']};
    border: 1px solid {c['border']};
    border-radius: 6px;
    padding: 6px 10px;
    font-size: 13px;
}}
QLineEdit:focus, QTextEdit:focus, QComboBox:focus,
QSpinBox:focus, QDoubleSpinBox:focus, QDateEdit:focus {{
    border: 2px solid {c['accent']};
    outline: none;
}}
QComboBox::drop-down {{
    border: none;
    width: 28px;
}}
QComboBox::down-arrow {{
    image: none;
    border-left: 5px solid transparent;
    border-right: 5px solid transparent;
    border-top: 6px solid {c['text2']};
    margin-right: 8px;
}}
QComboBox QAbstractItemView {{
    background-color: {c['surface']};
    color: {c['text']};
    border: 1px solid {c['border']};
    selection-background-color: {c['accent']};
}}
QTableWidget {{
    background-color: {c['surface']};
    color: {c['text']};
    border: 1px solid {c['border']};
    border-radius: 6px;
    gridline-color: {c['border']};
    alternate-background-color: {c['table_alt']};
}}
QTableWidget::item {{
    padding: 6px 10px;
}}
QTableWidget::item:selected {{
    background-color: {c['accent']};
    color: #FFFFFF;
}}
QHeaderView::section {{
    background-color: {c['surface2']};
    color: {c['text']};
    padding: 8px 10px;
    border: none;
    border-right: 1px solid {c['border']};
    border-bottom: 2px solid {c['accent']};
    font-weight: bold;
    font-size: 12px;
}}
QTabWidget::pane {{
    border: 1px solid {c['border']};
    border-radius: 6px;
    background-color: {c['surface']};
}}
QTabBar::tab {{
    background-color: {c['surface2']};
    color: {c['text2']};
    padding: 8px 18px;
    border-top-left-radius: 6px;
    border-top-right-radius: 6px;
    margin-right: 2px;
    font-size: 13px;
}}
QTabBar::tab:selected {{
    background-color: {c['accent']};
    color: #FFFFFF;
    font-weight: bold;
}}
QTabBar::tab:hover:!selected {{
    background-color: {c['border']};
    color: {c['text']};
}}
QGroupBox {{
    border: 1px solid {c['border']};
    border-radius: 8px;
    margin-top: 12px;
    padding: 12px 8px 8px 8px;
    font-weight: bold;
    color: {c['text']};
}}
QGroupBox::title {{
    subcontrol-origin: margin;
    subcontrol-position: top left;
    padding: 0 8px;
    color: {c['accent']};
    font-size: 13px;
}}
QScrollBar:vertical {{
    background: {c['surface2']};
    width: 8px;
    border-radius: 4px;
}}
QScrollBar::handle:vertical {{
    background: {c['border']};
    border-radius: 4px;
    min-height: 24px;
}}
QScrollBar::handle:vertical:hover {{
    background: {c['accent']};
}}
QScrollBar:horizontal {{
    background: {c['surface2']};
    height: 8px;
    border-radius: 4px;
}}
QScrollBar::handle:horizontal {{
    background: {c['border']};
    border-radius: 4px;
    min-width: 24px;
}}
QScrollBar::add-line, QScrollBar::sub-line {{
    width: 0; height: 0;
}}
QMessageBox {{
    background-color: {c['surface']};
    color: {c['text']};
}}
QStatusBar {{
    background-color: {c['header']};
    color: {c['text2']};
    border-top: 1px solid {c['border']};
    font-size: 12px;
    padding: 2px 8px;
}}
QRadioButton {{
    color: {c['text']};
    spacing: 8px;
}}
QRadioButton::indicator {{
    width: 16px; height: 16px;
    border-radius: 8px;
    border: 2px solid {c['border']};
    background: {c['input_bg']};
}}
QRadioButton::indicator:checked {{
    background: {c['accent']};
    border-color: {c['accent']};
}}
QCheckBox {{
    color: {c['text']};
    spacing: 8px;
}}
QCheckBox::indicator {{
    width: 16px; height: 16px;
    border-radius: 3px;
    border: 2px solid {c['border']};
    background: {c['input_bg']};
}}
QCheckBox::indicator:checked {{
    background: {c['accent']};
    border-color: {c['accent']};
}}
QSplitter::handle {{
    background: {c['border']};
}}
QToolTip {{
    background-color: {c['surface']};
    color: {c['text']};
    border: 1px solid {c['border']};
    padding: 4px 8px;
    border-radius: 4px;
}}
"""


# ═══════════════════════════════════════════════════════════════════
#  REUSABLE WIDGETS
# ═══════════════════════════════════════════════════════════════════

class MetricCard(QFrame):
    def __init__(self, title, value, icon="📦", color=None, parent=None):
        super().__init__(parent)
        self.setObjectName("card")
        self.setMinimumHeight(110)
        self.setMinimumWidth(180)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 14, 16, 14)
        layout.setSpacing(6)

        top = QHBoxLayout()
        icon_lbl = QLabel(icon)
        icon_lbl.setFont(QFont("Segoe UI Emoji", 22))
        self.val_lbl = QLabel(str(value))
        self.val_lbl.setObjectName("metric_value")
        if color:
            self.val_lbl.setStyleSheet(f"color: {color}; font-size: 26px; font-weight: bold;")
        top.addWidget(icon_lbl)
        top.addStretch()
        top.addWidget(self.val_lbl)
        layout.addLayout(top)

        self.title_lbl = QLabel(title)
        self.title_lbl.setObjectName("metric_label")
        layout.addWidget(self.title_lbl)

    def set_value(self, v):
        self.val_lbl.setText(str(v))


class SearchBar(QWidget):
    search_changed = pyqtSignal(str)

    def __init__(self, placeholder="Buscar...", parent=None):
        super().__init__(parent)
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        self.edit = QLineEdit()
        self.edit.setPlaceholderText(f"🔍  {placeholder}")
        self.edit.textChanged.connect(self.search_changed)
        layout.addWidget(self.edit)


class SectionTitle(QLabel):
    def __init__(self, text, parent=None):
        super().__init__(text, parent)
        self.setObjectName("section_title")
        font = QFont()
        font.setPointSize(13)
        font.setBold(True)
        self.setFont(font)


class StatusBadge(QLabel):
    COLORS = {
        "ativo": ("#00C853", "#003300"),
        "baixo": ("#FFD600", "#332b00"),
        "esgotado": ("#FF1744", "#330008"),
        "concluída": ("#00C853", "#003300"),
        "cancelada": ("#FF1744", "#330008"),
        "pendente": ("#FFD600", "#332b00"),
    }

    def __init__(self, status, parent=None):
        super().__init__(parent)
        self.set_status(status)

    def set_status(self, status):
        self.setText(f" {status.capitalize()} ")
        key = status.lower()
        fg, bg = self.COLORS.get(key, ("#AAAAAA", "#222222"))
        self.setStyleSheet(
            f"background:{bg}; color:{fg}; border:1px solid {fg}; "
            f"border-radius:10px; padding:2px 8px; font-size:11px; font-weight:bold;"
        )


# ═══════════════════════════════════════════════════════════════════
#  DIALOGS
# ═══════════════════════════════════════════════════════════════════

class ProductDialog(QDialog):
    CATEGORIES = ["Filtros","Freios","Óleos","Motor","Elétrica","Suspensão",
                  "Arrefecimento","Transmissão","Carroceria","Acessórios","Outros"]
    UNITS = ["UN","JG","LT","KG","MT","CX","PC","FR","KIT","PAR"]

    def __init__(self, parent=None, product=None):
        super().__init__(parent)
        self.product = product
        self.setWindowTitle("Novo Produto" if not product else "Editar Produto")
        self.setMinimumWidth(480)
        self.setModal(True)
        self._build_ui()
        if product:
            self._populate(product)

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(20, 20, 20, 20)

        title = SectionTitle("Novo Produto" if not self.product else "Editar Produto")
        layout.addWidget(title)
        layout.addSpacing(8)

        form = QGridLayout()
        form.setSpacing(10)

        def row(label, widget, r, c=0, cs=1):
            form.addWidget(QLabel(label), r, c)
            form.addWidget(widget, r, c+1, 1, cs)

        self.code_e = QLineEdit(); self.code_e.setPlaceholderText("Ex: FLT001")
        self.name_e = QLineEdit(); self.name_e.setPlaceholderText("Nome do produto")
        self.cat_e = QComboBox(); self.cat_e.addItems(self.CATEGORIES)
        self.brand_e = QLineEdit(); self.brand_e.setPlaceholderText("Fabricante")
        self.unit_e = QComboBox(); self.unit_e.addItems(self.UNITS)
        self.cost_e = QDoubleSpinBox()
        self.cost_e.setRange(0, 999999); self.cost_e.setPrefix("R$ "); self.cost_e.setDecimals(2)
        self.price_e = QDoubleSpinBox()
        self.price_e.setRange(0, 999999); self.price_e.setPrefix("R$ "); self.price_e.setDecimals(2)
        self.stock_e = QSpinBox(); self.stock_e.setRange(0, 99999)
        self.min_e = QSpinBox(); self.min_e.setRange(0, 99999)
        self.desc_e = QTextEdit(); self.desc_e.setMaximumHeight(70)
        self.desc_e.setPlaceholderText("Descrição opcional...")

        row("Código *", self.code_e, 0)
        row("Nome *", self.name_e, 1)
        row("Categoria", self.cat_e, 2)
        row("Marca/Fabricante", self.brand_e, 3)
        row("Unidade", self.unit_e, 4)
        row("Preço de Custo", self.cost_e, 5)
        row("Preço de Venda *", self.price_e, 6)
        row("Estoque Atual", self.stock_e, 7)
        row("Estoque Mínimo", self.min_e, 8)
        form.addWidget(QLabel("Descrição"), 9, 0)
        form.addWidget(self.desc_e, 9, 1)

        form.setColumnStretch(1, 1)
        layout.addLayout(form)
        layout.addSpacing(10)

        btns = QHBoxLayout()
        btns.addStretch()
        cancel = QPushButton("Cancelar")
        cancel.setObjectName("btn_secondary")
        cancel.clicked.connect(self.reject)
        save = QPushButton("💾  Salvar Produto")
        save.clicked.connect(self._save)
        btns.addWidget(cancel)
        btns.addWidget(save)
        layout.addLayout(btns)

    def _populate(self, p):
        self.code_e.setText(p.get("code",""))
        self.name_e.setText(p.get("name",""))
        idx = self.cat_e.findText(p.get("category",""))
        if idx >= 0: self.cat_e.setCurrentIndex(idx)
        self.brand_e.setText(p.get("brand",""))
        idx = self.unit_e.findText(p.get("unit","UN"))
        if idx >= 0: self.unit_e.setCurrentIndex(idx)
        self.cost_e.setValue(float(p.get("cost_price",0)))
        self.price_e.setValue(float(p.get("sale_price",0)))
        self.stock_e.setValue(int(p.get("stock",0)))
        self.min_e.setValue(int(p.get("min_stock",0)))
        self.desc_e.setPlainText(p.get("description",""))

    def _save(self):
        if not self.code_e.text().strip():
            QMessageBox.warning(self, "Atenção", "Informe o código do produto.")
            return
        if not self.name_e.text().strip():
            QMessageBox.warning(self, "Atenção", "Informe o nome do produto.")
            return
        if self.price_e.value() <= 0:
            QMessageBox.warning(self, "Atenção", "Informe o preço de venda.")
            return
        self.result_data = {
            "id": self.product["id"] if self.product else "",
            "code": self.code_e.text().strip().upper(),
            "name": self.name_e.text().strip(),
            "category": self.cat_e.currentText(),
            "brand": self.brand_e.text().strip(),
            "unit": self.unit_e.currentText(),
            "cost_price": self.cost_e.value(),
            "sale_price": self.price_e.value(),
            "stock": self.stock_e.value(),
            "min_stock": self.min_e.value(),
            "description": self.desc_e.toPlainText().strip()
        }
        self.accept()


class CustomerDialog(QDialog):
    def __init__(self, parent=None, customer=None):
        super().__init__(parent)
        self.customer = customer
        self.setWindowTitle("Novo Cliente" if not customer else "Editar Cliente")
        self.setMinimumWidth(440)
        self.setModal(True)
        self._build_ui()
        if customer:
            self._populate(customer)

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(20, 20, 20, 20)

        title = SectionTitle("Novo Cliente" if not self.customer else "Editar Cliente")
        layout.addWidget(title)
        layout.addSpacing(8)

        form = QGridLayout()
        form.setSpacing(10)

        self.name_e = QLineEdit(); self.name_e.setPlaceholderText("Nome completo ou razão social")
        self.doc_e = QLineEdit(); self.doc_e.setPlaceholderText("CPF ou CNPJ")
        self.phone_e = QLineEdit(); self.phone_e.setPlaceholderText("(DD) XXXXX-XXXX")
        self.email_e = QLineEdit(); self.email_e.setPlaceholderText("email@exemplo.com")
        self.addr_e = QLineEdit(); self.addr_e.setPlaceholderText("Endereço completo")

        fields = [
            ("Nome / Razão Social *", self.name_e),
            ("CPF / CNPJ", self.doc_e),
            ("Telefone / WhatsApp", self.phone_e),
            ("E-mail", self.email_e),
            ("Endereço", self.addr_e),
        ]
        for r, (lbl, w) in enumerate(fields):
            form.addWidget(QLabel(lbl), r, 0)
            form.addWidget(w, r, 1)
        form.setColumnStretch(1, 1)
        layout.addLayout(form)
        layout.addSpacing(10)

        btns = QHBoxLayout()
        btns.addStretch()
        cancel = QPushButton("Cancelar")
        cancel.setObjectName("btn_secondary")
        cancel.clicked.connect(self.reject)
        save = QPushButton("💾  Salvar Cliente")
        save.clicked.connect(self._save)
        btns.addWidget(cancel)
        btns.addWidget(save)
        layout.addLayout(btns)

    def _populate(self, c):
        self.name_e.setText(c.get("name",""))
        self.doc_e.setText(c.get("cpf_cnpj",""))
        self.phone_e.setText(c.get("phone",""))
        self.email_e.setText(c.get("email",""))
        self.addr_e.setText(c.get("address",""))

    def _save(self):
        if not self.name_e.text().strip():
            QMessageBox.warning(self, "Atenção", "Informe o nome do cliente.")
            return
        self.result_data = {
            "id": self.customer["id"] if self.customer else "",
            "name": self.name_e.text().strip(),
            "cpf_cnpj": self.doc_e.text().strip(),
            "phone": self.phone_e.text().strip(),
            "email": self.email_e.text().strip(),
            "address": self.addr_e.text().strip()
        }
        self.accept()


class SettingsDialog(QDialog):
    def __init__(self, parent, dm: DataManager):
        super().__init__(parent)
        self.dm = dm
        self.settings = dm.get_settings().copy()
        self.setWindowTitle("⚙️  Configurações do Sistema")
        self.setMinimumWidth(460)
        self.setMinimumHeight(360)
        self.setModal(True)
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(16)

        title = SectionTitle("Configurações de Aparência")
        layout.addWidget(title)

        # Theme
        theme_grp = QGroupBox("🎨  Tema da Interface")
        tl = QHBoxLayout(theme_grp)
        tl.setSpacing(20)

        self.rd_dark = QRadioButton("🌙  Modo Escuro (Dark)")
        self.rd_light = QRadioButton("☀️  Modo Claro (Light)")
        self.rd_dark.setChecked(self.settings.get("theme","dark") == "dark")
        self.rd_light.setChecked(self.settings.get("theme","dark") == "light")

        tl.addWidget(self.rd_dark)
        tl.addWidget(self.rd_light)
        tl.addStretch()
        layout.addWidget(theme_grp)

        # Background
        bg_grp = QGroupBox("🖼️  Papel de Parede / Background")
        bl = QVBoxLayout(bg_grp)
        bl.setSpacing(10)

        self.bg_preview = QLabel()
        self.bg_preview.setFixedHeight(120)
        self.bg_preview.setAlignment(Qt.AlignCenter)
        self.bg_preview.setStyleSheet("border:1px dashed #666; border-radius:6px;")

        current_bg = self.settings.get("background_image","")
        if current_bg and os.path.exists(current_bg):
            pix = QPixmap(current_bg).scaled(380, 120, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.bg_preview.setPixmap(pix)
        else:
            self.bg_preview.setText("Nenhuma imagem selecionada\n(o fundo padrão será utilizado)")

        bl.addWidget(self.bg_preview)

        btn_row = QHBoxLayout()
        self.btn_choose_bg = QPushButton("📁  Escolher Imagem")
        self.btn_choose_bg.clicked.connect(self._choose_bg)
        self.btn_clear_bg = QPushButton("✕  Remover Imagem")
        self.btn_clear_bg.setObjectName("btn_secondary")
        self.btn_clear_bg.clicked.connect(self._clear_bg)
        btn_row.addWidget(self.btn_choose_bg)
        btn_row.addWidget(self.btn_clear_bg)
        btn_row.addStretch()
        bl.addLayout(btn_row)

        note = QLabel("Formatos aceitos: JPG, PNG, BMP, WEBP")
        note.setObjectName("subtitle")
        bl.addWidget(note)
        layout.addWidget(bg_grp)

        layout.addStretch()

        btns = QHBoxLayout()
        btns.addStretch()
        cancel = QPushButton("Cancelar")
        cancel.setObjectName("btn_secondary")
        cancel.clicked.connect(self.reject)
        apply = QPushButton("✔  Aplicar Configurações")
        apply.clicked.connect(self._apply)
        btns.addWidget(cancel)
        btns.addWidget(apply)
        layout.addLayout(btns)

    def _choose_bg(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Selecionar Imagem de Fundo", "",
            "Imagens (*.jpg *.jpeg *.png *.bmp *.webp)")
        if path:
            self.settings["background_image"] = path
            pix = QPixmap(path).scaled(380, 120, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.bg_preview.setPixmap(pix)

    def _clear_bg(self):
        self.settings["background_image"] = ""
        self.bg_preview.clear()
        self.bg_preview.setText("Nenhuma imagem selecionada\n(o fundo padrão será utilizado)")

    def _apply(self):
        self.settings["theme"] = "dark" if self.rd_dark.isChecked() else "light"
        self.dm.save_settings(self.settings)
        self.accept()


# ═══════════════════════════════════════════════════════════════════
#  DASHBOARD TAB
# ═══════════════════════════════════════════════════════════════════

class DashboardTab(QWidget):
    def __init__(self, dm: DataManager, parent=None):
        super().__init__(parent)
        self.dm = dm
        self._build_ui()
        self.refresh()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(18)

        # Header
        hdr = QHBoxLayout()
        title = QLabel("📊  Dashboard")
        title.setObjectName("title")
        self.time_lbl = QLabel()
        self.time_lbl.setObjectName("subtitle")
        hdr.addWidget(title)
        hdr.addStretch()
        hdr.addWidget(self.time_lbl)
        layout.addLayout(hdr)

        timer = QTimer(self)
        timer.timeout.connect(self._update_time)
        timer.start(1000)
        self._update_time()

        # Metric Cards
        self.cards_row = QHBoxLayout()
        self.cards_row.setSpacing(14)
        c = DARK_COLORS
        self.card_today = MetricCard("Vendas Hoje", "R$ 0,00", "💰", c["accent"])
        self.card_month = MetricCard("Vendas do Mês", "R$ 0,00", "📈", c["success"])
        self.card_prods = MetricCard("Produtos Cadastrados", "0", "📦", c["warning"])
        self.card_alerts = MetricCard("Alertas de Estoque", "0", "⚠️", c["danger"])
        for card in [self.card_today, self.card_month, self.card_prods, self.card_alerts]:
            self.cards_row.addWidget(card)
        layout.addLayout(self.cards_row)

        # Bottom section: recent sales + low stock
        bottom = QHBoxLayout()
        bottom.setSpacing(16)

        # Recent Sales
        sales_frame = QFrame(); sales_frame.setObjectName("card")
        sl = QVBoxLayout(sales_frame)
        sl.setContentsMargins(14, 14, 14, 14)
        sl.setSpacing(8)
        sl.addWidget(SectionTitle("🛒  Últimas Vendas"))
        self.sales_table = QTableWidget(0, 5)
        self.sales_table.setHorizontalHeaderLabels(["Data","Cliente","Itens","Total","Pagamento"])
        self.sales_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.sales_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.sales_table.setAlternatingRowColors(True)
        self.sales_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.sales_table.verticalHeader().setVisible(False)
        self.sales_table.setMaximumHeight(260)
        sl.addWidget(self.sales_table)
        bottom.addWidget(sales_frame, 3)

        # Low stock
        stock_frame = QFrame(); stock_frame.setObjectName("card")
        stl = QVBoxLayout(stock_frame)
        stl.setContentsMargins(14, 14, 14, 14)
        stl.setSpacing(8)
        stl.addWidget(SectionTitle("⚠️  Estoque Crítico"))
        self.stock_table = QTableWidget(0, 4)
        self.stock_table.setHorizontalHeaderLabels(["Código","Produto","Estoque","Mínimo"])
        self.stock_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.stock_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.stock_table.setAlternatingRowColors(True)
        self.stock_table.verticalHeader().setVisible(False)
        self.stock_table.setMaximumHeight(260)
        stl.addWidget(self.stock_table)
        bottom.addWidget(stock_frame, 2)

        layout.addLayout(bottom)
        layout.addStretch()

    def _update_time(self):
        now = datetime.now()
        self.time_lbl.setText(now.strftime("📅  %d/%m/%Y   🕐  %H:%M:%S"))

    def refresh(self):
        products = self.dm.get_products()
        sales = self.dm.get_sales()
        now = datetime.now()

        today_total = sum(
            s.get("total", 0) for s in sales
            if s.get("date","")[:10] == now.strftime("%Y-%m-%d")
        )
        month_total = sum(
            s.get("total", 0) for s in sales
            if s.get("date","")[:7] == now.strftime("%Y-%m")
        )
        alerts = [p for p in products if p.get("stock",0) <= p.get("min_stock",0)]

        self.card_today.set_value(f"R$ {today_total:,.2f}".replace(",","X").replace(".",",").replace("X","."))
        self.card_month.set_value(f"R$ {month_total:,.2f}".replace(",","X").replace(".",",").replace("X","."))
        self.card_prods.set_value(str(len(products)))
        self.card_alerts.set_value(str(len(alerts)))

        # Recent sales
        recent = sorted(sales, key=lambda x: x.get("date",""), reverse=True)[:10]
        self.sales_table.setRowCount(len(recent))
        for r, s in enumerate(recent):
            dt = s.get("date","")[:16].replace("T"," ")
            self.sales_table.setItem(r, 0, QTableWidgetItem(dt))
            self.sales_table.setItem(r, 1, QTableWidgetItem(s.get("customer_name","—")))
            self.sales_table.setItem(r, 2, QTableWidgetItem(str(len(s.get("items",[])))))
            total_str = f"R$ {s.get('total',0):,.2f}".replace(",","X").replace(".",",").replace("X",".")
            item = QTableWidgetItem(total_str)
            item.setForeground(QColor("#00C853"))
            self.sales_table.setItem(r, 3, item)
            self.sales_table.setItem(r, 4, QTableWidgetItem(s.get("payment_method","—")))

        # Low stock
        self.stock_table.setRowCount(len(alerts))
        for r, p in enumerate(alerts):
            self.stock_table.setItem(r, 0, QTableWidgetItem(p.get("code","")))
            self.stock_table.setItem(r, 1, QTableWidgetItem(p.get("name","")))
            stk_item = QTableWidgetItem(str(p.get("stock",0)))
            stk_item.setForeground(QColor("#FF1744"))
            self.stock_table.setItem(r, 2, stk_item)
            self.stock_table.setItem(r, 3, QTableWidgetItem(str(p.get("min_stock",0))))


# ═══════════════════════════════════════════════════════════════════
#  PRODUCTS TAB
# ═══════════════════════════════════════════════════════════════════

class ProductsTab(QWidget):
    def __init__(self, dm: DataManager, parent=None):
        super().__init__(parent)
        self.dm = dm
        self._build_ui()
        self.refresh()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        # Header
        hdr = QHBoxLayout()
        hdr.addWidget(SectionTitle("📦  Gestão de Produtos"))
        hdr.addStretch()
        layout.addLayout(hdr)

        # Toolbar
        tb = QHBoxLayout()
        self.search = SearchBar("Buscar por código, nome ou marca...")
        self.search.search_changed.connect(self._filter)
        tb.addWidget(self.search, 3)

        self.cat_filter = QComboBox()
        self.cat_filter.addItem("Todas as Categorias")
        self.cat_filter.addItems(ProductDialog.CATEGORIES)
        self.cat_filter.currentTextChanged.connect(self._filter)
        tb.addWidget(self.cat_filter, 1)

        self.low_stock_cb = QCheckBox("⚠️ Apenas Estoque Crítico")
        self.low_stock_cb.stateChanged.connect(self._filter)
        tb.addWidget(self.low_stock_cb)
        tb.addStretch()

        btn_add = QPushButton("➕  Novo Produto")
        btn_add.clicked.connect(self._add_product)
        tb.addWidget(btn_add)

        self.btn_edit = QPushButton("✏️  Editar")
        self.btn_edit.setObjectName("btn_secondary")
        self.btn_edit.clicked.connect(self._edit_product)
        tb.addWidget(self.btn_edit)

        self.btn_del = QPushButton("🗑️  Excluir")
        self.btn_del.setObjectName("btn_danger")
        self.btn_del.clicked.connect(self._delete_product)
        tb.addWidget(self.btn_del)

        layout.addLayout(tb)

        # Table
        self.table = QTableWidget(0, 10)
        self.table.setHorizontalHeaderLabels(
            ["Código","Nome","Categoria","Marca","Unid.","Custo","Venda","Estoque","Mínimo","Status"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.verticalHeader().setVisible(False)
        self.table.doubleClicked.connect(self._edit_product)
        self.table.setColumnWidth(0, 80)
        self.table.setColumnWidth(2, 100)
        self.table.setColumnWidth(3, 100)
        self.table.setColumnWidth(4, 55)
        self.table.setColumnWidth(5, 90)
        self.table.setColumnWidth(6, 90)
        self.table.setColumnWidth(7, 70)
        self.table.setColumnWidth(8, 70)
        self.table.setColumnWidth(9, 90)
        layout.addWidget(self.table)

        self.count_lbl = QLabel()
        self.count_lbl.setObjectName("subtitle")
        layout.addWidget(self.count_lbl)

    def refresh(self):
        self._all_products = self.dm.get_products()
        self._filter()

    def _filter(self):
        txt = self.search.edit.text().lower()
        cat = self.cat_filter.currentText()
        only_low = self.low_stock_cb.isChecked()
        prods = self._all_products

        if txt:
            prods = [p for p in prods if
                     txt in p.get("code","").lower() or
                     txt in p.get("name","").lower() or
                     txt in p.get("brand","").lower()]
        if cat != "Todas as Categorias":
            prods = [p for p in prods if p.get("category","") == cat]
        if only_low:
            prods = [p for p in prods if p.get("stock",0) <= p.get("min_stock",0)]

        self._display_products(prods)
        self.count_lbl.setText(f"Exibindo {len(prods)} de {len(self._all_products)} produtos")

    def _display_products(self, prods):
        self.table.setRowCount(len(prods))
        self._row_ids = []
        for r, p in enumerate(prods):
            self._row_ids.append(p["id"])
            stk = p.get("stock", 0)
            min_stk = p.get("min_stock", 0)
            status = "Esgotado" if stk == 0 else ("Baixo" if stk <= min_stk else "Ativo")

            self.table.setItem(r, 0, QTableWidgetItem(p.get("code","")))
            self.table.setItem(r, 1, QTableWidgetItem(p.get("name","")))
            self.table.setItem(r, 2, QTableWidgetItem(p.get("category","")))
            self.table.setItem(r, 3, QTableWidgetItem(p.get("brand","")))
            self.table.setItem(r, 4, QTableWidgetItem(p.get("unit","")))

            cost = f"R$ {p.get('cost_price',0):,.2f}".replace(",","X").replace(".",",").replace("X",".")
            price = f"R$ {p.get('sale_price',0):,.2f}".replace(",","X").replace(".",",").replace("X",".")
            self.table.setItem(r, 5, QTableWidgetItem(cost))
            self.table.setItem(r, 6, QTableWidgetItem(price))

            stk_item = QTableWidgetItem(str(stk))
            stk_item.setTextAlignment(Qt.AlignCenter)
            if stk == 0:
                stk_item.setForeground(QColor("#FF1744"))
            elif stk <= min_stk:
                stk_item.setForeground(QColor("#FFD600"))
            else:
                stk_item.setForeground(QColor("#00C853"))
            self.table.setItem(r, 7, stk_item)

            min_item = QTableWidgetItem(str(min_stk))
            min_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(r, 8, min_item)

            stat_item = QTableWidgetItem(status)
            stat_item.setTextAlignment(Qt.AlignCenter)
            color_map = {"Ativo": "#00C853", "Baixo": "#FFD600", "Esgotado": "#FF1744"}
            stat_item.setForeground(QColor(color_map.get(status, "#AAAAAA")))
            self.table.setItem(r, 9, stat_item)
            self.table.setRowHeight(r, 36)

    def _get_selected_id(self):
        row = self.table.currentRow()
        if row < 0 or row >= len(self._row_ids):
            return None
        return self._row_ids[row]

    def _add_product(self):
        dlg = ProductDialog(self)
        if dlg.exec_() == QDialog.Accepted:
            self.dm.add_product(dlg.result_data)
            self.refresh()

    def _edit_product(self):
        pid = self._get_selected_id()
        if not pid:
            QMessageBox.information(self, "Atenção", "Selecione um produto para editar.")
            return
        prod = self.dm.get_product_by_id(pid)
        if not prod: return
        dlg = ProductDialog(self, prod)
        if dlg.exec_() == QDialog.Accepted:
            dlg.result_data["id"] = pid
            self.dm.update_product(pid, dlg.result_data)
            self.refresh()

    def _delete_product(self):
        pid = self._get_selected_id()
        if not pid:
            QMessageBox.information(self, "Atenção", "Selecione um produto para excluir.")
            return
        prod = self.dm.get_product_by_id(pid)
        if not prod: return
        resp = QMessageBox.question(
            self, "Confirmar Exclusão",
            f"Deseja excluir o produto:\n\n{prod['name']}?",
            QMessageBox.Yes | QMessageBox.No)
        if resp == QMessageBox.Yes:
            self.dm.delete_product(pid)
            self.refresh()


# ═══════════════════════════════════════════════════════════════════
#  SALES TAB
# ═══════════════════════════════════════════════════════════════════

class SalesTab(QWidget):
    sale_completed = pyqtSignal()

    def __init__(self, dm: DataManager, parent=None):
        super().__init__(parent)
        self.dm = dm
        self.cart = []
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        layout.addWidget(SectionTitle("🛒  Nova Venda"))

        splitter = QSplitter(Qt.Horizontal)
        splitter.setHandleWidth(8)

        # LEFT PANEL: product search + cart
        left = QWidget()
        ll = QVBoxLayout(left)
        ll.setContentsMargins(0, 0, 0, 0)
        ll.setSpacing(10)

        # Product search
        search_grp = QGroupBox("Adicionar Produto")
        sg = QGridLayout(search_grp)
        sg.setSpacing(8)

        sg.addWidget(QLabel("Buscar Produto:"), 0, 0)
        self.prod_search = QLineEdit()
        self.prod_search.setPlaceholderText("Digite código ou nome do produto...")
        self.prod_search.textChanged.connect(self._search_products)
        sg.addWidget(self.prod_search, 0, 1, 1, 3)

        self.prod_list = QTableWidget(0, 5)
        self.prod_list.setHorizontalHeaderLabels(["Código","Nome","Marca","Preço","Estoque"])
        self.prod_list.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.prod_list.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.prod_list.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.prod_list.verticalHeader().setVisible(False)
        self.prod_list.setMaximumHeight(160)
        sg.addWidget(self.prod_list, 1, 0, 1, 4)

        sg.addWidget(QLabel("Qtd:"), 2, 0)
        self.qty_spin = QSpinBox()
        self.qty_spin.setRange(1, 9999); self.qty_spin.setValue(1)
        sg.addWidget(self.qty_spin, 2, 1)

        sg.addWidget(QLabel("Preço Unit.:"), 2, 2)
        self.unit_price = QDoubleSpinBox()
        self.unit_price.setRange(0, 999999); self.unit_price.setPrefix("R$ "); self.unit_price.setDecimals(2)
        sg.addWidget(self.unit_price, 2, 3)

        sg.addWidget(QLabel("Desconto Item:"), 3, 0)
        self.disc_spin = QDoubleSpinBox()
        self.disc_spin.setRange(0, 100); self.disc_spin.setSuffix("%"); self.disc_spin.setDecimals(2)
        sg.addWidget(self.disc_spin, 3, 1)

        btn_add_item = QPushButton("➕  Adicionar ao Carrinho")
        btn_add_item.clicked.connect(self._add_to_cart)
        sg.addWidget(btn_add_item, 3, 2, 1, 2)

        ll.addWidget(search_grp)

        # Cart table
        cart_grp = QGroupBox("Carrinho de Compras")
        cl = QVBoxLayout(cart_grp)
        self.cart_table = QTableWidget(0, 7)
        self.cart_table.setHorizontalHeaderLabels(
            ["Código","Produto","Qtd","Preço Unit.","Desc%","Total",""])
        self.cart_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.cart_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.cart_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.cart_table.setAlternatingRowColors(True)
        self.cart_table.verticalHeader().setVisible(False)
        self.cart_table.setColumnWidth(0, 80)
        self.cart_table.setColumnWidth(2, 50)
        self.cart_table.setColumnWidth(3, 90)
        self.cart_table.setColumnWidth(4, 60)
        self.cart_table.setColumnWidth(5, 90)
        self.cart_table.setColumnWidth(6, 40)
        cl.addWidget(self.cart_table)

        btn_remove = QPushButton("🗑️  Remover Item Selecionado")
        btn_remove.setObjectName("btn_danger")
        btn_remove.clicked.connect(self._remove_from_cart)
        cl.addWidget(btn_remove)
        ll.addWidget(cart_grp)

        splitter.addWidget(left)

        # RIGHT PANEL: customer + payment + totals
        right = QWidget()
        rl = QVBoxLayout(right)
        rl.setContentsMargins(0, 0, 0, 0)
        rl.setSpacing(12)

        # Customer
        cust_grp = QGroupBox("👤  Cliente")
        cgl = QVBoxLayout(cust_grp)
        self.cust_combo = QComboBox()
        self.cust_combo.addItem("-- Consumidor Final --", None)
        for c in self.dm.get_customers():
            self.cust_combo.addItem(f"{c['name']} | {c.get('cpf_cnpj','')}", c["id"])
        cgl.addWidget(self.cust_combo)
        rl.addWidget(cust_grp)

        # Payment
        pay_grp = QGroupBox("💳  Pagamento")
        pgl = QFormLayout(pay_grp)
        self.pay_method = QComboBox()
        self.pay_method.addItems([
            "Dinheiro","Cartão de Crédito","Cartão de Débito",
            "PIX","Boleto","Transferência","Fiado / A Prazo"])
        pgl.addRow("Forma:", self.pay_method)

        self.discount_total = QDoubleSpinBox()
        self.discount_total.setRange(0, 100); self.discount_total.setSuffix("%"); self.discount_total.setDecimals(2)
        self.discount_total.valueChanged.connect(self._update_totals)
        pgl.addRow("Desconto Geral:", self.discount_total)

        self.obs_e = QTextEdit()
        self.obs_e.setMaximumHeight(60)
        self.obs_e.setPlaceholderText("Observações da venda...")
        pgl.addRow("Observações:", self.obs_e)
        rl.addWidget(pay_grp)

        # Totals
        totals_grp = QGroupBox("💰  Totais")
        tgl = QGridLayout(totals_grp)

        def tot_row(label, attr, r, big=False):
            lbl = QLabel(label)
            val = QLabel("R$ 0,00")
            val.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            if big:
                lbl.setStyleSheet("font-size:16px; font-weight:bold;")
                val.setStyleSheet("font-size:20px; font-weight:bold; color:#FF6B35;")
            setattr(self, attr, val)
            tgl.addWidget(lbl, r, 0)
            tgl.addWidget(val, r, 1)

        tot_row("Subtotal:", "lbl_sub", 0)
        tot_row("Desconto:", "lbl_disc", 1)
        sep = QFrame(); sep.setFrameShape(QFrame.HLine); sep.setStyleSheet("color:#444;")
        tgl.addWidget(sep, 2, 0, 1, 2)
        tot_row("TOTAL:", "lbl_total", 3, big=True)
        rl.addWidget(totals_grp)

        rl.addStretch()

        btn_finalize = QPushButton("✅  FINALIZAR VENDA")
        btn_finalize.setObjectName("btn_success")
        btn_finalize.setMinimumHeight(50)
        btn_finalize.setFont(QFont("Segoe UI", 14, QFont.Bold))
        btn_finalize.clicked.connect(self._finalize_sale)
        rl.addWidget(btn_finalize)

        btn_clear = QPushButton("🗑️  Limpar Venda")
        btn_clear.setObjectName("btn_secondary")
        btn_clear.clicked.connect(self._clear_sale)
        rl.addWidget(btn_clear)

        splitter.addWidget(right)
        splitter.setSizes([650, 320])

        layout.addWidget(splitter, 1)

        self._search_products("")

    def _search_products(self, txt=""):
        prods = self.dm.get_products()
        if txt:
            prods = [p for p in prods if txt.lower() in p.get("code","").lower()
                     or txt.lower() in p.get("name","").lower()]
        prods = prods[:50]
        self.prod_list.setRowCount(len(prods))
        self._prod_ids_in_list = []
        for r, p in enumerate(prods):
            self._prod_ids_in_list.append(p["id"])
            self.prod_list.setItem(r, 0, QTableWidgetItem(p.get("code","")))
            self.prod_list.setItem(r, 1, QTableWidgetItem(p.get("name","")))
            self.prod_list.setItem(r, 2, QTableWidgetItem(p.get("brand","")))
            price_str = f"R$ {p.get('sale_price',0):,.2f}".replace(",","X").replace(".",",").replace("X",".")
            self.prod_list.setItem(r, 3, QTableWidgetItem(price_str))
            stk_item = QTableWidgetItem(str(p.get("stock",0)))
            stk_item.setForeground(QColor("#FF1744") if p.get("stock",0) == 0 else QColor("#00C853"))
            self.prod_list.setItem(r, 4, stk_item)
            self.prod_list.setRowHeight(r, 30)

        self.prod_list.selectionModel().selectionChanged.connect(self._on_product_select)

    def _on_product_select(self):
        row = self.prod_list.currentRow()
        if row < 0 or row >= len(self._prod_ids_in_list): return
        pid = self._prod_ids_in_list[row]
        prod = self.dm.get_product_by_id(pid)
        if prod:
            self.unit_price.setValue(prod.get("sale_price", 0))

    def _add_to_cart(self):
        row = self.prod_list.currentRow()
        if row < 0 or row >= len(getattr(self, "_prod_ids_in_list", [])):
            QMessageBox.information(self, "Atenção", "Selecione um produto da lista.")
            return
        pid = self._prod_ids_in_list[row]
        prod = self.dm.get_product_by_id(pid)
        if not prod: return

        qty = self.qty_spin.value()
        price = self.unit_price.value()
        disc = self.disc_spin.value()

        if prod.get("stock", 0) < qty:
            QMessageBox.warning(self, "Estoque Insuficiente",
                                f"Estoque disponível: {prod.get('stock',0)} {prod.get('unit','UN')}")
            return

        # Check if already in cart
        for item in self.cart:
            if item["product_id"] == pid:
                item["quantity"] += qty
                item["unit_price"] = price
                item["discount"] = disc
                self._update_cart_table()
                self._update_totals()
                return

        self.cart.append({
            "product_id": pid,
            "code": prod.get("code",""),
            "name": prod.get("name",""),
            "quantity": qty,
            "unit_price": price,
            "discount": disc,
        })
        self._update_cart_table()
        self._update_totals()
        self.qty_spin.setValue(1)
        self.disc_spin.setValue(0)

    def _update_cart_table(self):
        self.cart_table.setRowCount(len(self.cart))
        for r, item in enumerate(self.cart):
            disc_factor = 1 - (item["discount"] / 100)
            total = item["quantity"] * item["unit_price"] * disc_factor
            self.cart_table.setItem(r, 0, QTableWidgetItem(item["code"]))
            self.cart_table.setItem(r, 1, QTableWidgetItem(item["name"]))
            self.cart_table.setItem(r, 2, QTableWidgetItem(str(item["quantity"])))
            p_str = f"R$ {item['unit_price']:,.2f}".replace(",","X").replace(".",",").replace("X",".")
            self.cart_table.setItem(r, 3, QTableWidgetItem(p_str))
            self.cart_table.setItem(r, 4, QTableWidgetItem(f"{item['discount']:.1f}%"))
            t_str = f"R$ {total:,.2f}".replace(",","X").replace(".",",").replace("X",".")
            t_item = QTableWidgetItem(t_str)
            t_item.setForeground(QColor("#00C853"))
            self.cart_table.setItem(r, 5, t_item)
            del_btn = QPushButton("✕")
            del_btn.setFixedSize(28, 28)
            del_btn.setObjectName("btn_danger")
            del_btn.clicked.connect(lambda _, i=r: self._remove_item(i))
            self.cart_table.setCellWidget(r, 6, del_btn)
            self.cart_table.setRowHeight(r, 36)

    def _remove_item(self, row):
        if 0 <= row < len(self.cart):
            self.cart.pop(row)
            self._update_cart_table()
            self._update_totals()

    def _remove_from_cart(self):
        row = self.cart_table.currentRow()
        self._remove_item(row)

    def _update_totals(self):
        sub = sum(i["quantity"] * i["unit_price"] * (1 - i["discount"]/100) for i in self.cart)
        disc_pct = self.discount_total.value()
        disc_val = sub * disc_pct / 100
        total = sub - disc_val

        def fmt(v): return f"R$ {v:,.2f}".replace(",","X").replace(".",",").replace("X",".")
        self.lbl_sub.setText(fmt(sub))
        self.lbl_disc.setText(f"-{fmt(disc_val)}")
        self.lbl_total.setText(fmt(total))

    def _finalize_sale(self):
        if not self.cart:
            QMessageBox.warning(self, "Carrinho Vazio", "Adicione produtos ao carrinho antes de finalizar.")
            return

        sub = sum(i["quantity"] * i["unit_price"] * (1 - i["discount"]/100) for i in self.cart)
        disc_pct = self.discount_total.value()
        disc_val = sub * disc_pct / 100
        total = sub - disc_val

        cust_id = self.cust_combo.currentData()
        cust_name = self.cust_combo.currentText().split("|")[0].strip()
        if cust_id is None:
            cust_name = "Consumidor Final"

        sale = {
            "customer_id": cust_id or "",
            "customer_name": cust_name,
            "items": self.cart.copy(),
            "subtotal": sub,
            "discount_pct": disc_pct,
            "discount_value": disc_val,
            "total": total,
            "payment_method": self.pay_method.currentText(),
            "observations": self.obs_e.toPlainText(),
            "status": "concluída"
        }

        sid = self.dm.add_sale(sale)
        self.sale_completed.emit()

        total_str = f"R$ {total:,.2f}".replace(",","X").replace(".",",").replace("X",".")
        QMessageBox.information(
            self, "✅  Venda Concluída",
            f"Venda registrada com sucesso!\n\n"
            f"Cliente: {cust_name}\n"
            f"Total: {total_str}\n"
            f"Pagamento: {self.pay_method.currentText()}\n"
            f"ID: {sid[:8]}...")

        self._clear_sale()

    def _clear_sale(self):
        self.cart = []
        self._update_cart_table()
        self._update_totals()
        self.cust_combo.setCurrentIndex(0)
        self.pay_method.setCurrentIndex(0)
        self.discount_total.setValue(0)
        self.obs_e.clear()

    def refresh_customers(self):
        self.cust_combo.clear()
        self.cust_combo.addItem("-- Consumidor Final --", None)
        for c in self.dm.get_customers():
            self.cust_combo.addItem(f"{c['name']} | {c.get('cpf_cnpj','')}", c["id"])


# ═══════════════════════════════════════════════════════════════════
#  CUSTOMERS TAB
# ═══════════════════════════════════════════════════════════════════

class CustomersTab(QWidget):
    def __init__(self, dm: DataManager, parent=None):
        super().__init__(parent)
        self.dm = dm
        self._build_ui()
        self.refresh()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        layout.addWidget(SectionTitle("👥  Gestão de Clientes"))

        tb = QHBoxLayout()
        self.search = SearchBar("Buscar por nome, CPF/CNPJ ou telefone...")
        self.search.search_changed.connect(self._filter)
        tb.addWidget(self.search, 3)
        tb.addStretch()

        btn_add = QPushButton("➕  Novo Cliente")
        btn_add.clicked.connect(self._add)
        tb.addWidget(btn_add)

        self.btn_edit = QPushButton("✏️  Editar")
        self.btn_edit.setObjectName("btn_secondary")
        self.btn_edit.clicked.connect(self._edit)
        tb.addWidget(self.btn_edit)

        self.btn_del = QPushButton("🗑️  Excluir")
        self.btn_del.setObjectName("btn_danger")
        self.btn_del.clicked.connect(self._delete)
        tb.addWidget(self.btn_del)

        self.btn_hist = QPushButton("📋  Histórico")
        self.btn_hist.setObjectName("btn_secondary")
        self.btn_hist.clicked.connect(self._history)
        tb.addWidget(self.btn_hist)

        layout.addLayout(tb)

        self.table = QTableWidget(0, 6)
        self.table.setHorizontalHeaderLabels(
            ["Nome","CPF / CNPJ","Telefone","E-mail","Endereço","Total Compras"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.verticalHeader().setVisible(False)
        self.table.doubleClicked.connect(self._edit)
        self.table.setColumnWidth(1, 130)
        self.table.setColumnWidth(2, 130)
        self.table.setColumnWidth(3, 160)
        self.table.setColumnWidth(4, 200)
        self.table.setColumnWidth(5, 110)
        layout.addWidget(self.table)

        self.count_lbl = QLabel()
        self.count_lbl.setObjectName("subtitle")
        layout.addWidget(self.count_lbl)

    def refresh(self):
        self._all = self.dm.get_customers()
        self._filter()

    def _filter(self):
        txt = self.search.edit.text().lower()
        custs = self._all
        if txt:
            custs = [c for c in custs if
                     txt in c.get("name","").lower() or
                     txt in c.get("cpf_cnpj","").lower() or
                     txt in c.get("phone","").lower()]
        self._display(custs)
        self.count_lbl.setText(f"Exibindo {len(custs)} de {len(self._all)} clientes")

    def _display(self, custs):
        self.table.setRowCount(len(custs))
        self._row_ids = []
        sales = self.dm.get_sales()
        for r, c in enumerate(custs):
            self._row_ids.append(c["id"])
            total = sum(s.get("total",0) for s in sales if s.get("customer_id") == c["id"])
            self.table.setItem(r, 0, QTableWidgetItem(c.get("name","")))
            self.table.setItem(r, 1, QTableWidgetItem(c.get("cpf_cnpj","")))
            self.table.setItem(r, 2, QTableWidgetItem(c.get("phone","")))
            self.table.setItem(r, 3, QTableWidgetItem(c.get("email","")))
            self.table.setItem(r, 4, QTableWidgetItem(c.get("address","")))
            tot_str = f"R$ {total:,.2f}".replace(",","X").replace(".",",").replace("X",".")
            t_item = QTableWidgetItem(tot_str)
            t_item.setForeground(QColor("#00C853"))
            self.table.setItem(r, 5, t_item)
            self.table.setRowHeight(r, 36)

    def _get_id(self):
        row = self.table.currentRow()
        if row < 0 or row >= len(getattr(self, "_row_ids", [])): return None
        return self._row_ids[row]

    def _add(self):
        dlg = CustomerDialog(self)
        if dlg.exec_() == QDialog.Accepted:
            self.dm.add_customer(dlg.result_data)
            self.refresh()

    def _edit(self):
        cid = self._get_id()
        if not cid:
            QMessageBox.information(self, "Atenção", "Selecione um cliente para editar.")
            return
        cust = self.dm.get_customer_by_id(cid)
        if not cust: return
        dlg = CustomerDialog(self, cust)
        if dlg.exec_() == QDialog.Accepted:
            dlg.result_data["id"] = cid
            self.dm.update_customer(cid, dlg.result_data)
            self.refresh()

    def _delete(self):
        cid = self._get_id()
        if not cid:
            QMessageBox.information(self, "Atenção", "Selecione um cliente para excluir.")
            return
        cust = self.dm.get_customer_by_id(cid)
        if not cust: return
        resp = QMessageBox.question(
            self, "Confirmar Exclusão",
            f"Deseja excluir o cliente:\n\n{cust['name']}?",
            QMessageBox.Yes | QMessageBox.No)
        if resp == QMessageBox.Yes:
            self.dm.delete_customer(cid)
            self.refresh()

    def _history(self):
        cid = self._get_id()
        if not cid:
            QMessageBox.information(self, "Atenção", "Selecione um cliente.")
            return
        cust = self.dm.get_customer_by_id(cid)
        if not cust: return
        sales = [s for s in self.dm.get_sales() if s.get("customer_id") == cid]
        dlg = QDialog(self)
        dlg.setWindowTitle(f"Histórico - {cust['name']}")
        dlg.setMinimumSize(640, 420)
        l = QVBoxLayout(dlg)
        l.setContentsMargins(16, 16, 16, 16)
        l.addWidget(QLabel(f"<b>Cliente:</b> {cust['name']} | Total de vendas: {len(sales)}"))
        tbl = QTableWidget(len(sales), 5)
        tbl.setHorizontalHeaderLabels(["Data","Itens","Subtotal","Desconto","Total"])
        tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        tbl.verticalHeader().setVisible(False)
        for r, s in enumerate(sorted(sales, key=lambda x: x.get("date",""), reverse=True)):
            fmt = lambda v: f"R$ {v:,.2f}".replace(",","X").replace(".",",").replace("X",".")
            tbl.setItem(r, 0, QTableWidgetItem(s.get("date","")[:16].replace("T"," ")))
            tbl.setItem(r, 1, QTableWidgetItem(str(len(s.get("items",[])))))
            tbl.setItem(r, 2, QTableWidgetItem(fmt(s.get("subtotal",0))))
            tbl.setItem(r, 3, QTableWidgetItem(fmt(s.get("discount_value",0))))
            t_item = QTableWidgetItem(fmt(s.get("total",0)))
            t_item.setForeground(QColor("#00C853"))
            tbl.setItem(r, 4, t_item)
        l.addWidget(tbl)
        total_gasto = sum(s.get("total",0) for s in sales)
        lbl = QLabel(f"<b>Total Gasto:</b> R$ {total_gasto:,.2f}".replace(",","X").replace(".",",").replace("X","."))
        lbl.setStyleSheet("font-size:14px; color:#FF6B35; font-weight:bold;")
        l.addWidget(lbl)
        btn = QPushButton("Fechar"); btn.clicked.connect(dlg.accept)
        l.addWidget(btn)
        dlg.exec_()


# ═══════════════════════════════════════════════════════════════════
#  REPORTS TAB
# ═══════════════════════════════════════════════════════════════════

class ReportsTab(QWidget):
    def __init__(self, dm: DataManager, parent=None):
        super().__init__(parent)
        self.dm = dm
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(14)

        layout.addWidget(SectionTitle("📊  Relatórios"))

        # Controls row
        ctrl = QHBoxLayout()
        ctrl.setSpacing(10)

        ctrl.addWidget(QLabel("Relatório:"))
        self.report_type = QComboBox()
        self.report_type.addItems([
            "Vendas por Período",
            "Vendas por Cliente",
            "Vendas por Produto",
            "Estoque Atual",
            "Produtos Críticos (Estoque Baixo)",
            "Resumo Financeiro",
            "Ranking de Produtos Mais Vendidos",
        ])
        self.report_type.setMinimumWidth(250)
        ctrl.addWidget(self.report_type)

        ctrl.addWidget(QLabel("De:"))
        self.date_from = QDateEdit(QDate.currentDate().addDays(-30))
        self.date_from.setCalendarPopup(True)
        self.date_from.setDisplayFormat("dd/MM/yyyy")
        ctrl.addWidget(self.date_from)

        ctrl.addWidget(QLabel("Até:"))
        self.date_to = QDateEdit(QDate.currentDate())
        self.date_to.setCalendarPopup(True)
        self.date_to.setDisplayFormat("dd/MM/yyyy")
        ctrl.addWidget(self.date_to)

        btn_gen = QPushButton("🔍  Gerar Relatório")
        btn_gen.clicked.connect(self._generate)
        ctrl.addWidget(btn_gen)

        if EXCEL_OK:
            btn_xls = QPushButton("📥  Exportar Excel")
            btn_xls.setObjectName("btn_success")
            btn_xls.clicked.connect(self._export_excel)
            ctrl.addWidget(btn_xls)

        ctrl.addStretch()
        layout.addLayout(ctrl)

        # Summary cards
        self.summary_row = QHBoxLayout()
        self.summary_row.setSpacing(12)
        self.s1 = MetricCard("Total de Registros","0","📋")
        self.s2 = MetricCard("Valor Total","R$ 0,00","💰",DARK_COLORS["success"])
        self.s3 = MetricCard("Média por Registro","R$ 0,00","📈",DARK_COLORS["accent"])
        self.s4 = MetricCard("Período","0 dias","📅",DARK_COLORS["warning"])
        for c in [self.s1, self.s2, self.s3, self.s4]:
            c.setMaximumHeight(100)
            self.summary_row.addWidget(c)
        layout.addLayout(self.summary_row)

        # Report table
        self.table = QTableWidget(0, 1)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.verticalHeader().setVisible(False)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.table, 1)

        self.info_lbl = QLabel("Selecione um tipo de relatório e clique em 'Gerar Relatório'")
        self.info_lbl.setObjectName("subtitle")
        self.info_lbl.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.info_lbl)

        self._current_report_data = []
        self._current_headers = []
        self._generate()

    def _generate(self):
        rtype = self.report_type.currentText()
        d_from = self.date_from.date().toString("yyyy-MM-dd")
        d_to = self.date_to.date().toString("yyyy-MM-dd") + "T23:59:59"

        sales = [s for s in self.dm.get_sales()
                 if d_from <= s.get("date","") <= d_to]

        days = self.date_from.date().daysTo(self.date_to.date()) + 1
        self.s4.set_value(f"{days} dias")

        if rtype == "Vendas por Período":
            self._report_by_period(sales)
        elif rtype == "Vendas por Cliente":
            self._report_by_customer(sales)
        elif rtype == "Vendas por Produto":
            self._report_by_product(sales)
        elif rtype == "Estoque Atual":
            self._report_stock()
        elif rtype == "Produtos Críticos (Estoque Baixo)":
            self._report_low_stock()
        elif rtype == "Resumo Financeiro":
            self._report_financial(sales)
        elif rtype == "Ranking de Produtos Mais Vendidos":
            self._report_top_products(sales)

    def _set_table(self, headers, rows, money_cols=None):
        self._current_headers = headers
        self._current_report_data = rows
        money_cols = money_cols or []
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                item = QTableWidgetItem(str(val))
                if c in money_cols:
                    item.setForeground(QColor("#00C853"))
                self.table.setItem(r, c, item)
                self.table.setRowHeight(r, 32)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.info_lbl.setText(f"{len(rows)} registro(s) encontrado(s).")

    def _fmt_money(self, v):
        return f"R$ {v:,.2f}".replace(",","X").replace(".",",").replace("X",".")

    def _report_by_period(self, sales):
        from collections import defaultdict
        by_day = defaultdict(lambda: {"count":0, "total":0})
        for s in sales:
            day = s.get("date","")[:10]
            by_day[day]["count"] += 1
            by_day[day]["total"] += s.get("total",0)
        rows = []
        for day in sorted(by_day.keys(), reverse=True):
            d = datetime.strptime(day, "%Y-%m-%d")
            rows.append([
                d.strftime("%d/%m/%Y"), d.strftime("%A").capitalize(),
                by_day[day]["count"],
                self._fmt_money(by_day[day]["total"])
            ])
        total = sum(s.get("total",0) for s in sales)
        self._set_table(["Data","Dia da Semana","Nº Vendas","Total"], rows, [3])
        self._update_summary(len(sales), total)

    def _report_by_customer(self, sales):
        from collections import defaultdict
        by_cust = defaultdict(lambda: {"count":0,"total":0,"name":""})
        for s in sales:
            cid = s.get("customer_id","") or "consumidor_final"
            by_cust[cid]["count"] += 1
            by_cust[cid]["total"] += s.get("total",0)
            by_cust[cid]["name"] = s.get("customer_name","Consumidor Final")
        rows = sorted(by_cust.values(), key=lambda x: x["total"], reverse=True)
        data = [[r["name"], r["count"], self._fmt_money(r["total"])] for r in rows]
        total = sum(s.get("total",0) for s in sales)
        self._set_table(["Cliente","Nº Vendas","Total Gasto"], data, [2])
        self._update_summary(len(sales), total)

    def _report_by_product(self, sales):
        from collections import defaultdict
        by_prod = defaultdict(lambda: {"qty":0,"total":0,"name":"","code":""})
        for s in sales:
            for it in s.get("items",[]):
                pid = it.get("product_id","")
                by_prod[pid]["qty"] += it.get("quantity",0)
                by_prod[pid]["total"] += it.get("quantity",0)*it.get("unit_price",0)*(1-it.get("discount",0)/100)
                by_prod[pid]["name"] = it.get("name","")
                by_prod[pid]["code"] = it.get("code","")
        rows = sorted(by_prod.values(), key=lambda x: x["total"], reverse=True)
        data = [[r["code"],r["name"],r["qty"],self._fmt_money(r["total"])] for r in rows]
        total = sum(r["total"] for r in by_prod.values())
        self._set_table(["Código","Produto","Qtd Vendida","Total"], data, [3])
        self._update_summary(len(data), total)

    def _report_stock(self):
        prods = self.dm.get_products()
        rows = []
        for p in sorted(prods, key=lambda x: x.get("name","")):
            stk = p.get("stock",0)
            status = "Esgotado" if stk==0 else ("Crítico" if stk<=p.get("min_stock",0) else "Normal")
            valor = stk * p.get("cost_price",0)
            rows.append([
                p.get("code",""), p.get("name",""), p.get("category",""),
                p.get("brand",""), str(stk), str(p.get("min_stock",0)),
                self._fmt_money(p.get("sale_price",0)),
                self._fmt_money(valor), status
            ])
        total_val = sum(p.get("stock",0)*p.get("cost_price",0) for p in prods)
        self._set_table(
            ["Código","Produto","Categoria","Marca","Estoque","Mínimo","Preço Venda","Val. Estoque","Status"],
            rows, [6,7])
        self.s1.set_value(str(len(prods)))
        self.s2.set_value(self._fmt_money(total_val))
        self.s3.set_value(self._fmt_money(total_val/max(1,len(prods))))
        self.info_lbl.setText(f"{len(prods)} produtos no estoque.")

    def _report_low_stock(self):
        prods = [p for p in self.dm.get_products() if p.get("stock",0) <= p.get("min_stock",0)]
        rows = []
        for p in sorted(prods, key=lambda x: x.get("stock",0)):
            rows.append([
                p.get("code",""), p.get("name",""), p.get("category",""),
                str(p.get("stock",0)), str(p.get("min_stock",0)),
                str(max(0, p.get("min_stock",0)-p.get("stock",0))),
                self._fmt_money(p.get("sale_price",0))
            ])
        self._set_table(
            ["Código","Produto","Categoria","Estoque Atual","Mínimo","Qtd p/ Repor","Preço Venda"],
            rows, [6])
        self.s1.set_value(str(len(prods)))
        self.s2.set_value("—")
        self.s3.set_value("—")
        self.info_lbl.setText(f"{len(prods)} produtos com estoque crítico.")

    def _report_financial(self, sales):
        total_bruto = sum(s.get("subtotal",0) for s in sales)
        total_desc = sum(s.get("discount_value",0) for s in sales)
        total_liq = sum(s.get("total",0) for s in sales)
        by_pay = defaultdict(float)
        for s in sales:
            by_pay[s.get("payment_method","—")] += s.get("total",0)
        rows = []
        for method, val in sorted(by_pay.items(), key=lambda x: x[1], reverse=True):
            pct = (val/max(1,total_liq))*100
            rows.append([method, self._fmt_money(val), f"{pct:.1f}%"])
        rows.append(["─"*20, "─"*12, "─"*8])
        rows.append(["TOTAL BRUTO", self._fmt_money(total_bruto), "100%"])
        rows.append(["DESCONTOS", f"-{self._fmt_money(total_desc)}", ""])
        rows.append(["TOTAL LÍQUIDO", self._fmt_money(total_liq), ""])
        self._set_table(["Forma de Pagamento","Valor","% do Total"], rows, [1])
        self._update_summary(len(sales), total_liq)

    def _report_top_products(self, sales):
        from collections import defaultdict
        by_prod = defaultdict(lambda: {"qty":0,"total":0,"name":"","code":""})
        for s in sales:
            for it in s.get("items",[]):
                pid = it.get("product_id","")
                by_prod[pid]["qty"] += it.get("quantity",0)
                by_prod[pid]["total"] += it.get("quantity",0)*it.get("unit_price",0)*(1-it.get("discount",0)/100)
                by_prod[pid]["name"] = it.get("name","")
                by_prod[pid]["code"] = it.get("code","")
        rows_sorted = sorted(by_prod.values(), key=lambda x: x["qty"], reverse=True)[:20]
        data = [[f"#{i+1}", r["code"], r["name"], r["qty"], self._fmt_money(r["total"])]
                for i, r in enumerate(rows_sorted)]
        total = sum(r["total"] for r in rows_sorted)
        self._set_table(["Pos.","Código","Produto","Qtd Vendida","Total Faturado"], data, [4])
        self._update_summary(len(data), total)

    def _update_summary(self, count, total):
        self.s1.set_value(str(count))
        self.s2.set_value(self._fmt_money(total))
        self.s3.set_value(self._fmt_money(total/max(1,count)))

    def _export_excel(self):
        if not EXCEL_OK:
            QMessageBox.warning(self, "Erro", "openpyxl não está instalado.")
            return
        if not self._current_report_data:
            QMessageBox.information(self, "Atenção", "Gere um relatório antes de exportar.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Salvar Relatório Excel", f"relatorio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel (*.xlsx)")
        if not path: return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório"

        # Header style
        header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        header_font = XLFont(bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Title
        report_name = self.report_type.currentText()
        ws.merge_cells(f"A1:{chr(64+len(self._current_headers))}1")
        title_cell = ws["A1"]
        title_cell.value = f"SISTEMA AUTO PEÇAS — {report_name.upper()}"
        title_cell.font = XLFont(bold=True, size=14, color="FF6B35")
        title_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 28

        ws.merge_cells(f"A2:{chr(64+len(self._current_headers))}2")
        period_cell = ws["A2"]
        period_cell.value = (f"Período: {self.date_from.date().toString('dd/MM/yyyy')} "
                             f"a {self.date_to.date().toString('dd/MM/yyyy')}  |  "
                             f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        period_cell.font = XLFont(size=10, color="888888")
        period_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 18

        # Headers
        for c, h in enumerate(self._current_headers, 1):
            cell = ws.cell(row=4, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border
        ws.row_dimensions[4].height = 22

        # Data rows
        alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        for r, row in enumerate(self._current_report_data, 5):
            for c, val in enumerate(row, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = border
                if (r - 5) % 2 == 1:
                    cell.fill = alt_fill
            ws.row_dimensions[r].height = 18

        # Auto-width
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except: pass
            ws.column_dimensions[col_letter].width = min(50, max(10, max_len + 4))

        # Summary at bottom
        last_row = len(self._current_report_data) + 6
        ws.cell(row=last_row, column=1, value="Total de Registros:").font = XLFont(bold=True)
        ws.cell(row=last_row, column=2, value=self.s1.val_lbl.text())
        ws.cell(row=last_row+1, column=1, value="Valor Total:").font = XLFont(bold=True)
        ws.cell(row=last_row+1, column=2, value=self.s2.val_lbl.text())

        try:
            wb.save(path)
            QMessageBox.information(self, "✅  Exportado com Sucesso",
                                    f"Relatório salvo em:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Erro ao Salvar", str(e))


# ═══════════════════════════════════════════════════════════════════
#  SALES HISTORY TAB
# ═══════════════════════════════════════════════════════════════════

class SalesHistoryTab(QWidget):
    def __init__(self, dm: DataManager, parent=None):
        super().__init__(parent)
        self.dm = dm
        self._build_ui()
        self.refresh()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        layout.addWidget(SectionTitle("📋  Histórico de Vendas"))

        tb = QHBoxLayout()
        self.search = SearchBar("Buscar por cliente ou ID...")
        self.search.search_changed.connect(self._filter)
        tb.addWidget(self.search, 2)

        tb.addWidget(QLabel("De:"))
        self.d_from = QDateEdit(QDate.currentDate().addDays(-30))
        self.d_from.setCalendarPopup(True)
        self.d_from.setDisplayFormat("dd/MM/yyyy")
        tb.addWidget(self.d_from)

        tb.addWidget(QLabel("Até:"))
        self.d_to = QDateEdit(QDate.currentDate())
        self.d_to.setCalendarPopup(True)
        self.d_to.setDisplayFormat("dd/MM/yyyy")
        tb.addWidget(self.d_to)

        btn_filter = QPushButton("🔍  Filtrar")
        btn_filter.clicked.connect(self.refresh)
        tb.addWidget(btn_filter)
        tb.addStretch()
        layout.addLayout(tb)

        self.table = QTableWidget(0, 7)
        self.table.setHorizontalHeaderLabels(
            ["Data/Hora","Cliente","Itens","Subtotal","Desconto","Total","Pagamento"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.verticalHeader().setVisible(False)
        self.table.doubleClicked.connect(self._view_detail)
        layout.addWidget(self.table, 1)

        self.info_lbl = QLabel()
        self.info_lbl.setObjectName("subtitle")
        layout.addWidget(self.info_lbl)

    def refresh(self):
        d_from = self.d_from.date().toString("yyyy-MM-dd")
        d_to = self.d_to.date().toString("yyyy-MM-dd") + "T23:59:59"
        txt = self.search.edit.text().lower()
        sales = self.dm.get_sales()
        sales = [s for s in sales if d_from <= s.get("date","") <= d_to]
        if txt:
            sales = [s for s in sales if
                     txt in s.get("customer_name","").lower() or
                     txt in s.get("id","").lower()]
        sales = sorted(sales, key=lambda x: x.get("date",""), reverse=True)

        self._all_sales = sales
        self.table.setRowCount(len(sales))
        total_period = 0
        for r, s in enumerate(sales):
            dt = s.get("date","")[:16].replace("T"," ")
            total_period += s.get("total",0)
            fmt = lambda v: f"R$ {v:,.2f}".replace(",","X").replace(".",",").replace("X",".")
            self.table.setItem(r, 0, QTableWidgetItem(dt))
            self.table.setItem(r, 1, QTableWidgetItem(s.get("customer_name","—")))
            self.table.setItem(r, 2, QTableWidgetItem(str(len(s.get("items",[])))))
            self.table.setItem(r, 3, QTableWidgetItem(fmt(s.get("subtotal",0))))
            self.table.setItem(r, 4, QTableWidgetItem(fmt(s.get("discount_value",0))))
            t_item = QTableWidgetItem(fmt(s.get("total",0)))
            t_item.setForeground(QColor("#00C853"))
            self.table.setItem(r, 5, t_item)
            self.table.setItem(r, 6, QTableWidgetItem(s.get("payment_method","—")))
            self.table.setRowHeight(r, 34)

        total_str = f"R$ {total_period:,.2f}".replace(",","X").replace(".",",").replace("X",".")
        self.info_lbl.setText(f"{len(sales)} venda(s) | Total do período: {total_str}")

    def _filter(self):
        self.refresh()

    def _view_detail(self):
        row = self.table.currentRow()
        if row < 0 or row >= len(self._all_sales): return
        s = self._all_sales[row]
        dlg = QDialog(self)
        dlg.setWindowTitle(f"Detalhes da Venda #{s.get('id','')[:8]}")
        dlg.setMinimumSize(560, 420)
        l = QVBoxLayout(dlg)
        l.setContentsMargins(16,16,16,16)
        l.setSpacing(10)

        fmt = lambda v: f"R$ {v:,.2f}".replace(",","X").replace(".",",").replace("X",".")
        info_html = (f"<b>Data:</b> {s.get('date','')[:16].replace('T',' ')}  &nbsp;&nbsp; "
                     f"<b>Cliente:</b> {s.get('customer_name','—')}  &nbsp;&nbsp; "
                     f"<b>Pagamento:</b> {s.get('payment_method','—')}")
        l.addWidget(QLabel(info_html))

        tbl = QTableWidget(len(s.get("items",[])), 5)
        tbl.setHorizontalHeaderLabels(["Código","Produto","Qtd","Preço Unit.","Total"])
        tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        tbl.verticalHeader().setVisible(False)
        for r, it in enumerate(s.get("items",[])):
            disc = 1 - it.get("discount",0)/100
            total = it["quantity"] * it["unit_price"] * disc
            tbl.setItem(r, 0, QTableWidgetItem(it.get("code","")))
            tbl.setItem(r, 1, QTableWidgetItem(it.get("name","")))
            tbl.setItem(r, 2, QTableWidgetItem(str(it["quantity"])))
            tbl.setItem(r, 3, QTableWidgetItem(fmt(it["unit_price"])))
            t_it = QTableWidgetItem(fmt(total))
            t_it.setForeground(QColor("#00C853"))
            tbl.setItem(r, 4, t_it)
        l.addWidget(tbl)

        sums = QHBoxLayout()
        sums.addStretch()
        sums_txt = (f"Subtotal: {fmt(s.get('subtotal',0))}  |  "
                    f"Desconto: -{fmt(s.get('discount_value',0))}  |  "
                    f"<b>TOTAL: {fmt(s.get('total',0))}</b>")
        sums_lbl = QLabel(sums_txt)
        sums_lbl.setStyleSheet("color:#FF6B35; font-size:13px;")
        sums.addWidget(sums_lbl)
        l.addLayout(sums)

        if s.get("observations",""):
            l.addWidget(QLabel(f"<b>Observações:</b> {s['observations']}"))

        btn = QPushButton("Fechar"); btn.clicked.connect(dlg.accept)
        l.addWidget(btn)
        dlg.exec_()


# ═══════════════════════════════════════════════════════════════════
#  MAIN WINDOW
# ═══════════════════════════════════════════════════════════════════

class MainWindow(QMainWindow):
    def __init__(self, dm: DataManager):
        super().__init__()
        self.dm = dm
        settings = dm.get_settings()
        self._current_theme = settings.get("theme","dark")
        self._bg_image = settings.get("background_image","")

        self.setWindowTitle("🔧  Sistema de Vendas — Auto Peças & Mecânica")
        self.setMinimumSize(1200, 750)
        self.resize(1400, 860)

        self._build_ui()
        self._apply_theme()

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QHBoxLayout(central)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # SIDEBAR
        self.sidebar = QFrame()
        self.sidebar.setObjectName("sidebar")
        self.sidebar.setFixedWidth(220)
        sidebar_layout = QVBoxLayout(self.sidebar)
        sidebar_layout.setContentsMargins(12, 16, 12, 12)
        sidebar_layout.setSpacing(4)

        # Logo area
        logo_frame = QFrame()
        logo_layout = QVBoxLayout(logo_frame)
        logo_layout.setContentsMargins(8, 8, 8, 16)

        logo_icon = QLabel("🔧")
        logo_icon.setFont(QFont("Segoe UI Emoji", 28))
        logo_icon.setAlignment(Qt.AlignCenter)
        logo_layout.addWidget(logo_icon)

        logo_text = QLabel("AUTO PEÇAS")
        logo_text.setAlignment(Qt.AlignCenter)
        logo_text.setStyleSheet("font-size:14px; font-weight:bold; letter-spacing:2px;")
        logo_layout.addWidget(logo_text)

        logo_sub = QLabel("Sistema de Vendas v1.0")
        logo_sub.setObjectName("subtitle")
        logo_sub.setAlignment(Qt.AlignCenter)
        logo_sub.setStyleSheet("font-size:10px;")
        logo_layout.addWidget(logo_sub)

        sep = QFrame(); sep.setFrameShape(QFrame.HLine); sep.setObjectName("card")
        logo_layout.addWidget(sep)
        sidebar_layout.addWidget(logo_frame)

        # Nav buttons
        nav_items = [
            ("🏠", "Dashboard", 0),
            ("📦", "Produtos", 1),
            ("🛒", "Nova Venda", 2),
            ("📋", "Histórico de Vendas", 3),
            ("👥", "Clientes", 4),
            ("📊", "Relatórios", 5),
        ]

        self.nav_buttons = []
        for icon, label, idx in nav_items:
            btn = QPushButton(f"{icon}  {label}")
            btn.setObjectName("nav_btn")
            btn.setMinimumHeight(42)
            btn.clicked.connect(lambda _, i=idx: self._navigate(i))
            sidebar_layout.addWidget(btn)
            self.nav_buttons.append(btn)

        sidebar_layout.addStretch()

        sep2 = QFrame(); sep2.setFrameShape(QFrame.HLine); sep2.setObjectName("card")
        sidebar_layout.addWidget(sep2)

        # Options section
        opt_lbl = QLabel("OPÇÕES")
        opt_lbl.setObjectName("metric_label")
        opt_lbl.setContentsMargins(8, 4, 0, 4)
        sidebar_layout.addWidget(opt_lbl)

        btn_settings = QPushButton("⚙️  Configurações")
        btn_settings.setObjectName("nav_btn")
        btn_settings.setMinimumHeight(42)
        btn_settings.clicked.connect(self._open_settings)
        sidebar_layout.addWidget(btn_settings)

        btn_backup = QPushButton("💾  Backup de Dados")
        btn_backup.setObjectName("nav_btn")
        btn_backup.setMinimumHeight(42)
        btn_backup.clicked.connect(self._backup)
        sidebar_layout.addWidget(btn_backup)

        btn_logout = QPushButton("🚪  Sair do Sistema")
        btn_logout.setObjectName("nav_btn")
        btn_logout.setMinimumHeight(42)
        btn_logout.clicked.connect(self._logout)
        sidebar_layout.addWidget(btn_logout)

        main_layout.addWidget(self.sidebar)

        # CONTENT AREA
        self.content_stack = QStackedWidget()
        main_layout.addWidget(self.content_stack, 1)

        self.tab_dashboard = DashboardTab(self.dm)
        self.tab_products = ProductsTab(self.dm)
        self.tab_sales = SalesTab(self.dm)
        self.tab_sales.sale_completed.connect(self._on_sale_completed)
        self.tab_history = SalesHistoryTab(self.dm)
        self.tab_customers = CustomersTab(self.dm)
        self.tab_reports = ReportsTab(self.dm)

        self.content_stack.addWidget(self.tab_dashboard)
        self.content_stack.addWidget(self.tab_products)
        self.content_stack.addWidget(self.tab_sales)
        self.content_stack.addWidget(self.tab_history)
        self.content_stack.addWidget(self.tab_customers)
        self.content_stack.addWidget(self.tab_reports)

        # Status bar
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage(
            f"👤 Usuário: admin  |  📁 Dados: {DATA_FILE}  |  "
            f"🕐 {datetime.now().strftime('%d/%m/%Y %H:%M')}")

        self._navigate(0)

    def _navigate(self, idx):
        self.content_stack.setCurrentIndex(idx)
        for i, btn in enumerate(self.nav_buttons):
            btn.setObjectName("nav_btn_active" if i == idx else "nav_btn")
            btn.style().unpolish(btn)
            btn.style().polish(btn)
        # Refresh on navigate
        if idx == 0: self.tab_dashboard.refresh()
        elif idx == 1: self.tab_products.refresh()
        elif idx == 3: self.tab_history.refresh()
        elif idx == 4: self.tab_customers.refresh()

    def _on_sale_completed(self):
        self.tab_dashboard.refresh()
        self.tab_history.refresh()
        self.tab_products.refresh()
        self.tab_sales.refresh_customers()

    def _open_settings(self):
        dlg = SettingsDialog(self, self.dm)
        if dlg.exec_() == QDialog.Accepted:
            settings = self.dm.get_settings()
            self._current_theme = settings.get("theme","dark")
            self._bg_image = settings.get("background_image","")
            self._apply_theme()

    def _apply_theme(self):
        style = build_stylesheet(self._current_theme)
        QApplication.instance().setStyleSheet(style)
        if self._bg_image and os.path.exists(self._bg_image):
            pix = QPixmap(self._bg_image)
            palette = self.palette()
            scaled = pix.scaled(self.size(), Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)
            brush = QBrush(scaled)
            palette.setBrush(QPalette.Window, brush)
            self.setPalette(palette)
            self.setAutoFillBackground(True)
        else:
            self.setPalette(QApplication.instance().palette())
            self.setAutoFillBackground(False)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        if self._bg_image and os.path.exists(self._bg_image):
            self._apply_theme()

    def _backup(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Salvar Backup",
            f"backup_autopecas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            "JSON (*.json)")
        if path:
            import shutil
            try:
                shutil.copy2(DATA_FILE, path)
                QMessageBox.information(self, "✅  Backup Realizado",
                                        f"Backup salvo em:\n{path}")
            except Exception as e:
                QMessageBox.critical(self, "Erro no Backup", str(e))

    def _logout(self):
        resp = QMessageBox.question(self, "Confirmar Saída",
                                    "Deseja sair do sistema?",
                                    QMessageBox.Yes | QMessageBox.No)
        if resp == QMessageBox.Yes:
            self.close()
            from login import show_login
            show_login()


# ═══════════════════════════════════════════════════════════════════
#  LOGIN WINDOW
# ═══════════════════════════════════════════════════════════════════

class LoginWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.dm = DataManager()
        self.setWindowTitle("Login — Sistema Auto Peças")
        self.setFixedSize(420, 540)
        self.setWindowFlags(Qt.FramelessWindowHint)
        self._drag_pos = None
        self._build_ui()
        QApplication.instance().setStyleSheet(build_stylesheet("dark"))

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # Background gradient panel
        bg = QFrame()
        bg_layout = QVBoxLayout(bg)
        bg_layout.setContentsMargins(40, 32, 40, 32)
        bg_layout.setSpacing(0)
        bg.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
                    stop:0 #1A1A1A, stop:1 #2A1A0A);
                border-radius: 14px;
            }
        """)

        # Title bar (drag)
        titlebar = QHBoxLayout()
        title_lbl = QLabel("● ● ●")
        title_lbl.setStyleSheet("color:#555; font-size:11px;")
        titlebar.addWidget(title_lbl)
        titlebar.addStretch()
        close_btn = QPushButton("✕")
        close_btn.setFixedSize(28, 28)
        close_btn.setStyleSheet("background:transparent; color:#888; font-size:14px; border:none;")
        close_btn.clicked.connect(QApplication.quit)
        titlebar.addWidget(close_btn)
        bg_layout.addLayout(titlebar)
        bg_layout.addSpacing(10)

        # Logo
        logo = QLabel("🔧")
        logo.setFont(QFont("Segoe UI Emoji", 48))
        logo.setAlignment(Qt.AlignCenter)
        bg_layout.addWidget(logo)

        title = QLabel("AUTO PEÇAS")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size:22px; font-weight:bold; color:#FF6B35; letter-spacing:4px;")
        bg_layout.addWidget(title)

        subtitle = QLabel("Sistema de Gestão e Vendas")
        subtitle.setAlignment(Qt.AlignCenter)
        subtitle.setStyleSheet("font-size:12px; color:#888; margin-bottom:30px;")
        bg_layout.addWidget(subtitle)
        bg_layout.addSpacing(24)

        # Form
        form_frame = QFrame()
        form_frame.setStyleSheet("""
            QFrame {
                background: rgba(255,255,255,0.05);
                border: 1px solid rgba(255,255,255,0.1);
                border-radius: 10px;
                padding: 8px;
            }
        """)
        form_layout = QVBoxLayout(form_frame)
        form_layout.setContentsMargins(20, 20, 20, 20)
        form_layout.setSpacing(14)

        user_lbl = QLabel("👤  USUÁRIO")
        user_lbl.setStyleSheet("font-size:11px; color:#888; font-weight:bold; letter-spacing:1px;")
        form_layout.addWidget(user_lbl)
        self.user_e = QLineEdit()
        self.user_e.setPlaceholderText("Digite seu usuário...")
        self.user_e.setText("admin")
        self.user_e.setMinimumHeight(42)
        self.user_e.setStyleSheet("""
            QLineEdit {
                background: rgba(255,255,255,0.08);
                border: 1px solid rgba(255,255,255,0.15);
                border-radius: 8px;
                color: white;
                padding: 8px 14px;
                font-size: 14px;
            }
            QLineEdit:focus {
                border: 2px solid #FF6B35;
            }
        """)
        form_layout.addWidget(self.user_e)

        pwd_lbl = QLabel("🔒  SENHA")
        pwd_lbl.setStyleSheet("font-size:11px; color:#888; font-weight:bold; letter-spacing:1px;")
        form_layout.addWidget(pwd_lbl)
        self.pwd_e = QLineEdit()
        self.pwd_e.setEchoMode(QLineEdit.Password)
        self.pwd_e.setPlaceholderText("Digite sua senha...")
        self.pwd_e.setText("admin")
        self.pwd_e.setMinimumHeight(42)
        self.pwd_e.setStyleSheet("""
            QLineEdit {
                background: rgba(255,255,255,0.08);
                border: 1px solid rgba(255,255,255,0.15);
                border-radius: 8px;
                color: white;
                padding: 8px 14px;
                font-size: 14px;
            }
            QLineEdit:focus {
                border: 2px solid #FF6B35;
            }
        """)
        self.pwd_e.returnPressed.connect(self._login)
        form_layout.addWidget(self.pwd_e)

        bg_layout.addWidget(form_frame)
        bg_layout.addSpacing(18)

        self.err_lbl = QLabel("")
        self.err_lbl.setAlignment(Qt.AlignCenter)
        self.err_lbl.setStyleSheet("color:#FF1744; font-size:12px;")
        bg_layout.addWidget(self.err_lbl)

        btn_login = QPushButton("ENTRAR  →")
        btn_login.setMinimumHeight(48)
        btn_login.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 #FF6B35, stop:1 #FF8C5A);
                color: white;
                border: none;
                border-radius: 10px;
                font-size: 15px;
                font-weight: bold;
                letter-spacing: 2px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 #FF8C5A, stop:1 #FFAA80);
            }
            QPushButton:pressed {
                background: #CC4F1F;
            }
        """)
        btn_login.clicked.connect(self._login)
        bg_layout.addWidget(btn_login)
        bg_layout.addStretch()

        version = QLabel("v1.0.0  |  Desenvolvido para Auto Peças & Mecânica")
        version.setAlignment(Qt.AlignCenter)
        version.setStyleSheet("font-size:10px; color:#444; margin-top:8px;")
        bg_layout.addWidget(version)

        layout.addWidget(bg)

    def _login(self):
        user = self.user_e.text().strip()
        pwd = self.pwd_e.text().strip()
        if user == "admin" and pwd == "admin":
            self.err_lbl.setText("")
            self.main_win = MainWindow(self.dm)
            self.main_win.show()
            self.close()
        else:
            self.err_lbl.setText("❌  Usuário ou senha incorretos!")
            self.pwd_e.clear()
            self.pwd_e.setFocus()

    def mousePressEvent(self, e):
        if e.button() == Qt.LeftButton:
            self._drag_pos = e.globalPos() - self.pos()
            e.accept()

    def mouseMoveEvent(self, e):
        if e.buttons() == Qt.LeftButton and self._drag_pos:
            self.move(e.globalPos() - self._drag_pos)
            e.accept()


# ═══════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════

def main():
    app = QApplication(sys.argv)
    app.setApplicationName("AutoPeças Sistema")
    app.setApplicationVersion("1.0.0")

    # High DPI support
    try:
        app.setAttribute(Qt.AA_EnableHighDpiScaling, True)
        app.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    except Exception:
        pass

    win = LoginWindow()
    win.show()

    # Center on screen
    screen = app.primaryScreen().geometry()
    x = (screen.width() - win.width()) // 2
    y = (screen.height() - win.height()) // 2
    win.move(x, y)

    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
